import pandas as pd
import numpy as np
import os
import io # Para trabalhar com CSV/Excel em memória
import unicodedata
import xlwt # Para escrever arquivos .xls
import csv
import openpyxl # Necessário para engine='openpyxl' do pandas
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill # Para possíveis estilos futuros
from openpyxl.worksheet.table import Table, TableStyleInfo
import re # Para limpeza numérica e busca de colunas
import traceback # Para erros detalhados
from flask import (
    Flask, render_template, request, redirect, url_for,
    send_file, flash, session, abort
)
from collections import defaultdict
from openpyxl.utils.cell import range_boundaries
from werkzeug.utils import secure_filename
from formatadores.tabela_preco_formatador import find_column_flexible, normalize_text_for_match, extract_block_number_safe, parse_flexible_float
from formatadores.tabela_preco_formatador import processar_tabela_precos_web                            
from formatadores.tabela_preco_importador import (
    processar_preco_incorporacao,
    processar_preco_lote_avista,
    processar_preco_lote_parcelado
)
from formatadores.incorporacao_formatador import processar_incorporacao_web


# --- Constantes ---
TIPOLOGIAS_PADRAO = {
    "51 - 2 quartos sem suíte": "51", "36 - 2 quartos sendo 1 suíte térreo": "36",
    "34 - 3 quartos sendo 1 suíte térreo": "34",
    "21 - 2 quartos sendo 1 suíte casa": "21",
    "20 - 3 quartos sendo 1 suíte casa": "20",
    "88 PCD - 2 quartos sendo 1 suíte casa (PCD)": "88" 
}
TIPOLOGIAS_SUPERIOR = {
    "52 - 2 quartos sem suíte": "52", "35 - 2 quartos sendo 1 suíte superior": "35",
    "33 - 3 quartos sendo 1 suíte superior": "33" 
}
TIPOLOGIAS_PCD = {
    "50 PCD - 2 QUARTOS SENDO UMA SUÍTE - TÉRREO (PCD)": "50", 
    "88 PCD - 2 quartos sendo 1 suíte casa (PCD)": "88" 
}
ENDERECO_FIXO = {
    "Endereço": "Av. Olívia Flores", "Bairro": "Candeias", "Número": "1265", "Estado": "Bahia",
    "Cidade": "Vitória da Conquista", "CEP": "45028610", "Região": "Nordeste",
    "Data da Entrega (Empreendimento)": "01/01/2028"
}

# --- Configuração do Flask ---
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY', 'chave-padrao-muito-insegura-trocar-!!!!')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024 # 16MB
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
if not os.path.exists(app.config['UPLOAD_FOLDER']): os.makedirs(app.config['UPLOAD_FOLDER'])

# --- Funções Auxiliares Globais ---
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def format_decimal_br(value, precision):
    if pd.isna(value):
        return "" # Ou talvez '--' ou 0.0? Depende do que prefere para vazios
    try:
        num = float(value) # Tenta converter direto
        format_string = f"{{:.{precision}f}}"
        return format_string.format(num).replace('.', ',')
    except (ValueError, TypeError):
        # Se falhar, tenta limpar como número BR e converter
        s_val = str(value).strip()
        s_val = re.sub(r'[^\d,.-]', '', s_val)
        if ',' in s_val and '.' in s_val: s_val = s_val.replace('.', '').replace(',', '.')
        elif ',' in s_val: s_val = s_val.replace(',', '.')
        try:
            num = float(s_val)
            format_string = f"{{:.{precision}f}}"
            return format_string.format(num).replace('.', ',')
        except (ValueError, TypeError):
            print(f"Aviso format_decimal_br: Não formatou '{value}' com precisão {precision}.")
            return str(value) # Retorna original se tudo falhar

def format_decimal_br_cv(value, precision=2):
    """
    Tenta converter valor para float e formata como string com vírgula decimal.
    Tenta substituir ponto por vírgula mesmo se a conversão falhar,
    se o valor original parecer um número com ponto.
    Retorna string vazia para nulos/vazios, ou original em último caso.
    """
    if pd.isna(value) or str(value).strip() == '':
        return ""

    original_str = str(value).strip()
    numeric_value = parse_flexible_float(original_str) # Tenta converter

    if numeric_value is not None:
        # CONVERSÃO OK: Formata o número
        try:
            format_string = f"{{:.{precision}f}}"
            # Formata para garantir a precisão correta
            formatted_num_str = format_string.format(numeric_value)
            # Substitui o ponto pela vírgula no resultado formatado
            return formatted_num_str.replace('.', ',')
        except (ValueError, TypeError):
            # Erro inesperado na formatação do número já convertido
            print(f"Aviso format_decimal_br_cv: Falha INESPERADA ao formatar número '{numeric_value}'")
            # Como último recurso, tenta a substituição no original se tiver ponto
            if '.' in original_str and ',' not in original_str:
                # Verifica se parece um número antes de substituir cegamente
                if re.fullmatch(r'-?\s*\d+(\.\d+)?\s*$', original_str):
                     print(f"  Fallback format: Substituindo ponto no original '{original_str}' após erro de formatação.")
                     return original_str.replace('.', ',')
            return original_str # Retorna original se tudo falhar
    else:
        # CONVERSÃO FALHOU: Verifica o formato da string original
        # Já tem vírgula? Provavelmente já está correto (ou é texto inválido)
        if ',' in original_str:
             # Pode ser que já esteja formatado ou seja "1,2,3" - retorna como está
             return original_str
        # Não tem vírgula, mas tem ponto? Tenta substituir.
        elif '.' in original_str:
             # Verifica se parece um número ANTES de substituir
             # Permite espaços no início/fim, mas o miolo deve ser numérico com ponto
             if re.fullmatch(r'-?\s*\d+(\.\d+)?\s*$', original_str):
                 print(f"Aviso format_decimal_br_cv: Parse falhou para '{original_str}', mas parece numérico com ponto. Substituindo ponto.")
                 # Tenta formatar para garantir a precisão, se possível (menos provável de funcionar)
                 try:
                     num_from_dot = float(original_str)
                     format_string = f"{{:.{precision}f}}"
                     formatted_num_str = format_string.format(num_from_dot)
                     return formatted_num_str.replace('.', ',')
                 except (ValueError, TypeError):
                     # Se formatar falhar, apenas substitui o ponto
                     return original_str.replace('.', ',')
             else:
                 # Tem ponto, mas não parece um número (ex: "Texto.com")
                 return original_str
        else:
             # Não tem vírgula nem ponto, conversão falhou (provavelmente inteiro ou texto)
             # Verifica se é um inteiro para adicionar casas decimais ",00"
             if original_str.isdigit() or (original_str.startswith('-') and original_str[1:].isdigit()):
                 try:
                     num_int = int(original_str)
                     format_string = f"{{:.{precision}f}}"
                     formatted_num_str = format_string.format(num_int)
                     return formatted_num_str.replace('.', ',')
                 except ValueError:
                     return original_str # Retorna se não for inteiro válido
             else:
                # É apenas texto sem separadores numéricos
                return original_str # Retorna original

# --- Funções Auxiliares CV ---
def normalize_text(text):
    if not isinstance(text, str): text = str(text)
    text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
    return text.upper().strip()
def normalize_column_name(column_name):
    if not isinstance(column_name, str): column_name = str(column_name)
    normalized = unicodedata.normalize('NFKD', column_name).encode('ASCII', 'ignore').decode('ASCII')
    return normalized.lower().replace(" ", "").strip()
def encontrar_coluna_garagem(df_columns):
    normalized_columns = {normalize_column_name(col): col for col in df_columns}
    for norm_name, orig_name in normalized_columns.items():
        if "garagem" in norm_name: return orig_name
    return None
def formatar_nome_unidade(row, bq_col_name, ca_col_name):
    """
    Formata o nome da unidade combinando Bloco/Quadra e Casa/Apto.
    Ex: BL01 - CASA 05, QD10 - APT 101
    """
    # Pega os valores das colunas corretas
    bq_val = row.get(bq_col_name, '')
    ca_val = row.get(ca_col_name, '')

    # Formata número do Bloco/Quadra com zero à esquerda
    bq_num_str = "??"
    bq_prefix = "BL" # Default prefix
    if pd.notna(bq_val) and str(bq_val).strip():
        s_bq = str(bq_val).strip()
        match = re.search(r'\d+', s_bq)
        if match:
            try:
                bq_num_str = f"{int(match.group(0)):02d}"
            except ValueError:
                bq_num_str = "??" # Mantém ?? se não for número
        else:
             bq_num_str = "??" # Mantém ?? se não achar número
        # Determina prefixo baseado no NOME da coluna
        if 'quadra' in str(bq_col_name).lower():
            bq_prefix = "QD"
        else:
            bq_prefix = "BL" # Default para bloco ou outros nomes
    else:
        # Caso não encontre valor na coluna Bloco/Quadra
        print(f"Aviso: Valor não encontrado para Bloco/Quadra na coluna '{bq_col_name}' na linha: {row.name if hasattr(row, 'name') else 'desconhecida'}")
        # Poderia retornar um erro ou um nome padrão
        # return "ERRO_BLOCO_QUADRA_AUSENTE"

    # Formata número da Casa/Apto com zero à esquerda
    ca_num_str = "??"
    ca_prefix = "UNID" # Default prefix
    if pd.notna(ca_val) and str(ca_val).strip():
        s_ca = str(ca_val).strip()
        # Extrai todos os dígitos
        digits = ''.join(filter(str.isdigit, s_ca))
        if digits:
            try:
                ca_num_str = f"{int(digits):02d}" # Formata com zero à esquerda
            except ValueError:
                ca_num_str = s_ca # Usa a string original se a conversão falhar (improvável)
        else:
            ca_num_str = s_ca # Usa a string original se não houver dígitos
        # Determina prefixo baseado no NOME da coluna
        if 'casa' in str(ca_col_name).lower():
            ca_prefix = "CASA"
        elif 'apt' in str(ca_col_name).lower() or 'apartamento' in str(ca_col_name).lower():
            ca_prefix = "APT"
        else:
            ca_prefix = "UNID" # Default para 'unidade' ou outros nomes
    else:
        # Caso não encontre valor na coluna Casa/Apto
         print(f"Aviso: Valor não encontrado para Casa/Apto na coluna '{ca_col_name}' na linha: {row.name if hasattr(row, 'name') else 'desconhecida'}")
        # return "ERRO_CASA_APTO_AUSENTE"


    # Lógica PCD (mantida)
    # Verifica TIPO e a coluna original de Casa/Apto para PCD
    tipo_val_pcd = row.get('TIPO', '') # Assume que TIPO foi encontrada e renomeada
    pcd = " (PCD)" if any('PCD' in normalize_text(str(v)) for v in [ca_val, tipo_val_pcd]) else ""

    # Combina tudo - Só retorna se ambos os números foram minimamente definidos
    if bq_num_str != "??" and ca_num_str != "??":
        return f"{bq_prefix}{bq_num_str} - {ca_prefix} {ca_num_str}{pcd}"
    elif bq_num_str != "??": # Retorna só o bloco/quadra se a unidade falhou
         print(f"Aviso: Nome da unidade incompleto (faltou Casa/Apto?) para Bloco/Quadra {bq_prefix}{bq_num_str} na linha {row.name if hasattr(row, 'name') else 'desconhecida'}")
         return f"{bq_prefix}{bq_num_str}{pcd}" # Adiciona PCD mesmo se incompleto
    elif ca_num_str != "??": # Retorna só a unidade se o bloco falhou
        print(f"Aviso: Nome da unidade incompleto (faltou Bloco/Quadra?) para Unidade {ca_prefix} {ca_num_str} na linha {row.name if hasattr(row, 'name') else 'desconhecida'}")
        return f"{ca_prefix} {ca_num_str}{pcd}"
    else:
        print(f"Erro: Não foi possível gerar nome da unidade para linha {row.name if hasattr(row, 'name') else 'desconhecida'}")
        return "NOME_UNIDADE_INVALIDO" # Ou retorna ""
def verificar_vaga(g, num_mode):
    if pd.isna(g): return "01 VAGA"
    if num_mode:
        s = str(g).strip(); n=0
        if s:
            for sep in [" e ",","]:
                if sep in s: n=len([v for v in s.split(sep) if v.strip()]); break
            else: n=1
        if n>=4: return "04 VAGAS";
        elif n==3: return "03 VAGAS"
        elif n==2: return "02 VAGAS";
        else: return "01 VAGA"
    else:
        try:
            gn=float(str(g).replace(',','.').strip())
            if abs(gn-int(gn))>0.001: # Metragem
                if gn>35: return"04 VAGAS";
                elif gn>25: return"03 VAGAS"
                elif gn>15: return"02 VAGAS";
                else: return"01 VAGA"
            else: # Int
                gi=int(gn)
                if gi>=4: return"04 VAGAS";
                elif gi==3: return"03 VAGAS"
                elif gi==2: return"02 VAGAS";
                else: return"01 VAGA"
        except: return verificar_vaga(g, True)
def formatar_jardim(v):
    if pd.isna(v): return ""
    try: vf=float(str(v).replace(',','.')); return f"{vf:.2f}".replace('.',',')+" m²" if vf!=0 else ""
    except: return ""
def mapear_tipologia_web(row, tip_map, is_casa):
    t_orig=str(row.get('TIPO','')).strip(); map_i=tip_map.get(t_orig,{}); unit=str(row.get('APT','')or row.get('CASA','')).strip().upper()
    if not t_orig or not map_i: return None
    is_pcd='PCD' in unit or 'PCD' in normalize_text(t_orig)
    if is_pcd: return map_i.get('pcd')
    if is_casa: return map_i.get('padrao')
    try: apt_n=int(''.join(filter(str.isdigit, unit))); return map_i.get('padrao') if apt_n<=6 else map_i.get('superior')
    except: return map_i.get('padrao')

# --- Funções Auxiliares CV Lote ---
def normalize_text_lote(t):
    if pd.isna(t): return ""
    t = str(t); t = unicodedata.normalize('NFKD', t).encode('ASCII','ignore').decode('ASCII')
    return t.upper().strip()
def normalize_column_name_lote(c):
    if pd.isna(c): return ""
    norm = unicodedata.normalize('NFKD', str(c)).encode('ASCII','ignore').decode('ASCII')
    return norm.lower().replace(" ","").strip()
def encontrar_coluna_similar_lote(cols, target):
    t=target.lower().strip(); norm_c={normalize_column_name_lote(col):col for col in cols}
    for n, o in norm_c.items():
        if t in n: print(f"(Lote) Found '{t}': '{o}'"); return o
    print(f"(Lote) Warn: Col '{t}' not found."); return None
def limpar_converter_numerico_lote(v):
    if pd.isna(v): return 0.0
    try:
        s=str(v); s=re.sub(r'M2|M²','',s,flags=re.IGNORECASE).strip(); sep='.'
        if s.rfind(',')>s.rfind('.'): sep=','
        elif s.rfind('.')==-1 and s.rfind(',')!=-1: sep=','
        s = s.replace('.','') if sep==',' else s.replace(',','')
        s=s.replace(sep,'.'); s=''.join(s.split())
        return float(s) if re.fullmatch(r'-?\d+(\.\d+)?',s) else 0.0
    except: return 0.0
def formatar_nome_bloco_lote(row, col_q):
    try:
        if col_q in row.index and pd.notna(row[col_q]):
            try: n=int(row[col_q])
            except ValueError: n=int(float(row[col_q]))
            return f"QUADRA {n:02d}"
        else: return "QUADRA_NA"
    except: return "QUADRA_ERR"
def formatar_nome_unidade_lote(row, col_q, col_l):
    qd_s, lt_s = "QD??", "LOTE ??"
    try:
        if col_q in row.index and pd.notna(row[col_q]):
            try: qd_s=f"QD{int(float(row[col_q])):02d}"
            except: qd_s="QD_INV"
        else: qd_s="QD_NA"
        if col_l in row.index and pd.notna(row[col_l]):
            lv=str(row[col_l]).strip(); ln=''.join(filter(str.isdigit, lv))
            if ln:
                try: lt_s=f"LOTE {int(ln):02d}"
                except: lt_s="LOTE_INV"
            else: lt_s=f"LOTE_{lv}" if lv else "LOTE_S/N"
        else: lt_s="LOTE_NA"
        return f"{qd_s} - {lt_s}"
    except: return "ERRO_NOME_UNIDADE"
def formatar_fracao_ideal_lote(v_num):
    try: return "" if pd.isna(v_num) else str(float(v_num)).replace('.',',')
    except: return "ERRO_FRAC"
def formatar_area_privativa_lote(v_num):
    try: return "" if pd.isna(v_num) else f"{float(v_num):.2f}".replace('.',',')+" m²"
    except: return "ERRO m²"

# --- Funções Auxiliares SIENGE ---
def normalize_column_name_sienge(c):
    if pd.isna(c): return ""
    return str(c).upper().strip()
def determinar_tipo_imovel_sienge(row, apt_col):
    if apt_col=="APT": return "APARTAMENTO";
    elif apt_col=="CASA": return "CASA";
    else: return "INDEFINIDO"
def formatar_unidade_sienge(row, bloco_coluna_nome, apt_coluna_nome):
    """
    Formata a unidade para o padrão SIENGE (ex: QD01 - CASA 01, BL01 - APT 01).
    Determina o prefixo (QD/BL) baseado no nome da coluna bloco_coluna_nome.
    """
    bloco_str = "00"
    apt_str = "00"
    bloco_prefix = "??" # Default/fallback
    apt_prefix = "??"   # Default/fallback

    try:
        # --- Processa Bloco/Quadra ---
        if bloco_coluna_nome and pd.notna(row.get(bloco_coluna_nome)):
            bloco_val = row[bloco_coluna_nome]

            # --- LÓGICA CORRIGIDA PARA O PREFIXO ---
            # Verifica se o NOME da coluna contém 'QUADRA' (ignorando case)
            if 'QUADRA' in str(bloco_coluna_nome).upper():
                bloco_prefix = "QD"
            else: # Assume que é Bloco ou usa BL como padrão
                bloco_prefix = "BL"
            # --- FIM DA LÓGICA CORRIGIDA ---

            # Formata a parte numérica do VALOR do Bloco/Quadra
            try:
                # Tenta converter via float primeiro para lidar com "1.0" etc.
                bloco_int = int(float(bloco_val))
                bloco_str = f"{bloco_int:02d}" # Formata com zero à esquerda
            except (ValueError, TypeError):
                # Se a conversão falhar, usa o valor original (removendo espaços)
                bloco_str = str(bloco_val).strip()
                print(f"Aviso SIENGE L{row.name if hasattr(row,'name') else 'Unk'}: Não converteu valor Bloco/Quadra '{bloco_val}' para número. Usando original: '{bloco_str}'")

        # --- Processa Apartamento/Casa (Mantém lógica anterior) ---
        if apt_coluna_nome and pd.notna(row.get(apt_coluna_nome)):
            apt_val = row[apt_coluna_nome]
            # Deriva o prefixo do NOME da coluna de apt/casa
            apt_prefix = str(apt_coluna_nome).upper() # Ex: "CASA", "APT"

            # Extrai e formata a parte numérica do VALOR do apt/casa
            apt_num_str = ''.join(filter(str.isdigit, str(apt_val)))
            if apt_num_str:
                try:
                    apt_int = int(apt_num_str)
                    apt_str = f"{apt_int:02d}" # Formata com zero à esquerda
                except ValueError:
                    # Caso raro após filtrar dígitos
                    apt_str = apt_num_str
            else:
                # Se não encontrar dígitos (ex: valor é "GARAGEM")
                apt_str = str(apt_val).strip() # Usa o valor original
                # Considerar se "S/N" seria melhor aqui ou o valor original
                # apt_str = "S/N"
                print(f"Aviso SIENGE L{row.name if hasattr(row,'name') else 'Unk'}: Não extraiu número de '{apt_val}' em {apt_coluna_nome}. Usando original: '{apt_str}'")

        # --- Combina as partes (Mantém lógica anterior) ---
        # Verifica se ambos os prefixos foram definidos (não são mais "??")
        if bloco_coluna_nome and apt_coluna_nome and bloco_prefix != "??" and apt_prefix != "??":
            return f"{bloco_prefix}{bloco_str} - {apt_prefix} {apt_str}"
        elif bloco_coluna_nome and bloco_prefix != "??":
            return f"{bloco_prefix}{bloco_str}" # Retorna só Bloco/Quadra se apt falhar
        elif apt_coluna_nome and apt_prefix != "??":
            return f"{apt_prefix} {apt_str}" # Retorna só Apt/Casa se bloco falhar
        else:
            return "N/D" # Caso nenhum seja encontrado/processado

    except Exception as e:
        print(f"(SIENGE) Erro formatar unidade: {e}")
        import traceback
        traceback.print_exc() # Ajuda a depurar erros inesperados
        return "ERRO_FORMAT"

# --- Funções Auxiliares SIENGE Lote ---
def normalize_column_name_sienge_lote(c):
    if pd.isna(c): return ""
    n=str(c).upper().strip(); n=n.replace('(M²)','(M2)').replace(' (M2)','(M2)'); n=n.replace('AREA(M2)','ÁREA(M2)')
    return n
def extrair_numero_sienge_lote(t, pref=None):
    if pd.isna(t): return None
    s=str(t).strip();
    if pref: s=re.sub(f'^{re.escape(pref)}\s*','',s,flags=re.IGNORECASE)
    m=re.search(r'\d+',s)
    try: return int(m.group(0)) if m else None
    except: return None
def formatar_unidade_sienge_lote(row, col_q, col_l):
    try:
        qv=row.get(col_q); qn=extrair_numero_sienge_lote(qv); pb=f"QD{qn:02d}" if qn is not None else ""
        lv=row.get(col_l); ln=extrair_numero_sienge_lote(lv,"LT") or extrair_numero_sienge_lote(lv,"LOTE"); pa=f"LOTE {ln:02d}" if ln is not None else ""
        if pb and pa: return f"{pb} - {pa}";
        elif pb: return pb;
        elif pa: return pa;
        else: return "LOCALIZACAO_INVALIDA"
    except: return "ERRO_FORMATACAO"
def limpar_converter_numerico_sienge_lote(v): return limpar_converter_numerico_lote(v) # Reutiliza

# --- Funções Auxiliares Formatador Incorporação ---
def processar_formatador_incorporacao_avancado(input_filepath):
    """
    Processa a planilha de incorporação (versão reestruturada):
    - Lê o cabeçalho da linha 3.
    - Usa a coluna 'QUADRA' dos dados para criar 'QUADRA_NUM'.
    - Remove colunas sem nome no cabeçalho original.
    - Aplica formatação numérica específica (casas decimais).
    """
    print(f"(Incorp Reestruturado) Processando: {input_filepath}")
    output = io.BytesIO()

    try:
        # 1. Leitura Direta com Cabeçalho na Linha 3 (índice 2)
        try:
            df_input = pd.read_excel(input_filepath, header=2, dtype=str) # Lê tudo como string inicialmente
            # Remove linhas que são completamente NA (podem aparecer no final)
            df_input.dropna(how='all', inplace=True)
            # Limpa nomes das colunas lidas
            df_input.columns = [str(col).strip() for col in df_input.columns]
            print(f"(Incorp Reestruturado) Lidas {len(df_input)} linhas de dados. Colunas lidas: {df_input.columns.tolist()}")
        except Exception as e:
            print(f"(Incorp Reestruturado) ERRO ao ler Excel (header=2): {e}")
            # Tenta ler sem header específico como fallback, pode não ser ideal
            try:
                print("(Incorp Reestruturado) Tentando ler sem header específico...")
                df_input = pd.read_excel(input_filepath, dtype=str)
                df_input.dropna(how='all', inplace=True)
                # Tenta encontrar header manualmente (complexo, melhor avisar usuário)
                raise ValueError(f"Falha ao ler o arquivo Excel com header na linha 3. Verifique o arquivo. Erro: {e}")
            except Exception as e2:
                 raise ValueError(f"Falha grave ao ler o arquivo Excel. Verifique o formato. Erros: {e}, {e2}")


        # 2. Filtrar Colunas "Unnamed" (geradas por cabeçalhos vazios)
        original_columns = df_input.columns.tolist()
        cols_to_keep = [col for col in original_columns if not str(col).startswith('Unnamed:')]
        if len(cols_to_keep) < len(original_columns):
             removed_cols = [col for col in original_columns if col not in cols_to_keep]
             print(f"(Incorp Reestruturado) Removendo colunas sem nome no cabeçalho: {removed_cols}")
        df_proc = df_input[cols_to_keep].copy() # Cria cópia com colunas válidas

        # 3. Criar Coluna QUADRA_NUM a partir da coluna original 'QUADRA'
        quadra_col_orig_name = find_column_flexible(df_proc.columns, ['quadra', 'bloco'], 'QUADRA', required=True)
        quadra_col_name = "QUADRA" # Nome da nova coluna a ser criada

        def format_quadra_num(q_val):
            if pd.isna(q_val) or str(q_val).strip() == '': return ''
            try:
                # Converte para float primeiro (caso seja "1.0") depois para int
                return f"{int(float(str(q_val))):02d}"
            except (ValueError, TypeError):
                return str(q_val).strip() # Retorna original se não converter

        # Aplica a formatação para criar a nova coluna
        df_proc[quadra_col_name] = df_proc[quadra_col_orig_name].apply(format_quadra_num)
        print(f"(Incorp Reestruturado) Coluna '{quadra_col_name}' criada a partir de '{quadra_col_orig_name}'.")

        # 4. Reordenar Colunas (QUADRA_NUM primeiro)
        # Garante que a coluna original 'QUADRA' não seja duplicada se tiver o mesmo nome normalizado
        final_columns_order = [quadra_col_name]
        for col in df_proc.columns:
             # Adiciona se não for a nova coluna de quadra e nem a original
             if col != quadra_col_name and col != quadra_col_orig_name:
                 final_columns_order.append(col)
        # OU se quiser MANTER a coluna QUADRA original também:
        # final_columns_order = [quadra_col_name] + [col for col in df_proc.columns if col != quadra_col_name]

        df_proc = df_proc[final_columns_order]
        print(f"(Incorp Reestruturado) Ordem final das colunas: {df_proc.columns.tolist()}")

        # 5. Escrever no Excel e Aplicar Formatação Numérica
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_proc.to_excel(writer, sheet_name='Incorporacao Formatada', index=False, header=True)

            workbook_out = writer.book
            worksheet_out = writer.sheets['Incorporacao Formatada']
            print("(Incorp Reestruturado) Aplicando formatação numérica no Excel...")

            # Mapeamento: Nome da coluna NORMALIZADO -> formato Excel
            col_formats = {
                'areaconstruida': '0.00',
                'quintal': '0.00',
                'areadescobertafrontal': '0.00', # Precisa ter essa coluna no header da linha 3
                'areaprivativa': '0.00', # Formato visual 'XXX,XX' não existe, 0.00 dá 2 casas
                'fracaoideal': '0.00000' # 5 casas decimais
            }
            # Mapeamento de nome normalizado para nome real no DataFrame FINAL
            df_cols_normalized = {normalize_text_for_match(col): col for col in df_proc.columns}

            # Encontra os índices (base 1) das colunas a serem formatadas no df_proc
            col_indices_to_format = {} # col_index (1-based) -> format_string
            found_formats_applied = []
            for norm_name, fmt_str in col_formats.items():
                real_col_name = df_cols_normalized.get(norm_name)
                if real_col_name:
                    try:
                        col_index_0based = df_proc.columns.get_loc(real_col_name)
                        col_indices_to_format[col_index_0based + 1] = fmt_str
                        found_formats_applied.append(real_col_name)
                    except KeyError:
                         print(f"  AVISO INTERNO: Coluna '{real_col_name}' (de '{norm_name}') não encontrada no índice do df_proc.")
                else:
                    # Só avisa se a coluna não foi encontrada entre as colunas válidas
                     if norm_name not in ['areadescobertafrontal']: # Exemplo: não avisa se esta for opcional
                         print(f"  AVISO: Coluna para formato '{norm_name}' não encontrada no cabeçalho (linha 3) ou foi filtrada.")

            print(f"(Incorp Reestruturado) Formatos serão aplicados para: {found_formats_applied}")
            if not col_indices_to_format:
                 print("(Incorp Reestruturado) AVISO: Nenhuma coluna encontrada para aplicar formatação numérica.")


            # Aplica o formato às células numéricas nas colunas corretas
            # Itera pelas linhas de dados (começando da linha 2 do Excel)
            for row_idx in range(2, worksheet_out.max_row + 1):
                # Itera pelas colunas que precisam de formatação
                for col_idx_1based, fmt_str in col_indices_to_format.items():
                    cell = worksheet_out.cell(row=row_idx, column=col_idx_1based)
                    # Tenta converter valor para float ANTES de aplicar formato
                    try:
                        numeric_value = parse_flexible_float(cell.value) # Usa a função importada
                        if numeric_value is not None:
                           cell.value = numeric_value # Garante que o valor na célula é numérico
                           cell.number_format = fmt_str
                        elif str(cell.value).strip() == '': # Se for vazio após parse, deixa vazio
                             cell.value = None
                        # Se não for conversível, mantém o valor original (string) sem formato
                    except Exception as e_fmt:
                        # Debug: Ajuda a entender por que falhou
                        # print(f"  Debug format error R{row_idx}C{col_idx_1based} Val:'{cell.value}' Err:{e_fmt}")
                        pass # Mantém valor original se parse_flexible_float falhar

            # Ajusta largura das colunas
            print("(Incorp Reestruturado) Ajustando largura das colunas...")
            for i, column_name in enumerate(df_proc.columns):
                column_letter = get_column_letter(i + 1)
                try:
                    # Lógica simples de largura baseada no conteúdo + cabeçalho
                    max_len_data = 0
                    if column_name in df_proc and not df_proc[column_name].empty:
                         # Calcula o máximo comprimento dos dados como string
                         max_len_data = df_proc[column_name].astype(str).map(len).max()

                    max_len_header = len(str(column_name))
                    # Pega o maior entre o dado mais longo e o cabeçalho, adiciona margem
                    width = max(max_len_data, max_len_header) + 3
                    # Limita a largura máxima
                    worksheet_out.column_dimensions[column_letter].width = min(width, 60)
                except Exception as e_width:
                     print(f"  Aviso: Falha ao ajustar largura da coluna {column_letter} ('{column_name}'): {e_width}")
                     worksheet_out.column_dimensions[column_letter].width = 15 # Fallback

        output.seek(0)
        print("(Incorp Reestruturado) Processamento concluído.")
        return output

    except ValueError as ve: # Erros de validação (leitura, coluna não encontrada)
         print(f"(Incorp Reestruturado) ERRO VALIDAÇÃO: {ve}")
         # Garante que o stream seja fechado se criado antes do erro
         if output: output.close()
         raise ve # Re-lança para o Flask mostrar
    except Exception as e:
        print(f"(Incorp Reestruturado) ERRO INESPERADO: {e}")
        traceback.print_exc()
        if output: output.close()
        raise RuntimeError(f"Erro inesperado no processamento do Formatador Incorporação: {e}") from e

# --- Funções Auxiliares Formatador Lote ---
def add_lt_prefix_if_needed_fmt_lote(v_str):
    if not isinstance(v_str,str): v_str=str(v_str)
    cleaned_value = v_str.strip() # Corrigido para usar cleaned_value
    if not cleaned_value: return ""
    if cleaned_value.isdigit(): return f"LT {cleaned_value}"
    elif cleaned_value.lower().startswith("lt"):
        if len(cleaned_value)>2 and cleaned_value[2].isspace(): return cleaned_value
        elif len(cleaned_value)>2 and cleaned_value[2].isdigit(): return f"{cleaned_value[:2]} {cleaned_value[2:].strip()}"
        else: return cleaned_value
    else: return cleaned_value
def format_measurement_fmt_lote(value_str, unit="m"):
    if not isinstance(value_str, str): value_str = str(value_str)
    cleaned_orig = value_str.strip() # Corrigido para usar cleaned_orig
    cleaned_float_attempt = cleaned_orig.lower().replace("m²", "").replace("m2", "").replace("m", "").strip()
    if not cleaned_float_attempt: return "N/A"
    try:
        numeric = float(cleaned_float_attempt.replace(',', '.'))
        formatted = f"{numeric:.2f}".replace('.', ',')
        return f"{formatted}{unit}"
    except (ValueError, TypeError):
        print(f"Aviso Fmt Lote: Valor '{cleaned_orig}' não convertido. Retornando original.") # Corrigido para cleaned_orig
        if cleaned_orig.lower().endswith(unit.lower()): return cleaned_orig
        else: return f"{cleaned_orig}{unit}" # Adiciona unidade se não tinha
def get_numeric_area_fmt_lote(a_str):
    if not isinstance(a_str,str): a_str=str(a_str)
    clean=a_str.strip().lower().replace("m²","").replace("m2","").replace("m","").strip()
    if not clean: return 0.0
    try: return float(clean.replace(',','.'))
    except: print(f"Warn Area (Fmt Lote): '{a_str}' -> 0.0"); return 0.0
def processar_formatador_lote_web(input_filepath):
    print(f"(Fmt Lote) Processando: {input_filepath}")
    try:
        df_raw=pd.read_excel(input_filepath, header=None, engine='openpyxl', dtype=str); df_raw.fillna("", inplace=True)
        dados_proc=[]; q_atual=None; q_val_num=None; cabecalho=None; map_hdr_rev={}; offset=0
        cols_esp={'lote':'LOTE', 'tipo':'TIPO', 'área(m²)':'AREA_M2','testada(m)':'TESTADA_M', 'fundo(m)':'FUNDO_M','lat. direita(n':'LAT_DIREITA_M', 'lat. direita(m)':'LAT_DIREITA_M','lat. esquerda(m)':'LAT_ESQUERDA_M','frente':'FRENTE_DESC', 'fundo':'FUNDO_DESC_CONFRONTANTE','direita':'DIREITA_DESC', 'esquerda':'ESQUERDA_DESC'}
        cols_medida={'AREA_M2':'m²', 'TESTADA_M':'m', 'FUNDO_M':'m','LAT_DIREITA_M':'m', 'LAT_ESQUERDA_M':'m'}
        cols_lt=['LOTE', 'FUNDO_DESC_CONFRONTANTE', 'DIREITA_DESC', 'ESQUERDA_DESC']
        print("(Fmt Lote) Varrendo...")
        for idx, row_s in df_raw.iterrows():
            linha=[str(v) for v in row_s.values]; linha_orig=idx+1; cel1=linha[0].strip()
            if cel1.lower().startswith(("quadra","bloco")):
                q_atual=cel1; m=re.search(r'\d+',q_atual)
                try: q_val_num=int(m.group(0)) if m else q_atual
                except: q_val_num=q_atual
                print(f" L{linha_orig}: QUADRA/BLOCO '{q_atual}' (Val:{q_val_num})"); cabecalho=None; map_hdr_rev={}; offset=linha_orig; continue
            if q_atual and not cabecalho:
                if cel1.lower()=='lote':
                    cabecalho=[h.strip() for h in linha]; print(f" L{linha_orig}: HEADER {cabecalho}")
                    map_hdr_rev={}; fundo_map={}
                    for i,hdr in enumerate(cabecalho):
                        hl=hdr.lower()
                        for el,ek in cols_esp.items():
                            if hl==el or (el=='lat. direita(n' and hl=='lat. direita(m)') or (el=='lat. direita(m)' and hl=='lat. direita(n)'):
                                if hl=='fundo':
                                    if ek=='FUNDO_M': fundo_map['FUNDO_M']=hdr
                                    elif ek=='FUNDO_DESC_CONFRONTANTE': fundo_map['FUNDO_DESC_CONFRONTANTE']=hdr
                                else: map_hdr_rev[ek]=hdr; break
                    if fundo_map.get('FUNDO_M')==fundo_map.get('FUNDO_DESC_CONFRONTANTE'):
                        if 'FUNDO_DESC_CONFRONTANTE' in fundo_map: map_hdr_rev['FUNDO_DESC_CONFRONTANTE']=fundo_map['FUNDO_DESC_CONFRONTANTE']
                    else:
                        if 'FUNDO_M' in fundo_map: map_hdr_rev['FUNDO_M']=fundo_map['FUNDO_M']
                        if 'FUNDO_DESC_CONFRONTANTE' in fundo_map: map_hdr_rev['FUNDO_DESC_CONFRONTANTE']=fundo_map['FUNDO_DESC_CONFRONTANTE']
                    print(f" Mapa Hdr Rev: {map_hdr_rev}"); offset=linha_orig; continue
                elif not any(c.strip() for c in linha): continue
                else: print(f" Warn L{linha_orig}: Ignorando linha antes do header '{cel1}'"); continue
            if q_atual and cabecalho:
                if any(c.strip() for c in linha) and cel1.lower()!='lote':
                    if len(linha)<len(cabecalho): linha.extend([""]*(len(cabecalho)-len(linha)))
                    if len(linha)>len(cabecalho): linha=linha[:len(cabecalho)]
                    d_lin=dict(zip(cabecalho,linha)); d_lin['QUADRA']=q_val_num
                    a_col=map_hdr_rev.get('AREA_M2'); a_str=d_lin.get(a_col,"") if a_col else ""; d_lin['_area_numerica']=get_numeric_area_fmt_lote(a_str)
                    for ek,u in cols_medida.items():
                        cr=map_hdr_rev.get(ek);
                        if cr and cr in d_lin: d_lin[cr]=format_measurement_fmt_lote(d_lin[cr],u)
                    for ek in cols_lt:
                        cr=map_hdr_rev.get(ek);
                        if cr and cr in d_lin: d_lin[cr]=add_lt_prefix_if_needed_fmt_lote(d_lin[cr])
                    try:
                        tm,fd = map_hdr_rev.get('TESTADA_M'),map_hdr_rev.get('FRENTE_DESC'); fum,fudc=map_hdr_rev.get('FUNDO_M'),map_hdr_rev.get('FUNDO_DESC_CONFRONTANTE'); ldm,dd=map_hdr_rev.get('LAT_DIREITA_M'),map_hdr_rev.get('DIREITA_DESC'); lem,ed=map_hdr_rev.get('LAT_ESQUERDA_M'),map_hdr_rev.get('ESQUERDA_DESC')
                        vt=d_lin.get(tm,"N/A") if tm else"N/A"; dfrente=d_lin.get(fd,"").strip() if fd else""; vf=d_lin.get(fum,"N/A") if fum else"N/A"; dfc=d_lin.get(fudc,"") if fudc else""; vld=d_lin.get(ldm,"N/A") if ldm else"N/A"; ddir=d_lin.get(dd,"") if dd else""; vle=d_lin.get(lem,"N/A") if lem else"N/A"; desq=d_lin.get(ed,"") if ed else""
                        pts=[]
                        if vt!="N/A" and dfrente and dfrente!="-": pts.append(f"Frente: {vt} - Confrontante: {dfrente}") # Corrigido para dfrente
                        if vf!="N/A" and dfc and dfc!="-": pts.append(f"Fundo: {vf} - Confrontante: {dfc}")
                        if vld!="N/A" and ddir and ddir!="-": pts.append(f"Lado Direito: {vld} - Confrontante: {ddir}")
                        if vle!="N/A" and desq and desq!="-": pts.append(f"Lado Esquerdo: {vle} - Confrontante: {desq}")
                        d_lin['CONFRONTANTES']=" <br>".join(pts)
                    except Exception as ec: print(f"Err CONF L{linha_orig}: {ec}"); d_lin['CONFRONTANTES']="Erro Conf."
                    dados_proc.append(d_lin)
        print(f"(Fmt Lote) Varredura FIM. {len(dados_proc)} linhas.")
        if not dados_proc: raise ValueError("Nenhum dado de lote encontrado.")
        df_final=pd.DataFrame(dados_proc); df_final['ETAPA']=1
        total_a=df_final['_area_numerica'].sum()
        df_final['FRAÇÃO IDEAL']=df_final['_area_numerica']/total_a if total_a>0 else 0.0
        df_final=df_final.drop(columns=['_area_numerica'])
        if cabecalho:
            orig_ord=[h for h in cabecalho if h in df_final.columns and h not in ['QUADRA','ETAPA','CONFRONTANTES','FRAÇÃO IDEAL']]
            final_ord=['QUADRA','ETAPA']+orig_ord+['FRAÇÃO IDEAL','CONFRONTANTES']
            final_ex=[c for c in final_ord if c in df_final.columns]; df_final=df_final[final_ex]
        else: print("(Fmt Lote) Warn: Header não detectado, não reordenado.")
        for c in ['QUADRA','ETAPA']:
            if c in df_final.columns: df_final[c]=pd.to_numeric(df_final[c],errors='coerce').astype('Int64')
        if 'FRAÇÃO IDEAL' in df_final.columns: df_final['FRAÇÃO IDEAL']=pd.to_numeric(df_final['FRAÇÃO IDEAL'],errors='coerce')
        print("(Fmt Lote) DF final pronto.")
        output=io.BytesIO(); df_final.to_excel(output,index=False,header=True,engine='openpyxl'); output.seek(0)
        print("(Fmt Lote) Excel em memória."); return output
    except Exception as e: print(f"(Fmt Lote) ERRO GERAL: {e}"); traceback.print_exc(); raise e

# --- Rotas Flask ---

@app.route('/')
def home():
    return render_template('home.html', active_page='home')

# === ROTAS IMPORTAÇÃO CV ===
@app.route('/importacao-cv')
def importacao_cv_index():
    session.pop('cv_uploaded_filename', None); session.pop('cv_basic_info', None)
    session.pop('cv_tipos_unicos', None); session.pop('cv_is_casa_project', None)
    return render_template('importacao_cv.html', active_page='importacao_cv')

@app.route('/upload-cv', methods=['POST'])
def upload_file_cv():
    tool_prefix = 'cv_'
    if 'arquivo_entrada' not in request.files: flash('Nenhum arquivo!', 'error'); return redirect(url_for('importacao_cv_index'))
    file = request.files['arquivo_entrada']
    if file.filename == '': flash('Nenhum arquivo!', 'error'); return redirect(url_for('importacao_cv_index'))
    if file and allowed_file(file.filename):
        filename = secure_filename(f"{tool_prefix}{file.filename}")
        temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        try:
            file.save(temp_filepath)
            df_chk = pd.read_excel(temp_filepath, engine="openpyxl")
            df_chk.columns = df_chk.columns.str.strip()
            is_casa = "CASA" in df_chk.columns or "Quadra" in df_chk.columns
            n_cols = {normalize_text(c): c for c in df_chk.columns}
            t_col_orig = n_cols.get("TIPO")
            if t_col_orig:
                if t_col_orig.upper() != 'TIPO': df_chk.rename(columns={t_col_orig: "TIPO"}, inplace=True)
            else:
                found_alternative = False; alternatives = ["Tipologia", "Tipo da Unidade"]
                for alt in alternatives:
                    t_col_alt_orig = n_cols.get(normalize_text(alt))
                    if t_col_alt_orig: df_chk.rename(columns={t_col_alt_orig: "TIPO"}, inplace=True); found_alternative = True; break
                if not found_alternative: raise ValueError(f"Coluna 'TIPO' (ou alt: {', '.join(alternatives)}) não encontrada!")
            if "TIPO" not in df_chk.columns: raise ValueError("Erro TIPO.")
            tipos = sorted(df_chk["TIPO"].dropna().astype(str).str.strip().unique())
            if not tipos: flash("Aviso: Coluna 'TIPO' vazia.", 'warning')
            session[f'{tool_prefix}uploaded_filename'] = filename; session[f'{tool_prefix}basic_info'] = request.form.to_dict()
            session[f'{tool_prefix}tipos_unicos'] = tipos; session[f'{tool_prefix}is_casa_project'] = is_casa
            return redirect(url_for('map_tipologias_cv_route'))
        except Exception as e:
            if 'temp_filepath' in locals() and os.path.exists(temp_filepath): os.remove(temp_filepath)
            flash(f'Erro CV Upload: {e}','error'); print(f"Err CV Up: {e}"); traceback.print_exc()
            return redirect(url_for('importacao_cv_index'))
    else: flash('Arquivo inválido.','error'); return redirect(url_for('importacao_cv_index'))

@app.route('/map-tipologias-cv')
def map_tipologias_cv_route():
    tool_prefix = 'cv_'
    if f'{tool_prefix}uploaded_filename' not in session: flash('Upload CV primeiro.','warning'); return redirect(url_for('importacao_cv_index'))
    tipos=session.get(f'{tool_prefix}tipos_unicos',[]); is_casa=session.get(f'{tool_prefix}is_casa_project', False)
    return render_template('map_tipologias_cv.html', active_page='importacao_cv', tipos_unicos=tipos, is_casa_project=is_casa,
                           tipologias_padrao=TIPOLOGIAS_PADRAO, tipologias_superior=TIPOLOGIAS_SUPERIOR, tipologias_pcd=TIPOLOGIAS_PCD)

@app.route('/process-cv', methods=['POST'])
def process_file_cv():
    tool_prefix = 'cv_'
    if f'{tool_prefix}uploaded_filename' not in session: flash('Sessão expirada CV.','error'); return redirect(url_for('importacao_cv_index'))
    fname=session[f'{tool_prefix}uploaded_filename']; basic=session[f'{tool_prefix}basic_info']; is_casa=session.get(f'{tool_prefix}is_casa_project',False); tipos_orig=session.get(f'{tool_prefix}tipos_unicos',[])
    fpath=os.path.join(app.config['UPLOAD_FOLDER'],fname)
    if not os.path.exists(fpath): flash('Arquivo temp sumiu.','error'); return redirect(url_for('importacao_cv_index'))
    try:
        tip_map={}
        for t in tipos_orig:
            p,pc,s=request.form.get(f'tipo_{t}_padrao','').strip(), request.form.get(f'tipo_{t}_pcd','').strip(), request.form.get(f'tipo_{t}_superior','').strip() if not is_casa else None
            cp,cpc,cs = TIPOLOGIAS_PADRAO.get(p,p or None), TIPOLOGIAS_PCD.get(pc,pc or None), TIPOLOGIAS_SUPERIOR.get(s,s or None) if s is not None else None
            tip_map[t]={'padrao':cp,'pcd':cpc,'superior':cs}
        df=pd.read_excel(fpath,engine="openpyxl", dtype=str); # Ler como string inicialmente é mais seguro
        df.columns=df.columns.str.strip()
        df = df.fillna('') # Preenche NaN com string vazia
        print("Procurando colunas Bloco/Quadra e Casa/Apto...")
        bloco_quadra_col_name = find_column_flexible(df.columns, ['bloco', 'quadra', 'blk', 'qd'], 'Bloco/Quadra', required=True)
        casa_apto_col_name = find_column_flexible(df.columns, ['casa', 'apto', 'apt', 'apartamento', 'unidade'], 'Casa/Apto', required=True)
        print(f"Coluna Bloco/Quadra encontrada: '{bloco_quadra_col_name}'")
        print(f"Coluna Casa/Apto encontrada: '{casa_apto_col_name}'")
        n_cols={normalize_text(c):c for c in df.columns}; t_col=n_cols.get("TIPO")
        if t_col and t_col.upper()!='TIPO': df.rename(columns={t_col:"TIPO"},inplace=True)
        elif "TIPO" not in df.columns: raise ValueError("Coluna TIPO sumiu.")
        g_col=encontrar_coluna_garagem(df.columns)
        if g_col: df.rename(columns={g_col:"GARAGEM_ORIG"},inplace=True); g_col="GARAGEM_ORIG"
        col_map_ui={"Nome do Empreendimento":"Nome (Empreendimento)","Sigla":"Sigla (Empreendimento)","Empresa":"Empresa (Empreendimento)","Tipo":"Tipo (Empreendimento)","Segmento":"Segmento (Empreendimento)"}
        for k,v in col_map_ui.items(): df[v]=basic.get(k,'')
        col_map_fix={"Endereço":"Endereço (Empreendimento)","CEP":"CEP (Empreendimento)","Região":"Região (Empreendimento)","Bairro":"Bairro (Empreendimento)","Número":"Número (Empreendimento)","Estado":"Estado (Empreendimento)","Cidade":"Cidade (Empreendimento)","Data da Entrega (Empreendimento)":"Data da Entrega (Empreendimento)"}
        for k,v in ENDERECO_FIXO.items():
            if k in col_map_fix: df[col_map_fix[k]]=v
        df["Matrícula (Empreendimento)"]="XXXXX"; df["Ativo no painel (Empreendimento)"]="Ativo"; df["Nome (Etapa)"]="ETAPA 01"; df["Nome (Bloco)"]="BLOCO 01"; df["Ativo no painel (Unidade)"]="Ativo"
        df["Nome (Unidade)"] = df.apply(
            lambda r: formatar_nome_unidade(
                r,
                bq_col_name=bloco_quadra_col_name,
                ca_col_name=casa_apto_col_name
            ),
            axis=1
        )

        g_col=encontrar_coluna_garagem(df.columns) # Função auxiliar existente
        original_garage_col_name = None # Para log
        if g_col:
            original_garage_col_name = g_col
            df.rename(columns={g_col:"GARAGEM_ORIG"},inplace=True)
            g_col="GARAGEM_ORIG" # Usa o nome renomeado internamente
            print(f"Coluna de Garagem encontrada como '{original_garage_col_name}' e renomeada para '{g_col}'")
        else:
            print("Aviso: Coluna de Garagem não encontrada.")
            g_col = None # Garante que g_col seja None se não for encontrada
            
        v_num_mode=basic.get('vaga_por_numero')=='on'
        df["Vagas de garagem (Unidade)"]=df[g_col].apply(lambda x: verificar_vaga(x, v_num_mode)) if g_col else "01 VAGA"
        if g_col: # Verifica se a coluna de garagem foi encontrada e renomeada
            # Aplica a formatação usando a coluna renomeada (GARAGEM_ORIG)
            df["Área de Garagem (Unidade)"] = df[g_col].apply(lambda x: format_decimal_br_cv(x, precision=2))
            print(f"Coluna 'Área de Garagem (Unidade)' criada e formatada a partir de '{original_garage_col_name}'.")
        else:
            # Se nenhuma coluna de garagem foi encontrada, cria a coluna vazia
            df["Área de Garagem (Unidade)"] = ""
        quintal_col_name = find_column_flexible(df.columns, ['quintal', 'jardim'], 'Quintal/Jardim', required=False)
        df["Jardim (Unidade)"]=df[quintal_col_name].apply(formatar_jardim) if quintal_col_name else ""
        area_const_col_name = find_column_flexible(df.columns, ['areaconstruida', 'área construída'], 'Área Construída/Privativa', required=False)
        new_area_col_name = "Área privativa m² (Unidade)"
        if area_const_col_name:
            # Aplica a função de formatação com 2 casas decimais
            df[new_area_col_name] = df[area_const_col_name].apply(lambda x: format_decimal_br_cv(x, precision=2))
            print(f"Coluna 'Área privativa (Unidade)' criada e formatada a partir de '{area_const_col_name}'.")
        else:
            df[new_area_col_name] = "" # Define como vazio se a coluna não for encontrada
            print("Aviso: Coluna de Área Construída/Privativa não encontrada.")

        fracao_col_name = find_column_flexible(df.columns, ['fracaoideal', 'fração ideal'], 'Fração Ideal', required=False)
        # Agora o IF pode usar a variável fracao_col_name
        if fracao_col_name:
             # Aplica a função de formatação com 9 casas decimais (ou quantas precisar)
            df["Fração Ideal (Unidade)"] = df[fracao_col_name].apply(lambda x: format_decimal_br_cv(x, precision=9))
            print(f"Coluna 'Fração Ideal (Unidade)' criada e formatada a partir de '{fracao_col_name}'.")
        else:
            df["Fração Ideal (Unidade)"] = "" # Define como vazio se não encontrar
            print("Aviso: Coluna de Fração Ideal não encontrada.")

        df["Tipo (Unidade)"]=df["TIPO"].astype(str).fillna(''); df["Tipologia (Unidade)"]=df.apply(lambda r: mapear_tipologia_web(r, tip_map, is_casa), axis=1)
        cols_out=["Nome (Empreendimento)","Sigla (Empreendimento)","Matrícula (Empreendimento)","Empresa (Empreendimento)","Tipo (Empreendimento)","Segmento (Empreendimento)","Ativo no painel (Empreendimento)","Região (Empreendimento)","CEP (Empreendimento)","Endereço (Empreendimento)","Bairro (Empreendimento)","Número (Empreendimento)","Estado (Empreendimento)","Cidade (Empreendimento)","Data da Entrega (Empreendimento)","Nome (Etapa)","Nome (Bloco)","Nome (Unidade)","Tipologia (Unidade)","Tipo (Unidade)","Área privativa m² (Unidade)","Jardim (Unidade)","Área de Garagem (Unidade)","Vagas de garagem (Unidade)","Fração Ideal (Unidade)","Ativo no painel (Unidade)"]
        df_final=df[[c for c in cols_out if c in df.columns]]
        output=io.StringIO(); df_final.to_csv(output,index=False,encoding='utf-8-sig',sep=';',quoting=csv.QUOTE_MINIMAL,decimal=',')
        output.seek(0)
        if os.path.exists(fpath): os.remove(fpath)
        session.pop(f'{tool_prefix}uploaded_filename',None); session.pop(f'{tool_prefix}basic_info',None); session.pop(f'{tool_prefix}tipos_unicos',None); session.pop(f'{tool_prefix}is_casa_project',None)
        out_fname=f"importacao_cv_{basic.get('Sigla','output')}.csv"
        return send_file(io.BytesIO(output.getvalue().encode('utf-8-sig')),mimetype='text/csv',as_attachment=True,download_name=out_fname)
    except Exception as e:
        flash(f'Erro CV Process: {e}','error'); print(f"Err CV Proc: {e}"); traceback.print_exc()
        if 'fpath' in locals() and os.path.exists(fpath): os.remove(fpath)
        session.pop(f'{tool_prefix}uploaded_filename',None); session.pop(f'{tool_prefix}basic_info',None); session.pop(f'{tool_prefix}tipos_unicos',None); session.pop(f'{tool_prefix}is_casa_project',None)
        return redirect(url_for('importacao_cv_index'))

# === ROTAS IMPORTAÇÃO CV LOTE ===
@app.route('/importacao-cv-lote', methods=['GET', 'POST'])
def importacao_cv_lote_tool():
    tool_prefix = 'cv_lote_'
    if request.method == 'POST':
        if 'arquivo_entrada' not in request.files: flash('Nenhum arquivo Lote!', 'error'); return redirect(url_for('importacao_cv_lote_tool'))
        file=request.files['arquivo_entrada']
        if file.filename=='': flash('Nenhum arquivo Lote!','error'); return redirect(url_for('importacao_cv_lote_tool'))
        if not file or not allowed_file(file.filename): flash('Tipo inválido Lote.','error'); return redirect(url_for('importacao_cv_lote_tool'))
        basic=request.form.to_dict(); fname=secure_filename(f"{tool_prefix}{file.filename}"); fpath=os.path.join(app.config['UPLOAD_FOLDER'],fname)
        try:
            file.save(fpath); print(f"(Lote) Salvo: {fpath}")
            df=pd.read_excel(fpath,engine="openpyxl",dtype=str); df.columns=df.columns.str.strip()
            n_cols={normalize_text_lote(c):c for c in df.columns}
            cols_norm={"QUADRA":"QUADRA","LOTE":"LOTE","AREACONSTRUIDA":"ÁREA(M2)","AREA CONSTRUIDA":"ÁREA(M2)","AREACONSTRUIDAM2":"ÁREA(M2)","AREA CONSTRUIDA M2":"ÁREA(M2)","AREA(M2)":"ÁREA(M2)","AREAM2":"ÁREA(M2)","FRACAOIDEAL":"FRAÇÃO IDEAL","FRACAO IDEAL":"FRAÇÃO IDEAL","TIPO":"TIPO","CONFRONTANTES":"CONFRONTANTES"}
            cols_found={}; needed=["QUADRA","LOTE","ÁREA(M2)","FRAÇÃO IDEAL","TIPO","CONFRONTANTES"]; missing=set(needed)
            for norm,concept in cols_norm.items():
                orig=n_cols.get(norm);
                if orig and orig in df.columns and concept not in cols_found: cols_found[concept]=orig;
                if concept in missing: missing.remove(concept)
            if missing: raise ValueError(f"Colunas Lote faltando: {', '.join(missing)}")
            print(f"(Lote) Colunas: {cols_found}")
            df["Nome (Empreendimento)"]=basic.get("Nome do Empreendimento",""); df["Sigla (Empreendimento)"]=basic.get("Sigla",""); df["Empresa (Empreendimento)"]=basic.get("Empresa",""); df["Tipo (Empreendimento)"]=basic.get("Tipo","Loteamento"); df["Segmento (Empreendimento)"]=basic.get("Segmento","Residencial");
            df["Matrícula (Empreendimento)"]="XXXXX";
            # <<< CORREÇÃO APLICADA AQUI >>>
            df["Ativo no painel (Empreendimento)"]="Ativo"; # Corrigido de "panel" para "painel"
            # <<< FIM DA CORREÇÃO >>>
            df["Região (Empreendimento)"]=ENDERECO_FIXO["Região"]; df["CEP (Empreendimento)"]=ENDERECO_FIXO["CEP"]; df["Endereço (Empreendimento)"]=ENDERECO_FIXO["Endereço"]; df["Bairro (Empreendimento)"]=ENDERECO_FIXO["Bairro"]; df["Número (Empreendimento)"]=ENDERECO_FIXO["Número"]; df["Estado (Empreendimento)"]=ENDERECO_FIXO["Estado"]; df["Cidade (Empreendimento)"]=ENDERECO_FIXO["Cidade"]; df["Data da Entrega (Empreendimento)"]=ENDERECO_FIXO["Data da Entrega (Empreendimento)"]; df["Nome (Etapa)"]="ETAPA ÚNICA"; df["Ativo no painel (Unidade)"]="Ativo"
            df["Nome (Bloco)"]=df.apply(lambda r: formatar_nome_bloco_lote(r,col_q=cols_found["QUADRA"]),axis=1)
            df["Nome (Unidade)"]=df.apply(lambda r: formatar_nome_unidade_lote(r,col_q=cols_found["QUADRA"],col_l=cols_found["LOTE"]),axis=1)
            a_num=df[cols_found["ÁREA(M2)"]].apply(limpar_converter_numerico_lote); df["Área privativa m² (Unidade)"]=a_num.apply(formatar_area_privativa_lote)
            f_num=df[cols_found["FRAÇÃO IDEAL"]].apply(limpar_converter_numerico_lote); df["Fração Ideal (Unidade)"]=f_num.apply(formatar_fracao_ideal_lote)
            df["Tipo (Unidade)"]=df[cols_found["TIPO"]].astype(str).fillna(''); df["Descrição do Lote (Unidade)"]=df[cols_found["CONFRONTANTES"]].astype(str).fillna('')
            cols_out=["Nome (Empreendimento)","Sigla (Empreendimento)","Matrícula (Empreendimento)","Empresa (Empreendimento)","Tipo (Empreendimento)","Segmento (Empreendimento)","Ativo no painel (Empreendimento)","Região (Empreendimento)","CEP (Empreendimento)","Endereço (Empreendimento)","Bairro (Empreendimento)","Número (Empreendimento)","Estado (Empreendimento)","Cidade (Empreendimento)","Data da Entrega (Empreendimento)","Nome (Etapa)","Nome (Bloco)","Nome (Unidade)","Área privativa m² (Unidade)","Ativo no painel (Unidade)","Fração Ideal (Unidade)","Tipo (Unidade)","Descrição do Lote (Unidade)"]
            missing_out=[c for c in cols_out if c not in df.columns];
            if missing_out: raise ValueError(f"Erro Lote: Colunas finais faltando: {', '.join(missing_out)}")
            df_final=df[cols_out].copy()
            output=io.StringIO(); df_final.astype(str).to_csv(output,index=False,encoding='utf-8-sig',sep=';',quoting=csv.QUOTE_MINIMAL)
            output.seek(0)
            if os.path.exists(fpath): os.remove(fpath)
            out_fname=f"importacao_cv_lote_{basic.get('Sigla','output')}.csv"
            return send_file(io.BytesIO(output.getvalue().encode('utf-8-sig')),mimetype='text/csv',as_attachment=True,download_name=out_fname)
        except Exception as e:
            flash(f"Erro Lote: {e}",'error'); print(f"(Lote) Err: {e}"); traceback.print_exc()
            if 'fpath' in locals() and os.path.exists(fpath): os.remove(fpath)
            return redirect(url_for('importacao_cv_lote_tool'))
    else: # GET
        return render_template('importacao_cv_lote.html', active_page='importacao_cv_lote')

# === ROTAS IMPORTAÇÃO SIENGE ===
@app.route('/importacao-sienge')
def importacao_sienge_index():
    session.pop('sienge_uploaded_filename', None); session.pop('sienge_etapas_unicas', None)
    return render_template('importacao_sienge.html', active_page='importacao_sienge')

@app.route('/upload-sienge', methods=['POST'])
def upload_file_sienge():
    tool_prefix = 'sienge_'
    if 'arquivo_entrada' not in request.files: flash('Nenhum arquivo SIENGE!', 'error'); return redirect(url_for('importacao_sienge_index'))
    file = request.files['arquivo_entrada']
    if file.filename == '': flash('Nenhum arquivo SIENGE!', 'error'); return redirect(url_for('importacao_sienge_index'))
    if not file or not allowed_file(file.filename): flash('Tipo inválido SIENGE.', 'error'); return redirect(url_for('importacao_sienge_index'))
    filename = secure_filename(f"{tool_prefix}{file.filename}")
    temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    try:
        file.save(temp_filepath); print(f"(SIENGE) Arquivo salvo: {temp_filepath}")
        df = pd.read_excel(temp_filepath, engine="openpyxl"); df.columns = df.columns.str.upper().str.strip()
        if "ETAPA" not in df.columns: raise ValueError("Coluna 'ETAPA' não encontrada!")
        try: etapas_u = sorted(df["ETAPA"].dropna().astype(str).unique())
        except Exception as e: raise ValueError(f"Erro ao processar coluna 'ETAPA': {e}")
        if not etapas_u: flash("Nenhuma etapa encontrada.", 'warning') # Permite continuar
        session[f'{tool_prefix}uploaded_filename'] = filename
        session[f'{tool_prefix}etapas_unicas'] = etapas_u
        print(f"(SIENGE) Etapas encontradas: {etapas_u}")
        return redirect(url_for('map_etapas_sienge_route'))
    except Exception as e:
        flash(f"Erro SIENGE: {e}", 'error'); print(f"(SIENGE) Erro upload: {e}"); traceback.print_exc()
        if 'temp_filepath' in locals() and os.path.exists(temp_filepath): os.remove(temp_filepath)
        return redirect(url_for('importacao_sienge_index'))

@app.route('/map-etapas-sienge')
def map_etapas_sienge_route():
    tool_prefix = 'sienge_'
    if f'{tool_prefix}uploaded_filename' not in session: flash('Faça upload SIENGE primeiro.', 'warning'); return redirect(url_for('importacao_sienge_index'))
    etapas = session.get(f'{tool_prefix}etapas_unicas', [])
    return render_template('map_etapas_sienge.html',
                           active_page='importacao_sienge',
                           etapas_unicas=etapas,
                           tool_name="SIENGE",
                           process_url=url_for('process_file_sienge'),
                           cancel_url=url_for('importacao_sienge_index')
                           )

@app.route('/process-sienge', methods=['POST'])
def process_file_sienge():
    tool_prefix = 'sienge_'
    if f'{tool_prefix}uploaded_filename' not in session: flash('Sessão expirada SIENGE.','error'); return redirect(url_for('importacao_sienge_index'))
    fname=session[f'{tool_prefix}uploaded_filename']; fpath=os.path.join(app.config['UPLOAD_FOLDER'],fname); etapas_orig=session.get(f'{tool_prefix}etapas_unicas',[])
    if not os.path.exists(fpath): flash('Arquivo temp SIENGE sumiu.','error'); return redirect(url_for('importacao_sienge_index'))
    try:
        etapas_map={}; any_map=False
        print("(SIENGE) Coletando mapeamento:");
        for et_o in etapas_orig:
            val=request.form.get(f'etapa_{et_o}','').strip(); et_dig=''.join(filter(str.isdigit,str(et_o)))
            if et_dig and val: etapas_map[et_dig]=val; any_map=True; print(f" '{et_dig}' -> '{val}'")
            elif et_dig: print(f" Aviso: Etapa '{et_o}' não mapeada.")
        if not any_map and etapas_orig: print(" Nenhum mapeamento.")
        print("-"*30)
        df=pd.read_excel(fpath,engine="openpyxl"); df.columns=df.columns.str.upper().str.strip()
        bloco_col="QUADRA" if "QUADRA" in df.columns else "BLOCO" if "BLOCO" in df.columns else None
        apt_col="CASA" if "CASA" in df.columns else "APT" if "APT" in df.columns else None
        cols_ess=["ETAPA","ÁREA CONSTRUIDA","FRAÇÃO IDEAL"]
        if not bloco_col and not apt_col: raise ValueError("Faltando cols ID (QUADRA/BLOCO ou CASA/APT)")
        if bloco_col: cols_ess.append(bloco_col);
        if apt_col: cols_ess.append(apt_col)
        missing=[c for c in cols_ess if c not in df.columns];
        if missing: raise ValueError(f"Colunas SIENGE faltando: {', '.join(missing)}")
        def map_et_int(row): et_s=str(row.get('ETAPA','')); et_n=''.join(filter(str.isdigit,et_s)); return etapas_map.get(et_n,None)
        df["EMPREENDIMENTO_CODIGO"]=df.apply(map_et_int,axis=1)
        df_out=pd.DataFrame(); df_out['EMPREENDIMENTO']=df['EMPREENDIMENTO_CODIGO']
        df_out['UNIDADE']=df.apply(lambda r: formatar_unidade_sienge(r,bloco_col,apt_col),axis=1)
        df_out['ÁREA PRIVATIVA']=pd.to_numeric(df['ÁREA CONSTRUIDA'],errors='coerce').fillna(0)
        df_out['ÁREA COMUM']=0; df_out['FRAÇÃO IDEAL']=pd.to_numeric(df['FRAÇÃO IDEAL'],errors='coerce').fillna(0)
        df_out['TIPO DE IMÓVEL']=df.apply(lambda r: determinar_tipo_imovel_sienge(r,apt_col),axis=1)
        df_out['ESTOQUE COMERCIAL']='D'; df_out['ESTOQUE LEGAL']='L'; df_out['ESTOQUE DE OBRA']='C'
        output=io.BytesIO(); wb=xlwt.Workbook(encoding='utf-8'); sheet=wb.add_sheet("Dados")
        for c,h in enumerate(df_out.columns): sheet.write(0,c,h)
        for r, row_data in enumerate(df_out.itertuples(index=False),start=1):
            for c,val in enumerate(row_data):
                if pd.isna(val): sheet.write(r,c,None)
                else:
                    try:
                        if isinstance(val,str) and len(val)>32767: val=val[:32767]
                        sheet.write(r,c,val)
                    except: sheet.write(r,c,str(val))
        wb.save(output); output.seek(0); print("(SIENGE) XLS gerado.")
        if os.path.exists(fpath): os.remove(fpath)
        session.pop(f'{tool_prefix}uploaded_filename',None); session.pop(f'{tool_prefix}etapas_unicas',None)
        out_fname=f"importacao_sienge_{fname.replace(tool_prefix,'').rsplit('.',1)[0]}.xls"
        return send_file(output,mimetype='application/vnd.ms-excel',as_attachment=True,download_name=out_fname)
    except Exception as e:
        flash(f'Erro SIENGE: {e}','error'); print(f"(SIENGE) Err: {e}"); traceback.print_exc()
        if 'fpath' in locals() and os.path.exists(fpath): os.remove(fpath)
        session.pop(f'{tool_prefix}uploaded_filename',None); session.pop(f'{tool_prefix}etapas_unicas',None)
        if f'{tool_prefix}uploaded_filename' in session: return redirect(url_for('map_etapas_sienge_route'))
        else: return redirect(url_for('importacao_sienge_index'))

# === ROTAS IMPORTAÇÃO SIENGE LOTE ===
@app.route('/importacao-sienge-lote')
def importacao_sienge_lote_index():
    session.pop('sienge_lote_uploaded_filename', None); session.pop('sienge_lote_etapas_unicas', None)
    return render_template('importacao_sienge_lote.html', active_page='importacao_sienge_lote')

@app.route('/upload-sienge-lote', methods=['POST'])
def upload_file_sienge_lote():
    tool_prefix = 'sienge_lote_'
    if 'arquivo_entrada' not in request.files: flash('Nenhum arquivo S Lote!', 'error'); return redirect(url_for('importacao_sienge_lote_index'))
    file = request.files['arquivo_entrada']
    if file.filename == '': flash('Nenhum arquivo S Lote!', 'error'); return redirect(url_for('importacao_sienge_lote_index'))
    if not file or not allowed_file(file.filename): flash('Tipo inválido S Lote.', 'error'); return redirect(url_for('importacao_sienge_lote_index'))
    filename = secure_filename(f"{tool_prefix}{file.filename}")
    temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    try:
        file.save(temp_filepath); print(f"(S Lote) Salvo: {temp_filepath}")
        df=pd.read_excel(temp_filepath, engine="openpyxl")
        orig_cols=df.columns.tolist(); df.columns=[normalize_column_name_sienge_lote(c) for c in df.columns] # Aplica normalização S Lote
        print(f"(S Lote) Cols Orig: {orig_cols}"); print(f"(S Lote) Cols Norm: {df.columns.tolist()}")
        cols_nec=["ETAPA","QUADRA","LOTE","ÁREA(M2)","FRAÇÃO IDEAL"]
        missing=[c for c in cols_nec if c not in df.columns]
        if missing: raise ValueError(f"Colunas S Lote faltando: {', '.join(missing)}")
        etapas_u=sorted(df["ETAPA"].dropna().astype(str).unique())
        if not etapas_u: flash("Nenhuma etapa encontrada.",'warning')
        session[f'{tool_prefix}uploaded_filename']=filename; session[f'{tool_prefix}etapas_unicas']=etapas_u
        print(f"(S Lote) Etapas: {etapas_u}")
        return redirect(url_for('map_etapas_sienge_lote_route'))
    except Exception as e:
        flash(f"Erro S Lote: {e}",'error'); print(f"(S Lote) Err Upload: {e}"); traceback.print_exc()
        if 'temp_filepath' in locals() and os.path.exists(temp_filepath): os.remove(temp_filepath)
        return redirect(url_for('importacao_sienge_lote_index'))

@app.route('/map-etapas-sienge-lote')
def map_etapas_sienge_lote_route():
    tool_prefix = 'sienge_lote_'
    if f'{tool_prefix}uploaded_filename' not in session: flash('Upload S Lote primeiro.','warning'); return redirect(url_for('importacao_sienge_lote_index'))
    etapas=session.get(f'{tool_prefix}etapas_unicas',[])
    return render_template('map_etapas_sienge.html', active_page='importacao_sienge_lote',
                           etapas_unicas=etapas, tool_name="SIENGE Lote",
                           process_url=url_for('process_file_sienge_lote'),
                           cancel_url=url_for('importacao_sienge_lote_index'))

@app.route('/process-sienge-lote', methods=['POST'])
def process_file_sienge_lote():
    tool_prefix = 'sienge_lote_'
    if f'{tool_prefix}uploaded_filename' not in session: flash('Sessão expirada S Lote.','error'); return redirect(url_for('importacao_sienge_lote_index'))
    fname=session[f'{tool_prefix}uploaded_filename']; fpath=os.path.join(app.config['UPLOAD_FOLDER'],fname); etapas_orig=session.get(f'{tool_prefix}etapas_unicas',[])
    if not os.path.exists(fpath): flash('Arquivo temp S Lote sumiu.','error'); return redirect(url_for('importacao_sienge_lote_index'))
    try:
        etapas_map={}; any_map=False
        print("(S Lote) Coletando mapeamento:");
        for et_o in etapas_orig:
            val=request.form.get(f'etapa_{et_o}','').strip(); et_dig=''.join(filter(str.isdigit,str(et_o)))
            if et_dig and val: etapas_map[et_dig]=val; any_map=True; print(f" '{et_dig}' -> '{val}'")
            elif et_dig: print(f" Aviso S Lote: Etapa '{et_o}' não mapeada.")
        if not any_map and etapas_orig: print(" Nenhum mapeamento S Lote.")
        print("-"*30)
        df=pd.read_excel(fpath,engine="openpyxl"); df.columns=[normalize_column_name_sienge_lote(c) for c in df.columns] # Renormaliza
        col_q="QUADRA"; col_l="LOTE"; col_a="ÁREA(M2)"; col_f="FRAÇÃO IDEAL"; col_e="ETAPA" # Nomes já normalizados
        cols_nec=[col_e,col_q,col_l,col_a,col_f]; missing=[c for c in cols_nec if c not in df.columns]
        if missing: raise ValueError(f"Colunas S Lote faltando no proc: {', '.join(missing)}")
        def map_et_int(row): et_s=str(row.get(col_e,'')); et_n=''.join(filter(str.isdigit,et_s)); return etapas_map.get(et_n,None)
        df["EMPREENDIMENTO_CODIGO"]=df.apply(map_et_int,axis=1)
        df_out=pd.DataFrame(); df_out['EMPREENDIMENTO']=df['EMPREENDIMENTO_CODIGO']
        df_out['UNIDADE']=df.apply(lambda r: formatar_unidade_sienge_lote(r,col_q,col_l),axis=1)
        df_out['ÁREA PRIVATIVA']=df[col_a].apply(limpar_converter_numerico_sienge_lote)
        df_out['ÁREA COMUM']=0; df_out['FRAÇÃO IDEAL']=df[col_f].apply(limpar_converter_numerico_sienge_lote)
        df_out['TIPO DE IMÓVEL']="LOTE"
        df_out['ESTOQUE COMERCIAL']='D'; df_out['ESTOQUE LEGAL']='L'; df_out['ESTOQUE DE OBRA']='C'
        output=io.BytesIO(); wb=xlwt.Workbook(encoding='utf-8'); sheet=wb.add_sheet("Dados")
        style_a=xlwt.XFStyle(); style_a.num_format_str='0.00'; style_e=xlwt.XFStyle(); style_e.num_format_str='0'; style_f=xlwt.XFStyle(); style_f.num_format_str='0.00000000'; def_s=xlwt.Style.default_style
        try: col_idx={n:i for i,n in enumerate(df_out.columns)}; emp_i,area_i,frac_i = col_idx.get('EMPREENDIMENTO'),col_idx.get('ÁREA PRIVATIVA'),col_idx.get('FRAÇÃO IDEAL')
        except: emp_i=area_i=frac_i=None
        for c,h in enumerate(df_out.columns): sheet.write(0,c,h)
        print(f"(S Lote) Escrevendo {len(df_out)} linhas...")
        for r,row_data in enumerate(df_out.itertuples(index=False),start=1):
            for c,val in enumerate(row_data):
                st=def_s; p_val=val
                if pd.isna(p_val): sheet.write(r,c,None); continue
                if emp_i is not None and c==emp_i:
                    try: p_val=int(float(val)); st=style_e
                    except: pass # Mantem string se não converter
                elif area_i is not None and c==area_i and isinstance(p_val,(int,float)): st=style_a
                elif frac_i is not None and c==frac_i and isinstance(p_val,(int,float)): st=style_f
                try:
                    if isinstance(p_val,str) and len(p_val)>32767: p_val=p_val[:32767]
                    sheet.write(r,c,p_val,st)
                except: sheet.write(r,c,str(p_val),def_s) # Fallback
        wb.save(output); output.seek(0); print("(S Lote) XLS gerado.")
        if os.path.exists(fpath): os.remove(fpath)
        session.pop(f'{tool_prefix}uploaded_filename',None); session.pop(f'{tool_prefix}etapas_unicas',None)
        out_fname=f"importacao_sienge_lote_{fname.replace(tool_prefix,'').rsplit('.',1)[0]}.xls"
        return send_file(output,mimetype='application/vnd.ms-excel',as_attachment=True,download_name=out_fname)
    except Exception as e:
        flash(f'Erro S Lote: {e}','error'); print(f"(S Lote) Err: {e}"); traceback.print_exc()
        if 'fpath' in locals() and os.path.exists(fpath): os.remove(fpath)
        session.pop(f'{tool_prefix}uploaded_filename',None); session.pop(f'{tool_prefix}etapas_unicas',None)
        if f'{tool_prefix}uploaded_filename' in session: return redirect(url_for('map_etapas_sienge_lote_route'))
        else: return redirect(url_for('importacao_sienge_lote_index'))

# === ROTA PARA FORMATADOR INCORPORAÇÃO ===
@app.route('/formatador-incorporacao', methods=['GET', 'POST'])
def formatador_incorporacao_tool():
    tool_prefix = 'incorp_'
    temp_filepath = None # Definir fora do try para garantir acesso no finally
    output_stream = None

    if request.method == 'POST':
        if 'arquivo_entrada' not in request.files: flash('Nenhum arquivo!', 'error'); return redirect(url_for('formatador_incorporacao_tool'))
        file = request.files['arquivo_entrada']
        if file.filename == '': flash('Nenhum arquivo!', 'error'); return redirect(url_for('formatador_incorporacao_tool'))
        # Valida extensão (somente Excel)
        if not file or file.filename.rsplit('.', 1)[1].lower() not in {'xlsx', 'xls'}:
            flash('Tipo de arquivo inválido. Use .xlsx ou .xls.', 'error')
            return redirect(url_for('formatador_incorporacao_tool'))

        filename = secure_filename(f"{tool_prefix}{file.filename}")
        file_stream = io.BytesIO(file.read()) # Lê o arquivo em memória
        file.close()

        try:
            print(f"(Formatador Incorporação Rota) Chamando processar_incorporacao_web...")
            # *** CHAMA A FUNÇÃO CORRETA ***
            output_stream = processar_incorporacao_web(file_stream) # <<< CHAMA A FUNÇÃO QUE REMOVE LINHAS

            input_basename = filename.replace(f"{tool_prefix}", "").rsplit('.', 1)[0]
            # Nome do arquivo de saída pode ser o mesmo
            output_filename = f"planilha_processada_{input_basename}.xlsx"
            print(f"(Formatador Incorporação Rota) Enviando: {output_filename}")

            return send_file(
                output_stream,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=output_filename
            )
        except ValueError as ve:
            flash(f"Erro ao processar: {ve}", 'error'); print(f"(Formatador Incorporação Rota) Erro Validação: {ve}")
            if file_stream and not file_stream.closed: file_stream.close()
            if output_stream and not output_stream.closed: output_stream.close()
            return redirect(url_for('formatador_incorporacao_tool'))
        except Exception as e:
            flash(f"Erro inesperado: {e}", 'error'); print(f"(Formatador Incorporação Rota) Erro Inesperado: {e}"); traceback.print_exc()
            if file_stream and not file_stream.closed: file_stream.close()
            if output_stream and not output_stream.closed: output_stream.close()
            return redirect(url_for('formatador_incorporacao_tool'))
        # Finally não é estritamente necessário aqui pois fechamos os streams no try/except

    else: # GET
        active_page = 'formatador_incorporacao'
        return render_template('formatador_incorporacao.html',
                               active_page=active_page,
                               ALLOWED_EXTENSIONS={'xlsx', 'xls'})

# === ROTA PARA FORMATADOR LOTE ===
@app.route('/formatador-lote', methods=['GET', 'POST'])
def formatador_lote_tool():
    tool_prefix = 'fmt_lote_'
    temp_filepath = None # Definir fora do try
    output_stream = None

    if request.method == 'POST':
        if 'arquivo_entrada' not in request.files: flash('Nenhum arquivo selecionado!', 'error'); return redirect(url_for('formatador_lote_tool'))
        file = request.files['arquivo_entrada']
        if file.filename == '': flash('Nenhum arquivo selecionado!', 'error'); return redirect(url_for('formatador_lote_tool'))
        if not file or not allowed_file(file.filename): flash('Tipo de arquivo inválido.', 'error'); return redirect(url_for('formatador_lote_tool'))
        filename = secure_filename(f"{tool_prefix}{file.filename}"); temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename);
        try:
            file.save(temp_filepath); print(f"(Fmt Lote) Arquivo salvo: {temp_filepath}")
            output_stream = processar_formatador_lote_web(temp_filepath) # Chama a função de processamento
            input_basename = file.filename.rsplit('.', 1)[0]; output_filename = f"{input_basename}_PROCESSADO.xlsx"
            print(f"(Fmt Lote) Enviando: {output_filename}")
            return send_file(output_stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=output_filename)
        except Exception as e:
            flash(f"Erro Fmt Lote: {e}", 'error'); print(f"(Fmt Lote) Erro: {e}"); traceback.print_exc()
             # Limpeza em caso de erro ANTES do send_file
            if os.path.exists(temp_filepath):
                try: os.remove(temp_filepath)
                except OSError: pass
            if output_stream: output_stream.close() # Fecha o stream se deu erro ANTES do send_file
            return redirect(url_for('formatador_lote_tool'))
        finally:
             # Limpeza do ARQUIVO temporário SEMPRE
             if temp_filepath and os.path.exists(temp_filepath):
                try: os.remove(temp_filepath); print(f"(Fmt Lote) Temp removido: {temp_filepath}")
                except OSError as oe: print(f"(Fmt Lote) Erro remover temp: {oe}")
            # NÃO FECHA output_stream aqui

    else: # GET
        return render_template('formatador_lote.html', active_page='formatador_lote')

# Função auxiliar (exemplo, adapte se estiver em utils.py)
def allowed_file(filename):
     ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
     return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    
# === ROTA PARA FORMATADOR TABELA DE PREÇOS ===
@app.route('/formatador-tabela-precos', methods=['GET', 'POST'])
def formatador_tabela_precos_upload():
    tool_prefix = 'tab_precos_'
    session_key = f'{tool_prefix}info'
    print(f"DEBUG: Entrando em {url_for('formatador_tabela_precos_upload')}, Método: {request.method}")

# Dentro da função formatador_tabela_precos_upload em app.py

    if request.method == 'POST':
        print("DEBUG: Processando POST")
        # --- Validações Iniciais ---
        if 'arquivo_entrada' not in request.files:
            flash('Nenhum arquivo selecionado!', 'error')
            print("DEBUG: Saindo - Nenhum arquivo_entrada")
            return redirect(url_for('formatador_tabela_precos_upload'))
        file = request.files['arquivo_entrada']
        if file.filename == '':
            flash('Nenhum arquivo selecionado!', 'error')
            print("DEBUG: Saindo - Nome de arquivo vazio")
            return redirect(url_for('formatador_tabela_precos_upload'))
        # Verifica extensão
        file_ext = '.' in file.filename and file.filename.rsplit('.', 1)[1].lower()
        if not file or file_ext not in {'xlsx', 'xls', 'csv'}:
            flash('Tipo de arquivo inválido. Use .xlsx, .xls ou .csv.', 'error')
            print(f"DEBUG: Saindo - Extensão inválida: {file_ext}")
            return redirect(url_for('formatador_tabela_precos_upload'))

        filename = secure_filename(f"{tool_prefix}{file.filename}")
        temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        is_csv = file_ext == 'csv'
        
        # <<< INÍCIO DAS MODIFICAÇÕES PARA LEITURA CONDICIONAL >>>
        # --- Bloco Try/Except para Leitura e Processamento Inicial ---
        df_check = None
        try:
            file.save(temp_filepath)
            print(f"(Tabela Preços Upload) Arquivo temporário salvo: {temp_filepath}")

            # --- Obter parâmetros CSV do formulário ---
            csv_sep = request.form.get('csv_separator', ';')
            csv_decimal = request.form.get('csv_decimal', ',')
            csv_encoding = request.form.get('csv_encoding', 'utf-8')

            # --- Leitura Condicional ---
            print(f"DEBUG: Tentando ler {'CSV' if is_csv else 'Excel'}")
            if is_csv:
                print(f"DEBUG: CSV Params - Sep: '{csv_sep}', Decimal: '{csv_decimal}', Encoding: '{csv_encoding}'")
                try:
                    df_check = pd.read_csv(temp_filepath, sep=csv_sep, decimal=csv_decimal, encoding=csv_encoding, header=0, dtype=str, skipinitialspace=True)
                except UnicodeDecodeError:
                    print(f"DEBUG: Falha com {csv_encoding}, tentando latin-1...")
                    try:
                        df_check = pd.read_csv(temp_filepath, sep=csv_sep, decimal=csv_decimal, encoding='latin-1', header=0, dtype=str, skipinitialspace=True)
                        print("DEBUG: Leitura com latin-1 ok.")
                    except Exception as e_fallback:
                         raise ValueError(f"Não foi possível ler o CSV (Encoding/Separador/Decimal?). Erro: {e_fallback}")
                except Exception as e_csv:
                    raise ValueError(f"Erro ao ler CSV: {e_csv}")
            else: # Excel
                print("DEBUG: Lendo Excel...")
                NUMERO_DA_LINHA_DO_CABECALHO_NO_EXCEL = 3
                linhas_para_pular = NUMERO_DA_LINHA_DO_CABECALHO_NO_EXCEL - 1
                df_check = pd.read_excel(temp_filepath, engine='openpyxl', skiprows=linhas_para_pular, header=0, dtype=str)

            df_check = df_check.fillna('')
            df_check.columns = df_check.columns.str.strip()

            # --- Lógica para encontrar blocos ---
            if df_check.empty:
                 raise ValueError("Arquivo vazio ou sem dados após cabeçalho.")

            # Certifique-se que find_column_flexible e extract_block_number_safe estão disponíveis/importadas
            bloco_col_name = find_column_flexible(df_check.columns, ['bloco', 'blk', 'quadra'], 'BLOCO', required=True)
            unique_blocks_raw = df_check[bloco_col_name].replace('', np.nan).ffill().dropna().unique().tolist()
            if not unique_blocks_raw:
                 raise ValueError("Nenhum valor de Bloco/Quadra encontrado.")
            unique_blocks = sorted(list(set(str(b).strip() for b in unique_blocks_raw if str(b).strip())),
                                   key=lambda b: extract_block_number_safe(b) if extract_block_number_safe(b) is not None else float('inf'))

            print(f"(Tabela Preços Upload) Blocos únicos: {unique_blocks}")

            # --- Armazena na sessão e redireciona ---
            session_data = {'filename': filename, 'unique_blocks': unique_blocks}
            session[session_key] = session_data
            print("DEBUG: Saindo - Sucesso POST, redirecionando para map_etapas")
            return redirect(url_for('formatador_tabela_precos_map_etapas'))

        except ValueError as ve:
             flash(f"Erro ao ler/processar arquivo: {ve}", 'error')
             print(f"(Tabela Preços Upload) Erro (ValueError): {ve}")
             # Limpeza
             if os.path.exists(temp_filepath): os.remove(temp_filepath)
             session.pop(session_key, None)
             print("DEBUG: Saindo - Erro (ValueError), redirecionando para upload")
             return redirect(url_for('formatador_tabela_precos_upload'))
        except Exception as e:
            flash(f"Erro inesperado: {e}", 'error')
            print(f"(Tabela Preços Upload) Erro Inesperado (Exception): {e}")
            traceback.print_exc()
            # Limpeza
            if 'temp_filepath' in locals() and os.path.exists(temp_filepath): os.remove(temp_filepath)
            session.pop(session_key, None)
            print("DEBUG: Saindo - Erro (Exception), redirecionando para upload")
            return redirect(url_for('formatador_tabela_precos_upload'))
        # Não deve chegar aqui no POST
        # print("!!!!! ALERTA: Fim do bloco POST sem return alcançado !!!!!")

    else: # Método GET
        print("DEBUG: Processando GET")
        print(f"DEBUG: Limpando sessão '{session_key}' (GET)")
        session.pop(session_key, None) # Limpa sessão ao carregar a página inicial
        print("DEBUG: Saindo - Renderizando template inicial (GET)")
        return render_template(
            'formatador_tabela_precos.html',
            active_page='formatador_tabela_precos'
        )
        
# --- NOVA ROTA para Mapeamento de Blocos e Etapas ---
@app.route('/formatador-tabela-precos/map-etapas', methods=['GET', 'POST'])
def formatador_tabela_precos_map_etapas():
    tool_prefix = 'tab_precos_'
    session_key = f'{tool_prefix}info'

    # --- Bloco de verificação de sessão e arquivo ---
    if session_key not in session or 'filename' not in session[session_key] or 'unique_blocks' not in session[session_key]:
        flash('Sessão expirada ou inválida. Por favor, faça o upload novamente.', 'warning')
        return redirect(url_for('formatador_tabela_precos_upload')) # Redireciona para a função de upload

    session_data = session[session_key]
    temp_filename = session_data['filename']
    unique_blocks = session_data['unique_blocks'] # Usado no GET e para pegar names no POST
    temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)

    if not os.path.exists(temp_filepath):
        flash('Arquivo temporário não encontrado. Por favor, faça o upload novamente.', 'error')
        session.pop(session_key, None)
        return redirect(url_for('formatador_tabela_precos_upload'))
    # --- Fim do bloco de verificação ---


    if request.method == 'POST':
        # --- Processamento Final após Mapeamento ---

        block_mapping = {} # Inicializa para evitar UnboundLocalError
        output_stream = None # Inicializa para o finally

        # --- Recria o mapeamento a partir dos dados enviados ---
        print(f"(Tabela Preços Mapeamento) Montando mapeamento a partir do form POST...")
        for block_name_original in unique_blocks:
             stage_assigned = request.form.get(f'stage_for_{block_name_original}', '').strip().upper()
             if stage_assigned:
                 block_mapping[block_name_original] = stage_assigned
             else:
                 # Tratamento se etapa estiver vazia (não deveria acontecer com validação JS)
                 print(f"AVISO BACKEND: Etapa VAZIA para o bloco '{block_name_original}' recebida no POST final.")
                 flash(f"Aviso: O bloco '{block_name_original}' foi enviado sem etapa definida. Verifique o resultado.", "warning")
                 # Considere impedir o processamento se isso for crítico:
                 # flash(f"Erro Crítico: O bloco '{block_name_original}' foi enviado sem etapa. Tente novamente.", "error")
                 # return redirect(url_for('formatador_tabela_precos_map_etapas'))

        print(f"(Tabela Preços Mapeamento) Mapeamento final montado: {block_mapping}")


        # --- Bloco try para processamento e envio ---
        try:
            # Validação extra (opcional, mas recomendada)
            if len(block_mapping) != len(unique_blocks):
                 raise ValueError(f"Inconsistência no mapeamento recebido. Esperados {len(unique_blocks)} blocos, recebidos {len(block_mapping)}. Tente novamente.")

            # Chama a função de processamento
            print(f"(Tabela Preços Mapeamento) Chamando processar_tabela_precos_web...")
            output_stream = processar_tabela_precos_web(temp_filepath, block_mapping)
            print(f"(Tabela Preços Mapeamento) Processamento concluído. Preparando envio...")

            # Define o nome do arquivo de saída
            input_basename = temp_filename.replace(f"{tool_prefix}", "").rsplit('.', 1)[0]
            output_filename = f"{input_basename}_PRECOS_ETAPAS_FORMATADO.xlsx"
            print(f"(Tabela Preços Processo Final) Enviando arquivo processado: {output_filename}")

            # Envia o arquivo processado para download
            response = send_file(
                output_stream,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=output_filename
            )
            session.pop(session_key, None) # Limpa sessão APÓS SUCESSO
            print(f"(Tabela Preços Processo Final) Arquivo enviado, sessão limpa.")
            return response # Retorna a resposta com o arquivo

        except ValueError as ve: # Erro de validação ou processamento esperado
            flash(f"Erro ao processar: {ve}", 'error')
            print(f"(Tabela Preços Processo Final) Erro de validação: {ve}")
            # Não limpa sessão, permite tentar corrigir no mapeamento
            return redirect(url_for('formatador_tabela_precos_map_etapas')) # Redireciona de volta para o mapeamento
        except Exception as e: # Erro inesperado
            flash(f"Erro inesperado ao processar Tabela de Preços: {e}", 'error')
            print(f"(Tabela Preços Processo Final) Erro inesperado: {e}")
            traceback.print_exc()
            session.pop(session_key, None) # Limpa sessão em erro grave
            return redirect(url_for('formatador_tabela_precos_upload')) # Redireciona para o início
        finally:
            # --- Bloco finally para limpar arquivo temporário e stream em caso de erro pré-envio ---
            if os.path.exists(temp_filepath):
                try:
                    os.remove(temp_filepath)
                    print(f"(Tabela Preços Processo Final - Finally) Temp removido: {temp_filepath}")
                except OSError as oe:
                    print(f"(Tabela Preços Processo Final - Finally) Erro remover temp: {oe}")
            # Fecha o stream SOMENTE se ocorreu erro ANTES do send_file ter sido retornado
            # Verifica se 'response' existe E se é um stream (para evitar fechar em erros que não geram stream)
            if output_stream and not ('response' in locals() and hasattr(response, 'is_streamed') and response.is_streamed):
                 try:
                     output_stream.close()
                     print(f"(Tabela Preços Processo Final - Finally) Output stream fechado devido a erro prévio.")
                 except Exception as close_err:
                     print(f"(Tabela Preços Processo Final - Finally) Erro ao fechar output stream: {close_err}")
            # --- Fim do bloco finally ---

    else: # Método GET para /map-etapas (ESSENCIAL!)
        # Renderiza a página de mapeamento, passando a lista de blocos da sessão
        print(f"(Tabela Preços Mapeamento - GET) Renderizando página de mapeamento com blocos: {unique_blocks}")
        return render_template(
            'map_etapas_blocos.html',
            active_page='formatador_tabela_precos',
            unique_blocks=unique_blocks # Passa a lista para o template renderizar os blocos pendentes
        )
    
# --- 1. Tabela Incorporação ---
@app.route('/importacao-preco-incorporacao', methods=['GET', 'POST'])
def importacao_preco_incorporacao_tool():
    tool_prefix = 'preco_incorp_'
    session_key = f'{tool_prefix}session_data' # Chave única para dados da sessão
    temp_filepath = None # Para garantir acesso no finally (embora não usado no finally aqui)

    if request.method == 'POST':
        # 1. Validação básica do arquivo
        if 'arquivo_entrada' not in request.files:
            flash('Nenhum arquivo selecionado!', 'error')
            return redirect(request.url)
        file = request.files['arquivo_entrada']
        if file.filename == '':
            flash('Nenhum arquivo selecionado!', 'error')
            return redirect(request.url)
        # Usando sua função global allowed_file e ALLOWED_EXTENSIONS
        # Certifique-se que ALLOWED_EXTENSIONS inclua 'csv' se for o caso
        if not file or not allowed_file(file.filename):
            flash(f'Tipo de arquivo inválido. Permitidos: {", ".join(ALLOWED_EXTENSIONS)}', 'error')
            return redirect(request.url)

        # 2. Salvar arquivo temporariamente
        filename = secure_filename(f"{tool_prefix}{file.filename}")
        temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        try:
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            file.save(temp_filepath)
            print(f"(Preço Incorp Upload) Arquivo salvo: {temp_filepath}")

            # 3. Ler Cabeçalho e Encontrar Colunas de Valor Candidatas
            linhas_para_ignorar = 2
            try:
                # Lê apenas a linha do cabeçalho após pular as iniciais
                df_header = pd.read_excel(
                    temp_filepath, engine='openpyxl',
                    skiprows=linhas_para_ignorar, header=0, nrows=0
                )
                df_header.columns = df_header.columns.str.strip()
                print(f"(Preço Incorp Upload) Colunas lidas para seleção: {df_header.columns.tolist()}")
            except Exception as e_read:
                 print(f"(Preço Incorp Upload) ERRO ao ler cabeçalho: {e_read}")
                 # Limpa arquivo temporário se a leitura falhar
                 if os.path.exists(temp_filepath): os.remove(temp_filepath)
                 flash(f"Falha ao ler cabeçalho do arquivo Excel: Verifique se o cabeçalho está na linha {linhas_para_ignorar + 1}.", 'error')
                 return redirect(request.url)

            # Encontra todas as colunas que contêm "valor" (normalizado)
            potential_valor_cols = []
            keyword_norm = normalize_text_for_match("valor") # Normaliza a keyword uma vez
            for col in df_header.columns:
                col_norm = normalize_text_for_match(str(col)) # Garante que col é string
                if keyword_norm in col_norm:
                    potential_valor_cols.append(col) # Adiciona o nome ORIGINAL da coluna

            print(f"(Preço Incorp Upload) Colunas candidatas para VALOR: {potential_valor_cols}")

            # 4. Validar se encontrou colunas candidatas
            if not potential_valor_cols:
                if os.path.exists(temp_filepath): os.remove(temp_filepath) # Limpa
                flash('Nenhuma coluna contendo "VALOR" foi encontrada no cabeçalho do arquivo (verificar linha 3).', 'error')
                return redirect(request.url)

            # 5. Armazenar na sessão e redirecionar
            session_data = {
                'temp_filepath': temp_filepath,
                'potential_valor_cols': potential_valor_cols,
                'original_filename': file.filename # Guarda nome original para mostrar
            }
            session[session_key] = session_data
            print(f"(Preço Incorp Upload) Dados salvos na sessão. Redirecionando para seleção.")

            # *** NÃO HÁ CHAMADA PARA processar_preco_incorporacao AQUI ***

            # Redireciona para a nova rota de seleção/confirmação
            return redirect(url_for('confirmar_processar_preco_incorporacao'))

        except Exception as e:
            # Tratamento genérico de erro durante o upload/leitura inicial
            flash(f"Erro inesperado durante o upload: {e}", 'error')
            print(f"(Preço Incorp Upload) Erro inesperado: {e}")
            traceback.print_exc()
            # Garante limpeza do arquivo temporário em caso de erro
            if temp_filepath and os.path.exists(temp_filepath):
                try: os.remove(temp_filepath)
                except OSError: pass
            session.pop(session_key, None) # Limpa sessão em caso de erro
            return redirect(request.url)
        # O finally para limpeza do arquivo agora está na rota de confirmação

    else: # GET (Mostrar formulário de upload inicial)
        active_page = 'importacao_preco_incorporacao'
        session.pop(session_key, None) # Limpa sessão antiga ao carregar a página de upload
        return render_template('importacao_preco_incorporacao.html',
                               active_page=active_page,
                               ALLOWED_EXTENSIONS=ALLOWED_EXTENSIONS)

# --- 2. Tabela Lote à Vista ---
# Dentro de app.py

# --- 2. Tabela Lote à Vista (ROTA AJUSTADA PARA PASSAR BytesIO) ---
@app.route('/importacao-preco-lote-avista', methods=['GET', 'POST'])
def importacao_preco_lote_avista_tool():
    tool_prefix = 'preco_lote_av_'
    temp_filepath = None
    output_bytes_stream = None # Para garantir acesso no finally em caso de erro

    if request.method == 'POST':
        # 1. Validação do arquivo (igual às outras rotas)
        if 'arquivo_entrada' not in request.files:
            flash('Nenhum arquivo selecionado!', 'error')
            return redirect(url_for('importacao_preco_lote_avista_tool'))
        file = request.files['arquivo_entrada']
        if file.filename == '':
            flash('Nenhum arquivo selecionado!', 'error')
            return redirect(url_for('importacao_preco_lote_avista_tool'))
        # Usando sua função global allowed_file e ALLOWED_EXTENSIONS
        if not file or not allowed_file(file.filename):
            flash(f'Tipo de arquivo inválido. Permitidos: {", ".join(ALLOWED_EXTENSIONS)}', 'error')
            return redirect(url_for('importacao_preco_lote_avista_tool'))

        # 2. Salvar arquivo temporariamente
        filename = secure_filename(f"{tool_prefix}{file.filename}")
        temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        try:
            # Garante que a pasta de uploads existe
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            # Salva o arquivo temporariamente
            file.save(temp_filepath)
            print(f"(Preço Lote AV Rota) Arquivo salvo: {temp_filepath}")

            # 3. Ler o conteúdo do arquivo salvo em BytesIO
            file_content_bytesio = None
            try:
                with open(temp_filepath, 'rb') as f: # Abre em modo binário ('rb')
                    file_content = f.read() # Lê todo o conteúdo binário
                file_content_bytesio = io.BytesIO(file_content) # Cria o objeto BytesIO em memória
                print("(Preço Lote AV Rota) Conteúdo do arquivo lido para BytesIO.")
            except Exception as read_err:
                 # Se falhar ao ler o arquivo salvo (permissão, disco cheio, etc.)
                 print(f"(Preço Lote AV Rota) Erro ao ler arquivo temporário '{temp_filepath}': {read_err}")
                 # Limpa o arquivo temporário se ele foi criado mas não pôde ser lido
                 if temp_filepath and os.path.exists(temp_filepath):
                     try: os.remove(temp_filepath)
                     except OSError: pass
                 raise ValueError("Erro ao ler o arquivo temporário após salvar.") from read_err

            # 4. Chama a função específica de processamento, passando o objeto BytesIO
            output_csv_stream = processar_preco_lote_avista(file_content_bytesio) # Passa o objeto, não o path

            # 5. Converte o StringIO retornado pela função para BytesIO para send_file
            output_bytes_stream = io.BytesIO(output_csv_stream.getvalue().encode('utf-8-sig'))
            output_csv_stream.close() # Fecha o StringIO intermediário

            # 6. Define o nome do arquivo de saída .csv
            input_basename = file.filename.rsplit('.', 1)[0]
            output_filename = f"{input_basename}_PREC_LOTE_AV_PROCESSADO.csv"
            print(f"(Preço Lote AV Rota) Enviando: {output_filename}")

            # 7. Envia o arquivo CSV para download
            return send_file(
                output_bytes_stream,
                mimetype='text/csv', # Mimetype correto para CSV
                as_attachment=True,
                download_name=output_filename
            )
        except ValueError as ve: # Captura erros de validação (ex: coluna não encontrada, erro de leitura)
            flash(f"Erro de Validação (Lote à Vista): {ve}", 'error')
            print(f"(Preço Lote AV Rota) Erro Validação: {ve}")
            # Redireciona de volta para a página de upload
            return redirect(url_for('importacao_preco_lote_avista_tool'))
        except Exception as e: # Captura outros erros inesperados durante o processamento
            flash(f"Erro ao processar Tabela Lote à Vista: {e}", 'error')
            print(f"(Preço Lote AV Rota) Erro Inesperado: {e}")
            traceback.print_exc()
            # Tenta fechar o stream de saída se ele foi criado antes do erro
            if output_bytes_stream and not output_bytes_stream.closed:
                 try: output_bytes_stream.close()
                 except Exception: pass
            # Redireciona de volta para a página de upload
            return redirect(url_for('importacao_preco_lote_avista_tool'))
        finally:
             # 8. Limpeza do arquivo temporário (SEMPRE executa, mesmo com erro)
             if temp_filepath and os.path.exists(temp_filepath):
                try:
                    os.remove(temp_filepath)
                    print(f"(Preço Lote AV Rota - Finally) Temp removido: {temp_filepath}")
                except OSError as oe:
                    print(f"(Preço Lote AV Rota - Finally) Erro remover temp '{temp_filepath}': {oe}")

    else: # GET - Mostrar o formulário de upload
        active_page = 'importacao_preco_lote_avista'
        return render_template('importacao_preco_lote_avista.html',
                               active_page=active_page,
                               ALLOWED_EXTENSIONS=ALLOWED_EXTENSIONS) # Passa extensões permitidas

# --- 3. Tabela Lote Parcelado ---
@app.route('/importacao-preco-lote-parcelado', methods=['GET', 'POST'])
def importacao_preco_lote_parcelado_tool():
    tool_prefix = 'preco_lote_parc_'
    temp_filepath = None
    output_bytes_stream = None

    if request.method == 'POST':
        # Validação básica do arquivo (igual)
        if 'arquivo_entrada' not in request.files: flash('Nenhum arquivo!', 'error'); return redirect(url_for('importacao_preco_lote_parcelado_tool'))
        file = request.files['arquivo_entrada']
        if file.filename == '': flash('Nenhum arquivo!', 'error'); return redirect(url_for('importacao_preco_lote_parcelado_tool'))
        if not file or not allowed_file(file.filename): flash(f'Tipo inválido. Permitidos: {", ".join(ALLOWED_EXTENSIONS)}', 'error'); return redirect(url_for('importacao_preco_lote_parcelado_tool'))

        # Obter e validar parâmetros do formulário
        try:
            quantidade_meses = int(request.form.get('quantidade_meses', 0))
            juros_anual_perc = float(request.form.get('juros_anual_perc', -1.0).replace(',', '.'))
            num_anos_parcelas = int(request.form.get('num_anos_parcelas', 0)) # <<< NOVO PARÂMETRO

            # Validações
            if quantidade_meses <= 0: raise ValueError('Quantidade de meses inválida.')
            if juros_anual_perc < 0: raise ValueError('Porcentagem de juros anual inválida.')
            if num_anos_parcelas <= 0: raise ValueError('Número de anos das parcelas inválido.') # <<< NOVA VALIDAÇÃO

        except (ValueError, TypeError) as val_err:
             flash(f'Valores inválidos nos parâmetros: {val_err}', 'error')
             return redirect(url_for('importacao_preco_lote_parcelado_tool'))

        # Salvar arquivo temporário
        filename = secure_filename(f"{tool_prefix}{file.filename}")
        temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        try:
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            file.save(temp_filepath)
            print(f"(Preço Lote Parc Rota) Arquivo salvo: {temp_filepath}")

            # Ler conteúdo em BytesIO
            file_content_bytesio = None
            try:
                with open(temp_filepath, 'rb') as f: file_content = f.read()
                file_content_bytesio = io.BytesIO(file_content)
                print("(Preço Lote Parc Rota) Conteúdo lido para BytesIO.")
            except Exception as read_err:
                 raise ValueError("Erro ao ler o arquivo temporário.") from read_err

            # Chama a função específica passando objeto e TODOS os parâmetros
            output_csv_stream = processar_preco_lote_parcelado(
                file_content_bytesio,
                quantidade_meses,
                juros_anual_perc,
                num_anos_parcelas # <<< PASSA NOVO PARÂMETRO
            )

            # Converte StringIO para BytesIO
            output_bytes_stream = io.BytesIO(output_csv_stream.getvalue().encode('utf-8-sig'))
            output_csv_stream.close()

            # Enviar arquivo CSV
            input_basename = file.filename.rsplit('.', 1)[0]
            output_filename = f"{input_basename}_PREC_LOTE_PARC_{num_anos_parcelas}anos_PROCESSADO.csv" # Nome dinâmico
            print(f"(Preço Lote Parc Rota) Enviando: {output_filename}")
            return send_file(output_bytes_stream, mimetype='text/csv', as_attachment=True, download_name=output_filename)

        except ValueError as ve:
            flash(f"Erro de Validação (Lote Parcelado): {ve}", 'error')
            print(f"(Preço Lote Parc Rota) Erro Validação: {ve}")
            return redirect(url_for('importacao_preco_lote_parcelado_tool'))
        except Exception as e:
            flash(f"Erro ao processar Tabela Lote Parcelado: {e}", 'error')
            print(f"(Preço Lote Parc Rota) Erro: {e}")
            traceback.print_exc()
            if output_bytes_stream and not output_bytes_stream.closed: output_bytes_stream.close()
            return redirect(url_for('importacao_preco_lote_parcelado_tool'))
        finally:
             if temp_filepath and os.path.exists(temp_filepath):
                try: os.remove(temp_filepath); print(f"(Preço Lote Parc Rota - Finally) Temp removido.")
                except OSError as oe: print(f"(Preço Lote Parc Rota - Finally) Erro remover temp: {oe}")

    else: # GET
        active_page = 'importacao_preco_lote_parcelado'
        return render_template('importacao_preco_lote_parcelado.html',
                               active_page=active_page,
                               ALLOWED_EXTENSIONS=ALLOWED_EXTENSIONS)

@app.route('/confirmar-preco-incorporacao', methods=['GET', 'POST'])
def confirmar_processar_preco_incorporacao():
    tool_prefix = 'preco_incorp_'
    session_key = f'{tool_prefix}session_data'
    temp_filepath = None # Para garantir acesso no finally

    # Verificar se os dados necessários estão na sessão
    if session_key not in session or 'temp_filepath' not in session[session_key] or 'potential_valor_cols' not in session[session_key]:
        flash('Sessão inválida ou expirada. Por favor, faça o upload novamente.', 'warning')
        return redirect(url_for('importacao_preco_incorporacao_tool'))

    session_data = session[session_key]
    temp_filepath = session_data['temp_filepath']
    potential_valor_cols = session_data['potential_valor_cols']

    # Verificar se o arquivo temporário ainda existe (importante!)
    if not os.path.exists(temp_filepath):
        flash('Arquivo temporário não encontrado. Por favor, faça o upload novamente.', 'error')
        session.pop(session_key, None)
        return redirect(url_for('importacao_preco_incorporacao_tool'))

    if request.method == 'POST':
        output_bytes_stream = None # Definir para o finally caso erro ocorra antes da criação
        # 1. Obter a coluna selecionada pelo usuário
        selected_col = request.form.get('selected_valor_col')
        if not selected_col:
            flash('Nenhuma coluna de valor foi selecionada.', 'error')
            # Permanece na página de seleção para o usuário corrigir
            return render_template('selecionar_valor_incorporacao.html',
                                   active_page='importacao_preco_incorporacao',
                                   potential_valor_cols=potential_valor_cols,
                                   session_key=session_key)

        # Validação extra: Verifica se a coluna selecionada estava entre as opções
        if selected_col not in potential_valor_cols:
             flash('Seleção inválida. Escolha uma das colunas listadas.', 'error')
             return render_template('selecionar_valor_incorporacao.html',
                                   active_page='importacao_preco_incorporacao',
                                   potential_valor_cols=potential_valor_cols,
                                   session_key=session_key)

        print(f"(Preço Incorp Process) Coluna de Valor Selecionada: '{selected_col}'")

        # 2. Chamar a função de processamento (passando os dois argumentos)
        try:
            # *** CHAMADA CORRETA COM DOIS ARGUMENTOS ***
            output_csv_stream = processar_preco_incorporacao(temp_filepath, selected_col)

            # Converte StringIO para BytesIO para send_file
            output_bytes_stream = io.BytesIO(output_csv_stream.getvalue().encode('utf-8-sig'))
            output_csv_stream.close() # Fecha o StringIO que não será mais usado

            # Define nome do arquivo de saída
            original_filename = session_data.get('original_filename', 'arquivo')
            input_basename = original_filename.rsplit('.', 1)[0]
            output_filename = f"{input_basename}_PREC_INCORP_PROCESSADO.csv"
            print(f"(Preço Incorp Process) Enviando: {output_filename}")

            # 3. Enviar o arquivo
            response = send_file(
                output_bytes_stream,
                mimetype='text/csv',
                as_attachment=True,
                download_name=output_filename
            )

            # 4. Limpar Sessão APÓS sucesso
            session.pop(session_key, None)
            print("(Preço Incorp Process) Sessão limpa após sucesso.")

            # Não limpar o arquivo temporário aqui, o finally cuidará disso

            return response

        except Exception as e:
            # Erro durante o processamento final
            flash(f"Erro ao processar o arquivo: {e}", 'error')
            print(f"(Preço Incorp Process) Erro: {e}")
            traceback.print_exc()
            session.pop(session_key, None) # Limpa sessão em caso de erro
            # Redireciona de volta para a página inicial de upload
            return redirect(url_for('importacao_preco_incorporacao_tool'))

        finally:
            # 5. Limpar Arquivo Temporário (SEMPRE, após tentativa de processamento)
            if temp_filepath and os.path.exists(temp_filepath):
                try:
                    os.remove(temp_filepath)
                    print(f"(Preço Incorp Process - Finally) Temp removido: {temp_filepath}")
                except OSError as oe:
                    print(f"(Preço Incorp Process - Finally) Erro remover temp: {oe}")
            # O stream de bytes será fechado pelo send_file ou já foi fechado (StringIO)

    else: # GET (Mostrar formulário de seleção)
        active_page = 'importacao_preco_incorporacao'
        # É importante passar session_key aqui se o template precisar acessar session[session_key]
        return render_template('selecionar_valor_incorporacao.html',
                               active_page=active_page,
                               potential_valor_cols=potential_valor_cols,
                               session_key=session_key)

# --- Roda a aplicação ---
if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5001)
