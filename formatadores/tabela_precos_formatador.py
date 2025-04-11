# formatadores/tabela_precos_formatador.py

import pandas as pd
import numpy as np
import io
import traceback
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color
from openpyxl.utils import get_column_letter
import unicodedata
import re
from collections import defaultdict # Para agrupar blocos por etapa

# --- Funções Auxiliares (parse_flexible_float, normalize_text, find_column, extract_block_number - Mantidas da v9) ---
def normalize_text_for_match(text):
    if not isinstance(text, str): text = str(text)
    try:
        text = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
        text = text.lower()
        text = re.sub(r'[^a-z0-9]', '', text)
        return text.strip()
    except Exception:
        return str(text).lower().strip()

def find_column_flexible(df_columns, concept_keywords, concept_name, required=True):
    normalized_input_cols = {normalize_text_for_match(col): col for col in df_columns}
    found_col_name = None
    for keyword in concept_keywords: # Match exato
        norm_keyword = normalize_text_for_match(keyword)
        if norm_keyword in normalized_input_cols:
            return normalized_input_cols[norm_keyword]
    potential_matches = []
    for keyword in concept_keywords: # Match parcial
        norm_keyword = normalize_text_for_match(keyword)
        for norm_col, orig_col in normalized_input_cols.items():
            if norm_keyword in norm_col:
                 priority = 0 if norm_col.startswith(norm_keyword) else 1
                 potential_matches.append((priority, orig_col))
    if potential_matches:
        potential_matches.sort()
        return potential_matches[0][1]
    if required: # Erro
        raise ValueError(f"Coluna obrigatória '{concept_name}' não encontrada. Keywords: {concept_keywords}. Cols. normalizadas: {list(normalized_input_cols.keys())}")
    return None

def extract_block_number_safe(block_value_str):
    if not isinstance(block_value_str, str): block_value_str = str(block_value_str)
    match = re.search(r'\d+', block_value_str)
    if match:
        try: return int(match.group(0))
        except ValueError: return None
    return None

def parse_flexible_float(value_str):
    if value_str is None: return None
    text = str(value_str).strip()
    if not text: return None
    cleaned_text = re.sub(r'[^\d,\.-]', '', text)
    if not cleaned_text: return None
    last_dot = cleaned_text.rfind('.')
    last_comma = cleaned_text.rfind(',')
    if last_comma > last_dot:
        num_str = cleaned_text.replace('.', '').replace(',', '.')
    elif last_dot > last_comma:
        num_str = cleaned_text.replace(',', '')
    elif last_comma != -1 and last_dot == -1:
        num_str = cleaned_text.replace(',', '.')
    else:
        num_str = cleaned_text
    try: return float(num_str)
    except (ValueError, TypeError): return None

def format_currency_brl(numeric_value):
    if numeric_value is None: return ""
    try:
        parts = f"{float(numeric_value):.2f}".split('.')
        integer_part = parts[0]; decimal_part = parts[1]
        integer_part_with_dots = ""
        n = len(integer_part)
        for i, digit in enumerate(reversed(integer_part)):
            integer_part_with_dots = digit + integer_part_with_dots
            if (i + 1) % 3 == 0 and i != n - 1: integer_part_with_dots = "." + integer_part_with_dots
        formatted = f"{integer_part_with_dots},{decimal_part}"
        return f"R$ {formatted}"
    except (ValueError, TypeError): return ""

def format_garagem_vagas(numeric_value):
    if numeric_value is None: return "01 VAGA"
    try:
        gn = float(numeric_value)
        if gn > 35: return "04 VAGAS"
        elif gn > 25: return "03 VAGAS"
        elif gn > 15: return "02 VAGAS"
        elif gn >= 0: return "01 VAGA"
        else: return "01 VAGA"
    except (ValueError, TypeError): return "01 VAGA"

def extract_stage_number(stage_name_str):
    """Tenta extrair um número do nome da etapa para ordenação."""
    match = re.search(r'\d+', stage_name_str)
    if match:
        try: return int(match.group(0))
        except ValueError: return float('inf') # Não numérico vai para o fim
    return float('inf') # Sem número vai para o fim

# --- Função Principal de Processamento (v10 - Aceita Mapeamento de Etapas) ---

def processar_tabela_precos_web(input_filepath, block_etapa_mapping): # <-- Novo argumento
    print(f"(Tabela Preços Formatador - v10 Mapeamento Etapas) Iniciando: {input_filepath}")
    try:
        # 1. Ler a planilha (igual v9)
        NUMERO_DA_LINHA_DO_CABECALHO_NO_EXCEL = 3
        linhas_para_pular = NUMERO_DA_LINHA_DO_CABECALHO_NO_EXCEL - 1
        try:
            df_input = pd.read_excel(input_filepath, engine='openpyxl', skiprows=linhas_para_pular, header=0, dtype=str).fillna('')
            df_input.columns = df_input.columns.str.strip()
            if df_input.empty: raise ValueError("Nenhum dado encontrado após cabeçalho.")
            print(f"  Lido {len(df_input)} linhas. Cabeçalho: {df_input.columns.tolist()}")
        except ValueError as ve: raise ve
        except Exception as read_err: raise ValueError(f"Erro ao ler Excel: {read_err}.")

        # 2. Definir Conceitos e Encontrar Colunas (igual v9)
        col_concepts = {
            'BLOCO': (['bloco', 'blk', 'quadra'], True),
            'UNIDADE': (['apt', 'apto', 'apartamento', 'unidade', 'casa', 'unid', 'ap'], True),
            'TIPOLOGIA': (['tipologia', 'tipo', 'descricao', 'descrição'], True),
            'AREA_CONSTRUIDA': (['area construida', 'área construída', 'areaconstruida', 'area util', 'área útil', 'area privativa', 'área privativa'], True),
            'QUINTAL': (['quintal', 'jardim', 'area descoberta', 'área descoberta', 'area externa', 'área externa', 'quintal m2', 'jardim m2'], False),
            'GARAGEM': (['garagem', 'vaga', 'vagas', 'estacionamento'], False),
            'VALOR': (['valor', 'preco', 'preço', 'valor imovel', 'valor do imóvel', 'valor venda'], True)
        }
        found_columns = {}
        print("  Buscando colunas...")
        for concept, (keywords, is_required) in col_concepts.items():
            found_columns[concept] = find_column_flexible(df_input.columns, keywords, concept, required=is_required)
        print(f"  Mapeamento de colunas: {found_columns}")

        # 3. Preparar DataFrame Intermediário (Adiciona Etapa e Bloco Original)
        df_intermediate = pd.DataFrame()
        output_concept_order = ['UNIDADE', 'TIPOLOGIA', 'AREA_CONSTRUIDA', 'QUINTAL', 'GARAGEM', 'VALOR']
        col_bloco_orig = found_columns.get('BLOCO')
        if not col_bloco_orig: raise ValueError("Coluna Bloco obrigatória não encontrada.")

        # Aplica ffill na coluna de bloco original
        df_input[col_bloco_orig] = df_input[col_bloco_orig].replace('', np.nan).ffill()

        # Cria coluna 'BLOCO_ORIGINAL' para manter o valor exato lido (após ffill)
        df_intermediate['BLOCO_ORIGINAL'] = df_input[col_bloco_orig].astype(str)

        # --- Aplica o Mapeamento de Etapas ---
        print(f"  Aplicando mapeamento de etapas fornecido: {block_etapa_mapping}")
        def map_etapa(bloco_original):
            # Trata caso de 'nan' string que pode vir do ffill
            if pd.isna(bloco_original) or str(bloco_original).lower() == 'nan':
                 return "ETAPA_NAO_MAPEADA" # Ou algum outro indicador
            return block_etapa_mapping.get(str(bloco_original).strip(), "ETAPA_NAO_MAPEADA")

        df_intermediate['ETAPA_MAPEADA'] = df_intermediate['BLOCO_ORIGINAL'].apply(map_etapa)

        # Verifica se alguma linha ficou sem etapa mapeada (pode indicar erro no ffill ou mapeamento incompleto)
        if "ETAPA_NAO_MAPEADA" in df_intermediate['ETAPA_MAPEADA'].unique():
            blocos_nao_mapeados = df_intermediate[df_intermediate['ETAPA_MAPEADA'] == "ETAPA_NAO_MAPEADA"]['BLOCO_ORIGINAL'].unique()
            print(f"Aviso: Os seguintes valores de bloco não foram encontrados no mapeamento ou ocorreram após linhas vazias: {blocos_nao_mapeados}. Serão agrupados separadamente.")
            # Poderia lançar um erro aqui se preferir:
            # raise ValueError(f"Blocos não mapeados encontrados: {blocos_nao_mapeados}")

        print("  Copiando dados originais das outras colunas...")
        for concept in output_concept_order:
            original_col_name = found_columns.get(concept)
            # Copia como string, mesmo que a coluna não seja encontrada (resultará em '')
            df_intermediate[concept] = df_input[original_col_name].astype(str).copy() if original_col_name else ''


        # 4. Agrupar e Ordenar por Etapa e Bloco
        etapas_agrupadas = defaultdict(list)
        # Agrupa os nomes originais dos blocos pela etapa mapeada
        for _, row in df_intermediate[['BLOCO_ORIGINAL', 'ETAPA_MAPEADA']].drop_duplicates().iterrows():
            etapa = row['ETAPA_MAPEADA']
            bloco = row['BLOCO_ORIGINAL']
            # Evita adicionar blocos 'nan' se existirem
            if pd.notna(bloco) and str(bloco).lower() != 'nan':
                 etapas_agrupadas[etapa].append(bloco)

        # Ordena as Etapas (numericamente se possível, depois alfabeticamente)
        etapas_ordenadas = sorted(etapas_agrupadas.keys(), key=lambda e: (extract_stage_number(e), e))
        print(f"  Etapas ordenadas para processamento: {etapas_ordenadas}")

        # Ordena os Blocos *dentro* de cada Etapa (numericamente)
        blocos_ordenados_por_etapa = {}
        for etapa in etapas_ordenadas:
            blocos_da_etapa = etapas_agrupadas[etapa]
            blocos_ordenados_da_etapa = sorted(
                blocos_da_etapa,
                key=lambda b: extract_block_number_safe(b) if extract_block_number_safe(b) is not None else float('inf')
            )
            blocos_ordenados_por_etapa[etapa] = blocos_ordenados_da_etapa
            print(f"    Blocos ordenados para {etapa}: {blocos_ordenados_da_etapa}")


        # 5. Construir a Estrutura da Planilha de Saída (COM ETAPAS)
        output_headers = ['UNIDADE', 'TIPOLOGIA', 'ÁREA CONSTRUIDA', 'QUINTAL', 'GARAGEM', 'VALOR']
        num_cols = len(output_headers)
        final_sheet_data = []
        # --- Adiciona título principal (igual antes) ---
        final_sheet_data.extend([([None] * num_cols)] * 2)
        output_title = "TABELA DE PREÇOS"
        final_sheet_data.append([output_title] + [None] * (num_cols - 1))
        final_sheet_data.append([None] * num_cols)
        # Não adiciona mais o "ETAPA ÚNICA" aqui

        row_map = {'title': 3, 'etapas': {}} # Estrutura do row_map ajustada
        current_excel_row = len(final_sheet_data) + 1

        print(f"  Construindo layout final com etapas e blocos ordenados...")
        # Itera sobre as ETAPAS ORDENADAS
        for etapa_idx, etapa_nome in enumerate(etapas_ordenadas):
            etapa_header_excel_row = current_excel_row
            row_map['etapas'][etapa_nome] = {'header_row': etapa_header_excel_row, 'blocks': {}}

            # --- Adiciona Cabeçalho da Etapa ---
            final_sheet_data.append([etapa_nome] + [None] * (num_cols - 1))
            final_sheet_data.append([None] * num_cols)
            current_excel_row += 2

            # Itera sobre os BLOCOS ORDENADOS DENTRO DA ETAPA
            blocos_desta_etapa = blocos_ordenados_por_etapa[etapa_nome]
            for bloco_idx, bloco_val_orig in enumerate(blocos_desta_etapa):
                bloco_header_excel_row = current_excel_row
                data_header_excel_row = current_excel_row + 2
                block_num = extract_block_number_safe(bloco_val_orig)
                # Usa o nome original do bloco para exibição, mas pode formatar se quiser
                block_display_name = f"BLOCO {block_num:02d}" if block_num is not None else str(bloco_val_orig).upper()

                # --- Adiciona Cabeçalhos do Bloco (igual antes) ---
                final_sheet_data.append([block_display_name] + [None] * (num_cols - 1))
                final_sheet_data.append([None] * num_cols)
                final_sheet_data.append(output_headers)

                # --- Filtra e FORMATA os dados do Bloco (lógica da v9 mantida) ---
                # Filtra usando o NOME ORIGINAL do bloco
                df_bloco_data = df_intermediate[df_intermediate['BLOCO_ORIGINAL'] == bloco_val_orig][output_concept_order]
                formatted_data_rows = []
                for _, row in df_bloco_data.iterrows():
                    formatted_row = []
                    for concept in output_concept_order:
                        original_value_str = str(row.get(concept, ''))
                        formatted_val = original_value_str # Fallback

                        if concept in ['AREA_CONSTRUIDA', 'QUINTAL']:
                            numeric_value = parse_flexible_float(original_value_str)
                            placeholder = "--"
                            if numeric_value is not None:
                                if np.isclose(numeric_value, 0): formatted_val = placeholder
                                else: formatted_val = f"{numeric_value:.2f}".replace('.', ',') + " m²"
                            else:
                                if not original_value_str.strip(): formatted_val = placeholder
                        elif concept == 'GARAGEM':
                            numeric_value = parse_flexible_float(original_value_str)
                            if numeric_value is not None: formatted_val = format_garagem_vagas(numeric_value)
                        elif concept == 'VALOR':
                            numeric_value = parse_flexible_float(original_value_str)
                            if numeric_value is not None: formatted_val = format_currency_brl(numeric_value)

                        formatted_row.append(formatted_val)
                    formatted_data_rows.append(formatted_row)

                final_sheet_data.extend(formatted_data_rows)

                # --- Guarda info para estilos (ajustado para aninhar sob etapa) ---
                row_map['etapas'][etapa_nome]['blocks'][bloco_val_orig] = {
                    'bloco_header': bloco_header_excel_row,
                    'blank_after_bloco': bloco_header_excel_row + 1,
                    'data_header': data_header_excel_row,
                    'data_start': data_header_excel_row + 1,
                    'data_end': data_header_excel_row + len(formatted_data_rows)
                }

                # --- Calcula próxima linha ---
                current_excel_row = data_header_excel_row + 1 + len(formatted_data_rows)
                # Adiciona espaço entre blocos (exceto o último da etapa)
                if bloco_idx < len(blocos_desta_etapa) - 1:
                    final_sheet_data.append([None] * num_cols)
                    current_excel_row += 1

            # Adiciona um espaço maior entre ETAPAS (exceto a última)
            if etapa_idx < len(etapas_ordenadas) - 1:
                 final_sheet_data.extend([([None] * num_cols)] * 2) # Duas linhas em branco
                 current_excel_row += 2


        print("  Layout com etapas e formatação concluídos.")

        # 6. Escrever no Excel e Aplicar Estilos Visuais (AJUSTADO PARA ETAPAS)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_final_sheet = pd.DataFrame(final_sheet_data)
            df_final_sheet.to_excel(writer, sheet_name='Tabela Formatada', index=False, header=False)

            workbook = writer.book
            worksheet = writer.sheets['Tabela Formatada']
            print("  Aplicando estilos visuais...")

            # --- Definição de Estilos ---
            title_header_bg_color = "DDEBF7" # Azul claro para Título e Cabeçalhos de dados
            etapa_header_bg_color = "FFF2CC" # Amarelo claro para Etapa
            bloco_header_bg_color = "E2EFDA" # Verde claro para Bloco

            title_fill = PatternFill(start_color=title_header_bg_color, end_color=title_header_bg_color, fill_type="solid")
            etapa_fill = PatternFill(start_color=etapa_header_bg_color, end_color=etapa_header_bg_color, fill_type="solid")
            bloco_fill = PatternFill(start_color=bloco_header_bg_color, end_color=bloco_header_bg_color, fill_type="solid")
            data_header_fill = PatternFill(start_color=title_header_bg_color, end_color=title_header_bg_color, fill_type="solid") # Mesmo do título

            title_font = Font(name='Calibri', size=11, bold=True, color="000000")
            etapa_font = Font(name='Calibri', size=11, bold=True, color="000000")
            bloco_font = Font(name='Calibri', size=11, bold=True, color="000000")
            data_header_font = Font(name='Calibri', size=11, bold=True, color="000000")
            data_font = Font(name='Calibri', size=11, bold=False, color="000000")

            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border_side = Side(style='thin', color="000000")
            thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            top_bottom_border = Border(top=thin_border_side, bottom=thin_border_side)

            # --- Aplicação de Estilos ---
            # Título Principal
            worksheet.merge_cells(start_row=row_map['title'], start_column=1, end_row=row_map['title'], end_column=num_cols)
            cell_title = worksheet.cell(row=row_map['title'], column=1)
            # cell_title.fill = title_fill # Opcional: pode deixar sem fundo
            cell_title.font = title_font
            cell_title.alignment = center_align

            # Itera sobre Etapas e Blocos para aplicar estilos
            for etapa_nome, etapa_info in row_map['etapas'].items():
                etapa_header_r = etapa_info['header_row']

                # Cabeçalho Etapa
                worksheet.merge_cells(start_row=etapa_header_r, start_column=1, end_row=etapa_header_r, end_column=num_cols)
                cell_etapa = worksheet.cell(row=etapa_header_r, column=1)
                cell_etapa.fill = etapa_fill
                cell_etapa.font = etapa_font
                cell_etapa.alignment = center_align
                cell_etapa.border = top_bottom_border # Borda só em cima/baixo

                for bloco_val_orig, rows_info in etapa_info['blocks'].items():
                    bloco_header_r = rows_info['bloco_header']
                    data_header_r = rows_info['data_header']
                    data_start_r = rows_info['data_start']
                    data_end_r = rows_info['data_end']

                    # Cabeçalho Bloco
                    worksheet.merge_cells(start_row=bloco_header_r, start_column=1, end_row=bloco_header_r, end_column=num_cols)
                    cell_bloco = worksheet.cell(row=bloco_header_r, column=1)
                    cell_bloco.fill = bloco_fill
                    cell_bloco.font = bloco_font
                    cell_bloco.alignment = center_align
                    cell_bloco.border = top_bottom_border

                    # Cabeçalho Dados
                    for c_idx in range(1, num_cols + 1):
                        cell = worksheet.cell(row=data_header_r, column=c_idx)
                        cell.fill = data_header_fill
                        cell.font = data_header_font
                        cell.alignment = center_align
                        cell.border = thin_border # Borda completa

                    # Linhas Dados
                    for r in range(data_start_r, data_end_r + 1):
                        for c_idx in range(1, num_cols + 1):
                            cell = worksheet.cell(row=r, column=c_idx)
                            cell.font = data_font
                            cell.alignment = center_align
                            cell.border = thin_border
                            # Aplica formato Texto se o valor for string e não for placeholder ou vazio
                            if cell.value is not None and isinstance(cell.value, str) and cell.value.strip() and cell.value != '--':
                                 cell.number_format = '@'

            # Ajuste de Largura (igual v9)
            col_widths = {'A': 10, 'B': 45, 'C': 18, 'D': 12, 'E': 15, 'F': 20}
            for i, col_letter in enumerate([get_column_letter(j+1) for j in range(num_cols)]):
                 width = col_widths.get(col_letter)
                 if width:
                     try: worksheet.column_dimensions[col_letter].width = width
                     except Exception as e: print(f"Aviso: Falha ao definir largura {col_letter}: {e}")
                 else: print(f"Aviso: Largura não definida para {col_letter}.")

            print("  Estilos visuais aplicados.")

        output.seek(0)
        print("(Tabela Preços Formatador - v10 Mapeamento Etapas) Processamento concluído.")
        return output

    except ValueError as ve:
        print(f"(Tabela Preços Formatador) ERRO VALIDAÇÃO: {ve}")
        raise ve # Re-lança para Flask exibir no flash
    except Exception as e:
        print(f"(Tabela Preços Formatador) ERRO INESPERADO: {e}")
        traceback.print_exc() # Log completo para erros inesperados
        raise RuntimeError(f"Erro inesperado ao formatar Tabela de Preços: {e}") from e