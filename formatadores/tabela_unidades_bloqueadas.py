# formatadores/tabela_unidades_bloqueadas.py

import pandas as pd
import io
import traceback
import re
import unicodedata
import requests

# Imports para estilização com openpyxl
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
)  # Border, Side não são mais tão usados com Tabelas
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo  # Para Tabelas do Excel
from openpyxl import Workbook  # Para criar Excel de mensagem de erro

# Credenciais da API
API_HEADERS = {
    "accept": "application/json",
    "email": "carlos.mauricio@vcaconstrutora.com.br",
    "token": "b3eb66cff818914ff41d0e538301727f3345fdd6",
}
API_BASE_URL_CVBOT = "https://vca.cvcrm.com.br/api/v1/cvbot"
API_BASE_URL_CVIO = "https://vca.cvcrm.com.br/api/v1/cvio"


# --- Funções Auxiliares (find_column_flexible, normalize_text_for_match) ---
def normalize_text_for_match(text):
    if not isinstance(text, str):
        text = str(text)
    try:
        text = (
            unicodedata.normalize("NFKD", text)
            .encode("ASCII", "ignore")
            .decode("ASCII")
        )
        text = text.lower()
        # Mantém espaços para melhor correspondência, remove outros caracteres não alfanuméricos
        text = re.sub(r"[^a-z0-9\s]", "", text)
        return text.strip()
    except Exception:
        # Fallback muito básico se a normalização falhar
        return str(text).lower().strip().replace(" ", "")


def find_column_flexible(df_columns, concept_keywords, concept_name, required=True):
    normalized_input_cols = {normalize_text_for_match(col): col for col in df_columns}
    # print(f"  Buscando '{concept_name}': Keywords={concept_keywords}. Colunas Norm.: {list(normalized_input_cols.keys())}") # Debug
    found_col_name = None

    # 1. Match exato normalizado
    for keyword in concept_keywords:
        norm_keyword = normalize_text_for_match(keyword)
        if norm_keyword in normalized_input_cols:
            found_col_name = normalized_input_cols[norm_keyword]
            # print(f"    -> Match exato norm. '{norm_keyword}' para '{concept_name}'. Col original: '{found_col_name}'")
            return found_col_name

    # 2. Match parcial normalizado (se não houve exato)
    # Prioriza keywords que estão no início do nome da coluna normalizada,
    # depois keywords que estão contidas, e por fim, considera o comprimento da coluna.
    potential_matches = []
    for keyword in concept_keywords:
        norm_keyword = normalize_text_for_match(keyword)
        if not norm_keyword:
            continue  # Pula keywords vazias após normalização

        for norm_col, orig_col in normalized_input_cols.items():
            if not norm_col:
                continue  # Pula colunas vazias após normalização

            if norm_keyword in norm_col:
                # Prioridade: 0 se começa, 1 se contém.
                priority = 0 if norm_col.startswith(norm_keyword) else 1
                potential_matches.append(
                    (priority, len(norm_col), orig_col)
                )  # (prioridade, comprimento da coluna, nome original)
                # Debug: mostra candidatos
                # print(f"    -> Match parcial candidato: '{keyword}' em '{orig_col}' (Norm: '{norm_keyword}' em '{norm_col}') Prio:{priority}, Len:{len(norm_col)}")

    if potential_matches:
        potential_matches.sort()  # Ordena por prioridade (0 vem primeiro), depois por comprimento (menor vem primeiro)
        found_col_name = potential_matches[0][
            2
        ]  # Pega o nome original da melhor correspondência
        # print(f"    -> Melhor match parcial para '{concept_name}'. Col original: '{found_col_name}'")
        return found_col_name

    if required:
        raise ValueError(
            f"Coluna obrigatória '{concept_name}' não encontrada. Keywords usadas: {concept_keywords}. Colunas originais: {list(df_columns)}"
        )
    else:
        # Se não for obrigatório, apenas informa e retorna None
        # print(f"    -> Coluna opcional '{concept_name}' não encontrada.")
        return None


# --- Fim Funções Auxiliares ---


def buscar_empreendimentos_api():
    """Busca todos os empreendimentos da API e retorna um dicionário {nome: id}"""
    url = f"{API_BASE_URL_CVBOT}/empreendimentos"
    try:
        response = requests.get(url, headers=API_HEADERS)
        response.raise_for_status()
        empreendimentos = response.json()
        return {emp["nome"].strip(): emp["idempreendimento"] for emp in empreendimentos}
    except Exception as e:
        print(f"Erro ao buscar empreendimentos: {e}")
        return {}


# Cache para armazenar dados dos empreendimentos
_cache_empreendimentos = {}


def buscar_dados_empreendimento(id_empreendimento):
    """Busca e cacheia os dados de todas as tabelas ativas de um empreendimento"""
    if id_empreendimento in _cache_empreendimentos:
        return _cache_empreendimentos[id_empreendimento]

    try:
        # Passo 1: Busca lista de tabelas ativas
        url_tabelas = f"https://vca.cvcrm.com.br/api/v1/cadastros/empreendimentos/{id_empreendimento}/tabelasdepreco"
        response = requests.get(url_tabelas, headers=API_HEADERS, timeout=10)

        if response.status_code != 200 or not response.text.strip():
            _cache_empreendimentos[id_empreendimento] = []
            return []

        tabelas = response.json()
        if not tabelas:
            _cache_empreendimentos[id_empreendimento] = []
            return []

        # Passo 2: Busca preços de TODAS as tabelas ativas
        todas_tabelas_dados = []
        for tabela in tabelas:
            id_tabela = tabela.get("idtabela")
            url_precos = f"https://vca.cvcrm.com.br/api/v1/cv/tabelasdepreco?idempreendimento={id_empreendimento}&idtabela={id_tabela}"
            response_precos = requests.get(url_precos, headers=API_HEADERS, timeout=10)

            if response_precos.status_code == 200 and response_precos.text.strip():
                dados_precos = response_precos.json()
                if dados_precos and len(dados_precos) > 0:
                    todas_tabelas_dados.append(dados_precos[0])

        _cache_empreendimentos[id_empreendimento] = todas_tabelas_dados
        print(f"  {len(todas_tabelas_dados)} tabela(s) carregada(s)")
        return todas_tabelas_dados

    except Exception as e:
        print(f"  Erro ao buscar dados do empreendimento {id_empreendimento}: {e}")
        _cache_empreendimentos[id_empreendimento] = []
        return []


def buscar_valor_e_vigencia_tabela_preco(id_empreendimento, unidade):
    """
    Busca o valor e a data de vigência da tabela de preço ativa para uma unidade específica.
    Retorna uma tupla (valor, data_vigencia_de).
    """
    todas_tabelas = buscar_dados_empreendimento(id_empreendimento)
    if not todas_tabelas:
        return None, None

    unidade_normalizada = normalize_text_for_match(str(unidade))

    # Busca a unidade em todas as tabelas ativas
    for tabela_dados in todas_tabelas:
        data_vigencia = tabela_dados.get(
            "data_vigencia_de"
        )  # Extrai a data de vigência da tabela
        unidades = tabela_dados.get("unidades", [])
        for un in unidades:
            nome_un = un.get("unidade", "")
            if normalize_text_for_match(str(nome_un)) == unidade_normalizada:
                series = un.get("series", [])
                if series and len(series) > 0:
                    valor = series[0].get("valor")  # Pega o valor da unidade
                    return valor, data_vigencia  # Retorna ambos

    return None, None


def ler_csv_e_extrair_filtros(input_filepath):
    """
    Lê o CSV, identifica colunas de empreendimento e motivo,
    e retorna o DataFrame e listas únicas para filtro, tratando vazios.
    """
    print(f"(Unidades Bloqueadas - Leitura CSV) Lendo: {input_filepath}")
    try:
        with open(input_filepath, "r", encoding="utf-8-sig", errors="replace") as f:
            primeira_linha = f.readline()
            if ";" in primeira_linha and "," not in primeira_linha:
                sep = ";"
            elif "," in primeira_linha and ";" not in primeira_linha:
                sep = ","
            else:
                sep = (
                    ";"
                    if primeira_linha.count(";") >= primeira_linha.count(",")
                    else ","
                )
            print(f"  Separador detectado: '{sep}'")

        df_input = pd.read_csv(
            input_filepath,
            dtype=str,
            sep=sep,
            encoding="utf-8-sig",
            keep_default_na=False,
            na_filter=False,
        )
        # keep_default_na=False e na_filter=False para tratar strings vazias como '' e não NaN inicialmente.

        df_input.columns = [str(col).strip() for col in df_input.columns]

        if df_input.empty:
            raise ValueError(
                "O arquivo CSV está vazio ou não pôde ser lido corretamente."
            )

        col_empreendimento_nome = find_column_flexible(
            df_input.columns,
            ["empreendimento", "projeto", "nome do empreendimento"],
            "Empreendimento",
            required=True,
        )
        col_motivo_bloqueio_nome = find_column_flexible(
            df_input.columns,
            ["motivo do bloqueio", "motivo bloqueio", "motivo"],
            "Motivo do Bloqueio",
            required=True,
        )

        placeholder_vazio = "<VAZIO>"

        # Tratar vazios na coluna de motivo:
        # 1. Strip de espaços em branco
        df_input[col_motivo_bloqueio_nome] = (
            df_input[col_motivo_bloqueio_nome].astype(str).str.strip()
        )
        # 2. Substituir strings vazias '' pelo placeholder
        df_input[col_motivo_bloqueio_nome] = df_input[col_motivo_bloqueio_nome].replace(
            "", placeholder_vazio
        )

        empreendimentos_unicos = sorted(
            list(
                df_input[col_empreendimento_nome]
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .dropna()
                .unique()
            )
        )
        motivos_unicos = sorted(list(df_input[col_motivo_bloqueio_nome].unique()))

        empreendimentos_unicos = [
            emp for emp in empreendimentos_unicos if emp
        ]  # Remove vazios residuais

        print(f"  Coluna de Empreendimento encontrada: '{col_empreendimento_nome}'")
        print(
            f"  Coluna de Motivo do Bloqueio encontrada: '{col_motivo_bloqueio_nome}'"
        )
        print(f"  Empreendimentos únicos: {empreendimentos_unicos}")
        print(f"  Motivos únicos (incluindo placeholder para vazios): {motivos_unicos}")

        return (
            df_input,
            empreendimentos_unicos,
            motivos_unicos,
            col_empreendimento_nome,
            col_motivo_bloqueio_nome,
            placeholder_vazio,
        )

    except Exception as e:
        print(f"(Unidades Bloqueadas - Leitura CSV) ERRO: {e}")
        traceback.print_exc()
        raise


def processar_unidades_bloqueadas_csv(
    df_input,
    col_empreendimento_input,
    col_motivo_input,
    placeholder_vazio_usado,
    empreendimentos_a_ignorar=None,
    motivos_a_ignorar=None,
    buscar_precos=True,
):
    print(f"(Unidades Bloqueadas - Processamento CSV com Tabelas Excel)")
    print(f"  Placeholder para vazios usado na leitura: '{placeholder_vazio_usado}'")
    print(f"  Empreendimentos a ignorar: {empreendimentos_a_ignorar}")
    print(f"  Motivos a ignorar (recebidos da UI): {motivos_a_ignorar}")
    print(f"  Buscar preços da API: {buscar_precos}")

    if empreendimentos_a_ignorar is None:
        empreendimentos_a_ignorar = []
    if motivos_a_ignorar is None:
        motivos_a_ignorar = []

    # Busca IDs dos empreendimentos da API
    empreendimentos_map = {}
    if buscar_precos:
        print("  Buscando IDs dos empreendimentos da API...")
        empreendimentos_map = buscar_empreendimentos_api()
        print(f"  {len(empreendimentos_map)} empreendimentos encontrados na API")

    try:
        col_etapa_input = find_column_flexible(
            df_input.columns, ["etapa"], "Etapa", required=False
        )
        col_bloco_input = find_column_flexible(
            df_input.columns, ["bloco", "quadra"], "Bloco/Quadra", required=True
        )
        col_unidade_input = find_column_flexible(
            df_input.columns,
            ["unidade", "lote", "identificacao da unidade", "identificação da unidade"],
            "Unidade/Lote",
            required=True,
        )
        col_descricao_input = find_column_flexible(
            df_input.columns,
            ["descrição", "descricao", "detalhes"],
            "Descrição",
            required=False,
        )
        col_data_bloqueio_input = find_column_flexible(
            df_input.columns,
            ["data do bloqueio", "data bloqueio", "data"],
            "Data do Bloqueio",
            required=False,
        )

        df_filtrado = df_input.copy()
        if empreendimentos_a_ignorar:
            df_filtrado = df_filtrado[
                ~df_filtrado[col_empreendimento_input]
                .astype(str)
                .str.strip()
                .isin([str(emp).strip() for emp in empreendimentos_a_ignorar])
            ]
        if motivos_a_ignorar:
            df_filtrado = df_filtrado[
                ~df_filtrado[col_motivo_input].isin(motivos_a_ignorar)
            ]

        if df_filtrado.empty:
            output_excel_stream = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultado"
            ws["A1"] = "Nenhum dado encontrado com os filtros aplicados."
            ws["A1"].font = Font(name="Calibri", size=12, bold=True)
            ws.column_dimensions["A"].width = 50
            wb.save(output_excel_stream)
            output_excel_stream.seek(0)
            return output_excel_stream

        # Reverte o placeholder para string vazia ANTES de gerar a saída, se não for para ignorar
        if placeholder_vazio_usado not in motivos_a_ignorar:
            df_filtrado.loc[
                df_filtrado[col_motivo_input] == placeholder_vazio_usado,
                col_motivo_input,
            ] = ""

        empreendimentos_no_df_filtrado = sorted(
            list(
                df_filtrado[col_empreendimento_input]
                .astype(str)
                .str.strip()
                .replace("", pd.NA)
                .dropna()
                .unique()
            )
        )

        colunas_saida_map = {
            col_etapa_input: "Etapa",
            col_bloco_input: "Bloco",
            col_unidade_input: "Unidade",
            col_motivo_input: "Motivo do Bloqueio",
            col_descricao_input: "Descrição",
            col_data_bloqueio_input: "Data do Bloqueio",
        }
        colunas_de_entrada_existentes = [
            col
            for col in [
                col_etapa_input,
                col_bloco_input,
                col_unidade_input,
                col_motivo_input,
                col_descricao_input,
                col_data_bloqueio_input,
            ]
            if col is not None
        ]
        nomes_colunas_saida_ordenadas = [
            colunas_saida_map[col]
            for col in colunas_de_entrada_existentes
            if col in colunas_saida_map
        ]

        # Adiciona coluna de valor da tabela se buscar_precos estiver ativo
        if buscar_precos:
            nomes_colunas_saida_ordenadas.append("Valor Tabela Atual")
            nomes_colunas_saida_ordenadas.append("DATA VIGÊNCIA")

        output_excel_stream = io.BytesIO()
        with pd.ExcelWriter(output_excel_stream, engine="openpyxl") as writer:
            current_row_excel_pandas = 0
            sheet_name_output = "Unidades Bloqueadas"
            empreendimento_style_info = []

            for i, nome_emp in enumerate(empreendimentos_no_df_filtrado):
                df_emp_data_original = df_filtrado[
                    df_filtrado[col_empreendimento_input] == nome_emp
                ].copy()
                if df_emp_data_original.empty:
                    continue

                df_output_emp = pd.DataFrame()
                for col_in, col_out_name in colunas_saida_map.items():
                    if col_in and col_in in df_emp_data_original.columns:
                        df_output_emp[col_out_name] = df_emp_data_original[col_in]
                    elif (
                        col_out_name in nomes_colunas_saida_ordenadas
                        and col_out_name not in ["Valor Tabela Atual", "DATA VIGÊNCIA"]
                    ):  # Se a coluna de saída é esperada mas não há dados de entrada
                        df_output_emp[col_out_name] = pd.Series(
                            [""] * len(df_emp_data_original),
                            index=df_emp_data_original.index,
                            dtype=str,
                        )

                # Busca valores da tabela de preço se habilitado
                if (
                    buscar_precos
                    and "Valor Tabela Atual" in nomes_colunas_saida_ordenadas
                ):
                    id_emp = empreendimentos_map.get(nome_emp.strip())
                    if id_emp:
                        print(f"  Buscando preços para {nome_emp} (ID: {id_emp})...")
                        # Busca dados do empreendimento uma única vez
                        dados_emp = buscar_dados_empreendimento(id_emp)
                        if dados_emp:
                            valores = []
                            vigencias = []
                            for idx, row in df_emp_data_original.iterrows():
                                unidade = row.get(col_unidade_input, "")
                                valor, vigencia = buscar_valor_e_vigencia_tabela_preco(
                                    id_emp, unidade
                                )
                                valores.append(valor if valor is not None else "")
                                vigencias.append(
                                    vigencia if vigencia is not None else ""
                                )
                            df_output_emp["Valor Tabela Atual"] = valores
                            df_output_emp["DATA VIGÊNCIA"] = vigencias
                        else:
                            print(f"  Dados não disponíveis para {nome_emp}")
                            df_output_emp["Valor Tabela Atual"] = [""] * len(
                                df_emp_data_original
                            )
                            df_output_emp["DATA VIGÊNCIA"] = [""] * len(
                                df_emp_data_original
                            )
                    else:
                        print(f"  Empreendimento '{nome_emp}' não encontrado na API")
                        df_output_emp["Valor Tabela Atual"] = [""] * len(
                            df_emp_data_original
                        )
                        df_output_emp["DATA VIGÊNCIA"] = [""] * len(
                            df_emp_data_original
                        )

                # Reordena as colunas conforme esperado
                colunas_existentes = [
                    col
                    for col in nomes_colunas_saida_ordenadas
                    if col in df_output_emp.columns
                ]
                if not df_output_emp.empty:
                    df_output_emp = df_output_emp[colunas_existentes]
                elif (
                    colunas_existentes
                ):  # Se df_output_emp está vazio mas temos nomes de colunas
                    df_output_emp = pd.DataFrame(columns=colunas_existentes)
                else:  # Caso extremo: nenhuma coluna de saída definida
                    continue  # Pula este empreendimento se não houver colunas para mostrar

                if i > 0:
                    current_row_excel_pandas += 1

                empreendimento_title_row_pyxl = current_row_excel_pandas + 1
                current_row_excel_pandas += 1
                start_row_data_table_pandas = current_row_excel_pandas

                # Escreve os nomes das colunas manualmente para a Tabela Excel
                if (
                    nomes_colunas_saida_ordenadas
                ):  # Só escreve cabeçalho se houver colunas
                    temp_header_df = pd.DataFrame([nomes_colunas_saida_ordenadas])
                    temp_header_df.to_excel(
                        writer,
                        sheet_name=sheet_name_output,
                        startrow=start_row_data_table_pandas,
                        index=False,
                        header=False,
                    )

                # Escreve os dados SEM o cabeçalho do Pandas
                if not df_output_emp.empty:
                    df_output_emp.to_excel(
                        writer,
                        sheet_name=sheet_name_output,
                        startrow=start_row_data_table_pandas
                        + (1 if nomes_colunas_saida_ordenadas else 0),
                        index=False,
                        header=False,
                    )

                empreendimento_style_info.append(
                    {
                        "nome": nome_emp,
                        "title_row_pyxl": empreendimento_title_row_pyxl,
                        "table_header_row_pyxl": (
                            start_row_data_table_pandas + 1
                            if nomes_colunas_saida_ordenadas
                            else 0
                        ),
                        "table_data_start_row_pyxl": start_row_data_table_pandas
                        + (2 if nomes_colunas_saida_ordenadas else 1),
                        "num_data_rows": len(df_output_emp),
                        "table_name": f"Tabela_{re.sub(r'[^A-Za-z0-9_]', '', nome_emp.replace(' ', '_'))}_{i}",
                    }
                )

                # Atualiza current_row_excel_pandas
                # Linhas ocupadas: 1 (título) + (1 se header + N dados OU 0 se sem header e sem dados)
                rows_for_table_content = 0
                if nomes_colunas_saida_ordenadas:
                    rows_for_table_content += 1  # Header
                if not df_output_emp.empty:
                    rows_for_table_content += len(df_output_emp)  # Dados

                current_row_excel_pandas += rows_for_table_content
                current_row_excel_pandas += 1  # Linha em branco ABAIXO

            workbook = writer.book
            worksheet = workbook[sheet_name_output]
            font_white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            fill_gray_dark = PatternFill(
                start_color="404040", end_color="404040", fill_type="solid"
            )
            center_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            left_alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

            # Usa as colunas que realmente existem no DataFrame final
            num_cols_tabela = (
                len(nomes_colunas_saida_ordenadas)
                if nomes_colunas_saida_ordenadas
                else 0
            )

            for info in empreendimento_style_info:
                effective_num_cols = max(
                    1, num_cols_tabela
                )  # Pelo menos 1 coluna para o título

                worksheet.merge_cells(
                    start_row=info["title_row_pyxl"],
                    start_column=1,
                    end_row=info["title_row_pyxl"],
                    end_column=effective_num_cols,
                )
                title_cell = worksheet.cell(row=info["title_row_pyxl"], column=1)
                title_cell.value = info["nome"].upper()
                title_cell.font = font_white_bold
                title_cell.fill = fill_gray_dark
                title_cell.alignment = center_alignment
                worksheet.row_dimensions[info["title_row_pyxl"]].height = 20

                if (
                    num_cols_tabela > 0 and info["table_header_row_pyxl"] > 0
                ):  # Só cria tabela e estiliza dados se houver colunas
                    table_range_start_col_letter = get_column_letter(1)
                    table_range_end_col_letter = get_column_letter(num_cols_tabela)
                    table_range_end_row = (
                        info["table_header_row_pyxl"] + info["num_data_rows"]
                    )
                    if table_range_end_row < info["table_header_row_pyxl"]:
                        table_range_end_row = info["table_header_row_pyxl"]
                    table_ref = f"{table_range_start_col_letter}{info['table_header_row_pyxl']}:{table_range_end_col_letter}{table_range_end_row}"

                    excel_table = Table(displayName=info["table_name"], ref=table_ref)
                    style = TableStyleInfo(
                        name="TableStyleLight1",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False,
                    )
                    excel_table.tableStyleInfo = style
                    worksheet.add_table(excel_table)

                    for r_offset in range(info["num_data_rows"]):
                        data_row_num_pyxl = info["table_data_start_row_pyxl"] + r_offset
                        for col_num_pyxl in range(1, num_cols_tabela + 1):
                            cell = worksheet.cell(
                                row=data_row_num_pyxl, column=col_num_pyxl
                            )
                            cell.alignment = left_alignment  # Estilo padrão para dados
                        worksheet.row_dimensions[data_row_num_pyxl].height = 15

            if nomes_colunas_saida_ordenadas:  # Só ajusta largura se houver colunas
                col_valor_index = None
                if "Valor Tabela Atual" in nomes_colunas_saida_ordenadas:
                    col_valor_index = (
                        nomes_colunas_saida_ordenadas.index("Valor Tabela Atual") + 1
                    )

                col_data_vigencia_index = None
                if "DATA VIGÊNCIA" in nomes_colunas_saida_ordenadas:
                    col_data_vigencia_index = (
                        nomes_colunas_saida_ordenadas.index("DATA VIGÊNCIA") + 1
                    )

                for col_idx_1based, col_name_output in enumerate(
                    nomes_colunas_saida_ordenadas, 1
                ):
                    column_letter = get_column_letter(col_idx_1based)
                    max_length = 0
                    if len(str(col_name_output)) > max_length:
                        max_length = len(str(col_name_output))
                    # Itera sobre as linhas REAIS escritas para aquela coluna
                    for row_info in empreendimento_style_info:
                        if (
                            num_cols_tabela > 0
                            and row_info["table_header_row_pyxl"] > 0
                        ):  # Se há tabela
                            # Cabeçalho da Tabela
                            cell_value_header = worksheet.cell(
                                row=row_info["table_header_row_pyxl"],
                                column=col_idx_1based,
                            ).value
                            if cell_value_header:
                                max_length = max(
                                    max_length, len(str(cell_value_header))
                                )
                            # Dados da Tabela - aplica formato moeda se for coluna de valor
                            for r_offset in range(row_info["num_data_rows"]):
                                data_row_num_pyxl = (
                                    row_info["table_data_start_row_pyxl"] + r_offset
                                )
                                cell = worksheet.cell(
                                    row=data_row_num_pyxl, column=col_idx_1based
                                )
                                cell_value_data = cell.value
                                if cell_value_data:
                                    max_length = max(
                                        max_length, len(str(cell_value_data))
                                    )
                                # Aplica formato de moeda brasileira
                                if (
                                    col_idx_1based == col_valor_index
                                    and cell_value_data
                                ):
                                    try:
                                        cell.value = float(cell_value_data)
                                        cell.number_format = "R$ #,##0.00"
                                    except (ValueError, TypeError):
                                        pass
                                # Aplica formato de data
                                if (
                                    col_idx_1based == col_data_vigencia_index
                                    and cell_value_data
                                ):
                                    try:
                                        # Converte a data YYYY-MM-DD para o formato DD/MM/YYYY
                                        cell.value = pd.to_datetime(
                                            cell_value_data
                                        ).strftime("%d/%m/%Y")
                                    except (ValueError, TypeError):
                                        pass  # Mantém o valor original se não for uma data válida

                    adjusted_width = (
                        max_length + 3
                    ) * 1.1  # Um pouco mais de padding e fator
                    if adjusted_width > 60:
                        adjusted_width = 60
                    if adjusted_width < 12:
                        adjusted_width = 12
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        output_excel_stream.seek(0)
        print("(Unidades Bloqueadas - Processamento CSV com Tabelas Excel) Concluído.")
        return output_excel_stream

    except ValueError as ve:
        print(
            f"(Unidades Bloqueadas - Processamento CSV com Estilização) ERRO VALIDAÇÃO: {ve}"
        )
        traceback.print_exc()
        raise ve
    except Exception as e:
        print(
            f"(Unidades Bloqueadas - Processamento CSV com Estilização) ERRO INESPERADO: {e}"
        )
        traceback.print_exc()
        raise RuntimeError(f"Erro inesperado na estilização Excel: {e}") from e
