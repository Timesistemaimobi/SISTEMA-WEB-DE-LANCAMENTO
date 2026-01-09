# formatadores/tabela_precos_formatador.py

import pandas as pd
import numpy as np
import io
import traceback
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import unicodedata
import re
from collections import defaultdict
from openpyxl.utils.cell import range_boundaries


# --- Funções Auxiliares (sem alterações significativas, apenas a adição de .strip() em map_etapa abaixo) ---
def normalize_text_for_match(text):
    """Normaliza texto para busca: minúsculo, sem acentos, sem não-alfanuméricos."""
    if not isinstance(text, str):
        text = str(text)
    try:
        text = (
            unicodedata.normalize("NFKD", text)
            .encode("ASCII", "ignore")
            .decode("ASCII")
        )
        text = text.lower()
        text = re.sub(r"[^a-z0-9]", "", text)
        return text.strip()
    except Exception:
        # Fallback para casos onde a normalização ASCII falha
        text = str(text).lower()
        text = re.sub(r"\s+", "", text)  # Remove espaços
        text = re.sub(r"[^a-z0-9]", "", text)  # Remove não alfanuméricos restantes
        return text.strip()


def find_column_flexible(df_columns, concept_keywords, concept_name, required=True):
    """Encontra a coluna de forma flexível por keywords."""
    normalized_input_cols = {normalize_text_for_match(col): col for col in df_columns}
    # print(f"  Buscando '{concept_name}': Keywords={concept_keywords}. Colunas Norm.: {list(normalized_input_cols.keys())}") # Debug
    found_col_name = None
    # 1. Match exato normalizado
    for keyword in concept_keywords:
        norm_keyword = normalize_text_for_match(keyword)
        if norm_keyword in normalized_input_cols:
            found_col_name = normalized_input_cols[norm_keyword]
            # print(f"    -> Match exato norm. '{norm_keyword}' para '{concept_name}'. Col original: '{found_col_name}'") # Debug
            return found_col_name
    # 2. Match parcial normalizado (prioriza início)
    potential_matches = []
    for keyword in concept_keywords:
        norm_keyword = normalize_text_for_match(keyword)
        for norm_col, orig_col in normalized_input_cols.items():
            if norm_keyword in norm_col:
                priority = 0 if norm_col.startswith(norm_keyword) else 1
                potential_matches.append((priority, orig_col))
                # print(f"    -> Match parcial candidato: '{keyword}' em '{orig_col}' (Norm: '{norm_keyword}' em '{norm_col}') Prio:{priority}") # Debug
    if potential_matches:
        potential_matches.sort()
        found_col_name = potential_matches[0][1]
        # print(f"    -> Melhor match parcial para '{concept_name}'. Col original: '{found_col_name}'") # Debug
        return found_col_name
    if required:
        raise ValueError(
            f"Coluna obrigatória '{concept_name}' não encontrada. Keywords usadas: {concept_keywords}. Colunas disponíveis: {df_columns.tolist()}"
        )
    else:
        # print(f"    -> Coluna opcional '{concept_name}' não encontrada.") # Debug
        return None


def extract_block_number_safe(block_value_str):
    """Extrai o primeiro número de uma string de bloco/quadra."""
    if not isinstance(block_value_str, str):
        block_value_str = str(block_value_str)
    match = re.search(r"\d+", block_value_str)
    if match:
        try:
            return int(match.group(0))
        except ValueError:
            return None
    return None


def parse_flexible_float(value_str):
    """Tenta converter uma string (com R$, m², ',', '.') para float."""
    if value_str is None:
        return None
    text = str(value_str).strip()
    if not text:
        return None

    # Limpeza inicial para remover símbolos comuns não numéricos
    cleaned_text = (
        text.upper().replace("R$", "").replace("M²", "").replace("M2", "").strip()
    )

    # Verifica se contém letras após a limpeza inicial (exceto E para notação científica)
    if re.search(r"[A-DF-Z]", cleaned_text, re.IGNORECASE):
        return None  # Contém letras não permitidas

    # Remove espaços internos que podem atrapalhar a conversão
    parse_ready_text = cleaned_text.replace(" ", "")

    # Lógica de conversão baseada no último separador (vírgula ou ponto)
    last_dot = parse_ready_text.rfind(".")
    last_comma = parse_ready_text.rfind(",")

    try:
        if last_comma > last_dot:  # Provável decimal BR (,) - ex: 1.234,56
            num_str = parse_ready_text.replace(".", "").replace(",", ".")
        elif last_dot > last_comma:  # Provável decimal US (.) - ex: 1,234.56
            num_str = parse_ready_text.replace(",", "")
        # Casos onde só há um tipo de separador ou nenhum
        elif last_comma != -1 and last_dot == -1:  # Só vírgula
            # Trata múltiplos como separador de milhar US (ex: 1,234,567)
            if parse_ready_text.count(",") > 1:
                num_str = parse_ready_text.replace(",", "")
            else:  # Trata como decimal BR (ex: 1,5)
                num_str = parse_ready_text.replace(",", ".")
        elif last_dot != -1 and last_comma == -1:  # Só ponto
            # Trata múltiplos como separador de milhar BR (ex: 1.234.567)
            if parse_ready_text.count(".") > 1:
                num_str = parse_ready_text.replace(".", "")
            else:  # Trata como decimal US (ex: 1.5)
                num_str = parse_ready_text
        else:  # Nenhum separador decimal claro ou apenas dígitos
            num_str = parse_ready_text

        # Remove qualquer caractere não numérico restante exceto '-' no início e 'E'/'e' para notação científica
        num_str = re.sub(r"[^-0-9.eE]", "", num_str)

        result = float(num_str)
        return result
    except (ValueError, TypeError):
        # print(f"AVISO: Falha na conversão final de '{text}' para float.") # Debug
        return None


def format_garagem_vagas(original_value_str, numeric_value):
    """Formata informação de garagem baseado na QUANTIDADE EXATA de vagas."""
    original_clean_str = str(original_value_str).strip()

    # Se vazio ou None, assume 1 vaga
    if not original_clean_str or original_clean_str.lower() == "none":
        return "01 VAGA"

    # Se conseguimos converter para número, usa a quantidade exata
    if numeric_value is not None:
        try:
            count = int(numeric_value)

            # Se for 0 ou negativo, retorna o original
            if count <= 0:
                return original_clean_str

            # Formata com zero-padding e singular/plural correto
            if count == 1:
                return f"{count:02d} VAGA"
            else:
                return f"{count:02d} VAGAS"

        except Exception as e:
            print(f"AVISO: Erro ao formatar vagas com valor {numeric_value}: {e}")
            return original_clean_str
    else:
        # Se não conseguiu converter para número, retorna o texto original
        return original_clean_str


def extract_stage_number(stage_name_str):
    """Extrai o primeiro número de uma string de etapa."""
    if not isinstance(stage_name_str, str):
        stage_name_str = str(stage_name_str)
    match = re.search(r"\d+", stage_name_str)
    if match:
        try:
            return int(match.group(0))
        except ValueError:
            return float("inf")  # Retorna infinito se não converter
    return float("inf")  # Retorna infinito se não achar número


def format_area_m2(numeric_value):
    """Formata um valor numérico como 'XXX,XX m²'. Retorna '--' se None, 0 ou inválido."""
    if numeric_value is None or pd.isna(numeric_value):
        return "--"
    try:
        val = float(numeric_value)
        # Usar np.isclose para comparar float com zero de forma segura
        if np.isclose(val, 0):
            return "--"
        # Usa formatação de string f para garantir duas casas decimais e substitui ponto por vírgula
        return f"{val:.2f}".replace(".", ",") + " m²"
    except (ValueError, TypeError):
        print(
            f"AVISO: Erro ao formatar valor de área {numeric_value}. Retornando '--'."
        )
        return "--"


# <<< INÍCIO DA MODIFICAÇÃO: Nova função auxiliar para o formato composto >>>
def format_composite_unit_name(row):
    """Gera o nome da unidade no formato BL/US-QD-CS, lendo direto do DataFrame intermediário."""
    try:
        # Lê os valores das colunas usando os nomes dos CONCEITOS
        quadra_val = str(row.get("QUADRA_COMPOSITE", ""))
        bloco_val = str(row.get("BLOCO_COMPOSITE", ""))
        casa_val = str(row.get("CASA_COMPOSITE", ""))

        quadra_num = extract_block_number_safe(quadra_val)
        bloco_num = extract_block_number_safe(bloco_val)
        casa_num = extract_block_number_safe(casa_val)

        quadra_str = f"QD{quadra_num:02d}" if quadra_num is not None else "QD??"
        casa_str = f"CS{casa_num:02d}" if casa_num is not None else "CS??"

        prefixo_str = ""
        # Verifica se o valor da coluna Bloco é exatamente 'US' (ignorando case e espaços)
        if bloco_val.strip().upper() == "US":
            # Se for 'US', o número vem da casa
            prefixo_str = f"US{casa_num:02d}" if casa_num is not None else "US??"
        else:
            # Para qualquer outro valor (número do bloco), usa 'BL'
            prefixo_str = f"BL{bloco_num:02d}" if bloco_num is not None else "BL??"

        return f"{prefixo_str}-{quadra_str}-{casa_str}"

    except Exception as e:
        print(f"AVISO: Erro ao gerar nome de unidade composto: {e}")
        return "ERRO_UNIDADE"


# --- Função Principal de Processamento (REVISADA) ---
def format_brl(numeric_value):
    """Formata um valor numérico como moeda BRL (R$ X.XXX,XX)."""
    if numeric_value is None:
        return ""
    try:
        val = float(numeric_value)
        return f"R$ {val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return str(numeric_value)


def processar_tabela_precos_web(
    input_filepath,
    block_etapa_mapping,
    selected_columns_order=None,
    column_format_map=None,
):
    """
    Processa a tabela de preços.
    MODIFICADO: Aceita selected_columns_order e column_format_map para controle manual.
    """
    print(f"(Tabela Preços Formatador - v_Refatorada) Iniciando: {input_filepath}")

    try:
        # 1. Leitura Robusta do Arquivo (Excel/CSV)
        df_input = None
        file_type_used = None
        try:
            df_input = pd.read_excel(
                input_filepath, engine="openpyxl", skiprows=2, header=0, dtype=str
            )
            file_type_used = "Excel (header linha 3)"
        except Exception as e_excel_h3:
            try:
                df_input = pd.read_excel(
                    input_filepath, engine="openpyxl", header=0, dtype=str
                )
                file_type_used = "Excel (header linha 1)"
            except Exception as e_excel_h1:
                try:
                    df_input = pd.read_csv(
                        input_filepath,
                        sep=";",
                        decimal=",",
                        encoding="utf-8",
                        header=0,
                        dtype=str,
                        skipinitialspace=True,
                    )
                    file_type_used = "CSV (sep=';')"
                except Exception as e_csv_1:
                    try:
                        df_input = pd.read_csv(
                            input_filepath,
                            sep=",",
                            decimal=".",
                            encoding="utf-8",
                            header=0,
                            dtype=str,
                            skipinitialspace=True,
                        )
                        file_type_used = "CSV (sep=',')"
                    except Exception as e_csv_2:
                        raise ValueError(f"Falha na leitura do arquivo: {e_csv_2}")

        print(f"  Arquivo lido com sucesso como: {file_type_used}")

        # Limpeza básica
        df_input.columns = df_input.columns.str.strip()
        df_input.dropna(how="all", inplace=True)
        if df_input.empty:
            raise ValueError("Arquivo vazio.")

        # --- Identificação de Colunas para Formatação (Apenas para aplicar R$ ou m²) ---
        # Não filtra colunas, apenas identifica para saber como formatar VALUES
        # --- Identificação de Colunas para Formatação (Apenas para aplicar R$ ou m²) ---
        # Não filtra colunas, apenas identifica para saber como formatar VALUES
        col_valor_detected = find_column_flexible(
            df_input.columns,
            ["valor", "preco", "preço", "total"],
            "VALOR",
            required=False,
        )

        # <<< AMPLIAÇÃO: Detectar múltiplas colunas de Área >>>
        area_keywords = [
            "area",
            "área",
            "privativa",
            "construida",
            "quintal",
            "frontal",
            "garagem",
            "terreno",
        ]
        detected_area_columns = []
        for col in df_input.columns:
            col_lower = str(col).lower()
            if any(k in col_lower for k in area_keywords):
                detected_area_columns.append(col)

        # Também precisamos identificar Bloco/Quadra para o Agrupamento de Etapas
        col_bloco_agrupamento = find_column_flexible(
            df_input.columns,
            ["bloco", "blk", "quadra", "pavimento"],
            "AGRUPADOR",
            required=False,
        )
        if not col_bloco_agrupamento:
            # Se não achar, não conseguiremos agrupar etapas corretamente, mas seguimos
            print(
                "AVISO: Coluna de Bloco/Quadra/Pavimento não encontrada automaticamente. O agrupamento por etapas pode falhar."
            )

        # --- FFILL em colunas comuns de agrupamento ---
        cols_to_ffill = []
        if col_bloco_agrupamento:
            cols_to_ffill.append(col_bloco_agrupamento)

        # Tenta achar outras comuns para ffill mesmo que não escolhidas explicitamente, para garantir dados
        for c_key in ["quadra", "pavimento", "andar"]:
            c_found = find_column_flexible(
                df_input.columns, [c_key], c_key, required=False
            )
            if c_found and c_found not in cols_to_ffill:
                cols_to_ffill.append(c_found)

        if cols_to_ffill:
            print(f"  Aplicando ffill em: {cols_to_ffill}")
            df_input[cols_to_ffill] = (
                df_input[cols_to_ffill].replace("", np.nan).ffill()
            )

        df_input = df_input.fillna("")

        # --- Definição das Colunas de Saída ---
        # Se receiving selected_columns_order, usa ele. Senão, usa todas as colunas do input.
        final_output_cols = (
            selected_columns_order
            if selected_columns_order
            else df_input.columns.tolist()
        )

        # Filtra para ter certeza que existem no DF (evita o erro key error)
        valid_output_cols = [c for c in final_output_cols if c in df_input.columns]
        if "ETAPA" not in valid_output_cols:
            valid_output_cols.insert(0, "ETAPA")

        # Lógica de Classificação DEPOIS de selecionar as colunas (para saber como formatar)
        # 1. Valor: "valor", "preco", "total"
        # 2. Vagas: "vaga"
        # 3. Área: "area", "quintal", "frontal", "terreno", "construida", "privativa", "garagem" (se não for vaga)

        cols_to_format_valor = []
        cols_to_format_area = []
        cols_to_format_vagas = []

        # Se o usuário forneceu um mapa de formatação manual, usa ele como prioridade
        if column_format_map:
            print(f"  Usando formatações manuais do usuário: {column_format_map}")
            for col in valid_output_cols:
                if col == "ETAPA":
                    continue
                manual_format = column_format_map.get(col)
                if manual_format == "price":
                    cols_to_format_valor.append(col)
                elif manual_format == "area":
                    cols_to_format_area.append(col)
                elif manual_format == "vagas":
                    cols_to_format_vagas.append(col)
        else:
            # Fallback: Detecção automática (lógica original)
            print("  Usando detecção automática de formatação")
            valor_keywords = ["valor", "preco", "preço", "total"]
            vaga_keywords = ["vaga"]
            area_keywords = [
                "area",
                "área",
                "privativa",
                "construida",
                "quintal",
                "frontal",
                "garagem",
                "terreno",
            ]

            for col in valid_output_cols:
                if col == "ETAPA":
                    continue
                c_lower = str(col).lower()

                # Prioridade 1: Vagas (pois Garagem pode ser confundida com Área)
                if any(k in c_lower for k in vaga_keywords):
                    cols_to_format_vagas.append(col)
                    continue

                # Prioridade 2: Valor
                if any(k in c_lower for k in valor_keywords) and not "m²" in c_lower:
                    cols_to_format_valor.append(col)
                    continue

                # Prioridade 3: Área
                if any(k in c_lower for k in area_keywords):
                    cols_to_format_area.append(col)
                    continue

        print(f"  Colunas selecionadas para saída: {valid_output_cols}")

        # --- Processamento dos Dados ---
        df_output = pd.DataFrame()

        # Gera a coluna ETAPA (Lógica de Mapeamento)
        if col_bloco_agrupamento and block_etapa_mapping:

            def map_etapa(val):
                s = str(val).strip()
                return block_etapa_mapping.get(s, "ETAPA_NAO_MAPEADA")

            df_output["ETAPA"] = df_input[col_bloco_agrupamento].apply(map_etapa)
        else:
            df_output["ETAPA"] = ""

        # Copia e Formata as colunas selecionadas
        for col in valid_output_cols:
            if col == "ETAPA":
                continue  # Já tratada

            # Aplica formatações baseadas nas listas preenchidas acima
            if col in cols_to_format_valor:
                # Tenta converter e formatar moeda
                df_output[col] = df_input[col].apply(
                    lambda x: format_brl(parse_flexible_float(x))
                )
            elif col in cols_to_format_area:
                # Formatação de Área
                df_output[col] = df_input[col].apply(
                    lambda x: format_area_m2(parse_flexible_float(x))
                )
            elif col in cols_to_format_vagas:
                # Formatação de Vagas (Lógica Específica: X VAGAS)
                # Tenta ler como número primeiro para categorizar se necessário
                df_output[col] = df_input[col].apply(
                    lambda x: format_garagem_vagas(x, parse_flexible_float(x))
                )
            else:
                # Copia as is
                df_output[col] = df_input[col].astype(str)

        # --- Ordenação (Lógica Complexa de Etapas) ---
        # Requer que tenhamos ETAPA e alguma coluna de identificação (Bloco/Unidade)
        # Se tivermos col_bloco_agrupamento, usamos ele para ordenar dentro da etapa

        if "ETAPA" in df_output.columns:
            # Ordena por Etapa (numérica se possível)
            df_output["__sort_etapa"] = df_output["ETAPA"].apply(extract_stage_number)

            sort_cols = ["__sort_etapa", "ETAPA"]
            if col_bloco_agrupamento and col_bloco_agrupamento in df_output.columns:
                # Ordenar por Agrupador como string mas tentando numérico se possível
                df_output["__sort_gp"] = df_output[col_bloco_agrupamento].apply(
                    lambda x: (
                        extract_block_number_safe(x)
                        if extract_block_number_safe(x) is not None
                        else 99999
                    )
                )
                sort_cols.extend(["__sort_gp", col_bloco_agrupamento])

            # Adicionar ordenação por Unidade ou similar se possível para ficar bonito
            if "UNIDADE" in df_output.columns:
                pass  # Já está bom

            df_output.sort_values(by=sort_cols, inplace=True)
            df_output.drop(
                columns=[c for c in df_output.columns if c.startswith("__sort_")],
                inplace=True,
            )

        # --- Geração do Excel com openpyxl (Estilização) ---
        num_cols = len(valid_output_cols)
        final_sheet_data = []

        # Título
        final_sheet_data.extend([[None] * num_cols] * 2)
        final_sheet_data.append(["TABELA DE PREÇOS"] + [None] * (num_cols - 1))
        final_sheet_data.append([None] * num_cols)

        current_excel_row = len(final_sheet_data) + 1
        row_map = {"title": 3, "etapas": {}}

        # Agrupamento para Output Visual (Cabeçalhos de Etapa e Blocos)
        # Estrutura: Etapa -> Bloco -> Cabeçalho Colunas -> Dados

        # Pega lista única de etapas na ordem ordenada
        unique_etapas = df_output["ETAPA"].unique()

        for etapa in unique_etapas:
            # Linha de Cabeçalho da Etapa
            etapa_header_row = current_excel_row
            row_map["etapas"][str(etapa)] = {
                "header_row": etapa_header_row,
                "blocks": {},
            }

            final_sheet_data.append([str(etapa)] + [None] * (num_cols - 1))
            final_sheet_data.append([None] * num_cols)  # Linha vazia pós etapa header
            current_excel_row += 2

            # Dados da Etapa - agrupados por Bloco se disponível
            subset = df_output[df_output["ETAPA"] == etapa]

            # Verifica se temos coluna de bloco/agrupamento para criar sub-grupos
            if col_bloco_agrupamento and col_bloco_agrupamento in df_output.columns:
                # Agrupa por bloco dentro da etapa
                unique_blocos = subset[col_bloco_agrupamento].unique()

                for bloco in unique_blocos:
                    # Linha de Cabeçalho do Bloco
                    bloco_header_row = current_excel_row
                    row_map["etapas"][str(etapa)]["blocks"][str(bloco)] = {
                        "header_row": bloco_header_row,
                    }

                    # Formata o nome do bloco como "BLOCO XX"
                    bloco_num = extract_block_number_safe(bloco)
                    bloco_texto = (
                        f"BLOCO {bloco_num:02d}"
                        if bloco_num is not None
                        else str(bloco)
                    )
                    final_sheet_data.append([bloco_texto] + [None] * (num_cols - 1))
                    final_sheet_data.append(
                        [None] * num_cols
                    )  # Linha vazia pós bloco header
                    current_excel_row += 2

                    # Cabeçalho das colunas para este bloco
                    columns_header_row = current_excel_row
                    header_row_vals = [str(c).upper() for c in valid_output_cols]
                    final_sheet_data.append(header_row_vals)
                    current_excel_row += 1

                    # Dados do bloco
                    bloco_subset = subset[subset[col_bloco_agrupamento] == bloco]
                    data_start_row = current_excel_row
                    data_end_row = (
                        data_start_row + len(bloco_subset) - 1
                        if len(bloco_subset) > 0
                        else None
                    )
                    row_map["etapas"][str(etapa)]["blocks"][str(bloco)].update(
                        {
                            "columns_header_row": columns_header_row,
                            "data_start_row": data_start_row,
                            "data_end_row": data_end_row,
                        }
                    )
                    for _, row_data in bloco_subset.iterrows():
                        row_vals = [row_data[c] for c in valid_output_cols]
                        final_sheet_data.append(row_vals)
                        current_excel_row += 1

                    final_sheet_data.append(
                        [None] * num_cols
                    )  # Espaço após tabela do bloco
                    current_excel_row += 1
            else:
                # Sem agrupamento por bloco - comportamento original
                columns_header_row = current_excel_row
                header_row_vals = [str(c).upper() for c in valid_output_cols]
                final_sheet_data.append(header_row_vals)
                current_excel_row += 1
                data_start_row = current_excel_row
                data_end_row = (
                    data_start_row + len(subset) - 1 if len(subset) > 0 else None
                )
                row_map["etapas"][str(etapa)].update(
                    {
                        "columns_header_row": columns_header_row,
                        "data_start_row": data_start_row,
                        "data_end_row": data_end_row,
                    }
                )
                for _, row_data in subset.iterrows():
                    row_vals = [row_data[c] for c in valid_output_cols]
                    final_sheet_data.append(row_vals)
                    current_excel_row += 1

                final_sheet_data.append(
                    [None] * num_cols
                )  # Espaço após tabela da etapa
                current_excel_row += 1

        # Remove empty rows at the end caused by our logic (if any)
        while final_sheet_data and all(x is None for x in final_sheet_data[-1]):
            final_sheet_data.pop()

        # 6. Criar Workbook e Worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tabela de Preços"

        # Preenche a planilha
        for r_idx, row_data in enumerate(final_sheet_data, 1):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                # Estilo básico: centralizado
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Estilização ---
        # Estilos e helpers (definir ANTES de usar)
        header_etapa_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        header_bloco_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )
        col_header_fill = PatternFill(
            start_color="BFBFBF", end_color="BFBFBF", fill_type="solid"
        )
        col_header_font = Font(color="000000", bold=True)

        medium_side = Side(style="medium", color="000000")

        def apply_medium_border_merged_row(row_idx: int):
            """Aplica borda média contínua na largura toda (linha mesclada)."""
            for c_idx in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=c_idx).border = Border(
                    left=medium_side if c_idx == 1 else None,
                    right=medium_side if c_idx == num_cols else None,
                    top=medium_side,
                    bottom=medium_side,
                )

        def apply_medium_outline_table(
            header_row_idx: int, data_start_row_idx: int, data_end_row_idx: int | None
        ):
            """Aplica apenas bordas externas (médias) do header até a última linha de dados."""
            if not data_end_row_idx or data_end_row_idx < data_start_row_idx:
                return

            # Topo do bloco: linha de cabeçalho
            for c_idx in range(1, num_cols + 1):
                ws.cell(row=header_row_idx, column=c_idx).border = Border(
                    left=medium_side if c_idx == 1 else None,
                    right=medium_side if c_idx == num_cols else None,
                    top=medium_side,
                    bottom=None,
                )

            # Laterais (todas as linhas de dados) + base (somente última linha)
            for r_idx in range(data_start_row_idx, data_end_row_idx + 1):
                is_last = r_idx == data_end_row_idx
                for c_idx in range(1, num_cols + 1):
                    ws.cell(row=r_idx, column=c_idx).border = Border(
                        left=medium_side if c_idx == 1 else None,
                        right=medium_side if c_idx == num_cols else None,
                        top=None,
                        bottom=medium_side if is_last else None,
                    )

        # Título Principal
        if len(final_sheet_data) >= 3:
            # Merge título
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
            title_cell = ws.cell(row=3, column=1)
            title_cell.font = Font(size=12, bold=True, color="000000")
            title_cell.fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )
            apply_medium_border_merged_row(3)

        for etapa_nome, info in row_map["etapas"].items():
            h_row = info["header_row"]
            # Merge Header Etapa
            ws.merge_cells(
                start_row=h_row, start_column=1, end_row=h_row, end_column=num_cols
            )
            c = ws.cell(row=h_row, column=1)
            c.font = Font(bold=True, size=12)
            c.fill = header_etapa_fill
            apply_medium_border_merged_row(h_row)

            # Processa os blocos dentro da etapa
            if "blocks" in info and info["blocks"]:
                # Tem blocos: estiliza cada bloco
                for bloco_nome, bloco_info in info["blocks"].items():
                    bloco_h_row = bloco_info["header_row"]
                    # Merge Header Bloco
                    ws.merge_cells(
                        start_row=bloco_h_row,
                        start_column=1,
                        end_row=bloco_h_row,
                        end_column=num_cols,
                    )
                    bloco_cell = ws.cell(row=bloco_h_row, column=1)
                    bloco_cell.font = Font(bold=True, size=10)
                    bloco_cell.fill = header_bloco_fill
                    apply_medium_border_merged_row(bloco_h_row)

                    # Header das Colunas do Bloco (Bloco Header -> Vazia -> Colunas Header)
                    col_h_row = bloco_h_row + 2
                    if col_h_row <= ws.max_row:
                        for c_idx in range(1, num_cols + 1):
                            cell = ws.cell(row=col_h_row, column=c_idx)
                            cell.font = col_header_font
                            cell.fill = col_header_fill

                    apply_medium_outline_table(
                        header_row_idx=bloco_info.get("columns_header_row", col_h_row),
                        data_start_row_idx=bloco_info.get(
                            "data_start_row", col_h_row + 1
                        ),
                        data_end_row_idx=bloco_info.get("data_end_row"),
                    )
            else:
                # Sem blocos: estiliza diretamente o cabeçalho de colunas da etapa
                col_h_row = h_row + 2
                if col_h_row <= ws.max_row:
                    for c_idx in range(1, num_cols + 1):
                        cell = ws.cell(row=col_h_row, column=c_idx)
                        cell.font = col_header_font
                        cell.fill = col_header_fill

                apply_medium_outline_table(
                    header_row_idx=info.get("columns_header_row", col_h_row),
                    data_start_row_idx=info.get("data_start_row", col_h_row + 1),
                    data_end_row_idx=info.get("data_end_row"),
                )

        # Ajuste de Largura de Coluna (Auto-size simples)
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = min(adjusted_width, 50)  # Max 50 chars

        # 7. Salvar em BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        print("(Tabela Preços Formatador) Processamento concluído com sucesso.")
        return output

    except ValueError as ve:
        print(f"(Tabela Preços Formatador) ERRO VALIDAÇÃO: {ve}")
        traceback.print_exc()
        raise ve
    except Exception as e:
        print(f"(Tabela Preços Formatador) ERRO INESPERADO: {e}")
        traceback.print_exc()
        raise RuntimeError(f"Erro inesperado no processamento: {e}") from e
