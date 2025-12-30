# formatadores/incorporacao_formatador.py

import pandas as pd
import io
import re
import traceback
import openpyxl
import unicodedata
import numpy as np
from openpyxl.utils import get_column_letter


def normalize_text_for_match(text):
    """Normaliza texto para busca: minúsculo, sem acentos, sem não-alfanuméricos."""
    if not isinstance(text, str):
        text = str(text)
    try:
        text = (
            unicodedata.normalize("NFKD", text)
            .encode("ASCII", "ignore")
            .decode("ASCII")
        )
        text = text.lower()
        text = re.sub(r"[^a-z0-9]", "", text)
        return text.strip()
    except Exception:
        return str(text).lower().strip().replace(" ", "")


def find_column_flexible(df_columns, concept_keywords, concept_name, required=True):
    """Encontra coluna de forma flexível."""
    normalized_input_cols = {normalize_text_for_match(col): col for col in df_columns}
    print(
        f"  Buscando '{concept_name}': Keywords={concept_keywords}. Colunas Norm.: {list(normalized_input_cols.keys())}"
    )
    found_col_name = None
    for keyword in concept_keywords:
        norm_keyword = normalize_text_for_match(keyword)
        if norm_keyword in normalized_input_cols:
            found_col_name = normalized_input_cols[norm_keyword]
            print(
                f"    -> Match exato norm. '{norm_keyword}' para '{concept_name}'. Col original: '{found_col_name}'"
            )
            return found_col_name
    potential_matches = []
    for keyword in concept_keywords:
        norm_keyword = normalize_text_for_match(keyword)
        if not norm_keyword:
            continue
        for norm_col, orig_col in normalized_input_cols.items():
            if not norm_col:
                continue
            if norm_keyword in norm_col:
                priority = 0 if norm_col.startswith(norm_keyword) else 1
                potential_matches.append((priority, orig_col))
    if potential_matches:
        potential_matches.sort()
        found_col_name = potential_matches[0][1]
        print(
            f"    -> Melhor match parcial para '{concept_name}'. Col original: '{found_col_name}'"
        )
        return found_col_name
    if required:
        raise ValueError(
            f"Coluna obrigatória '{concept_name}' não encontrada. Keywords: {concept_keywords}. Colunas: {list(df_columns)}"
        )
    else:
        print(f"    -> Coluna opcional '{concept_name}' não encontrada.")
        return None


def format_decimal_br(value, precision):
    if pd.isna(value) or str(value).strip() == "":
        return ""
    s_val = str(value).strip()
    s_val = re.sub(r"[^\d,.-]", "", s_val)
    if "," in s_val and "." in s_val:
        s_val = s_val.replace(".", "").replace(",", ".")
    elif "," in s_val:
        s_val = s_val.replace(",", ".")
    try:
        num = float(s_val)
        format_string = f"{{:.{precision}f}}"
        formatted = format_string.format(num).replace(".", ",")
        parts = formatted.split(",")
        int_part = parts[0]
        dec_part = parts[1] if len(parts) > 1 else ""
        int_part_fmt = ""
        n_dig = len(int_part)
        for i, d in enumerate(int_part):
            int_part_fmt += d
            if (n_dig - 1 - i) > 0 and (n_dig - 1 - i) % 3 == 0:
                int_part_fmt += "."
        return f"{int_part_fmt},{dec_part}"
    except (ValueError, TypeError):
        return str(value)


def parse_flexible_float(value):
    if value is None:
        return None
    s_val = str(value).strip()
    if not s_val:
        return None
    s_val = re.sub(r"[^\d,.-]+", "", s_val)
    num_commas = s_val.count(",")
    num_dots = s_val.count(".")
    if num_commas == 1 and num_dots >= 1:
        if s_val.rfind(",") > s_val.rfind("."):
            s_val = s_val.replace(".", "").replace(",", ".")
        else:
            s_val = s_val.replace(",", "")
    elif num_commas >= 2 and num_dots == 1:
        s_val = s_val.replace(",", "")
    elif num_dots >= 2 and num_commas == 1:
        s_val = s_val.replace(".", "").replace(",", ".")
    elif num_commas == 1 and num_dots == 0:
        s_val = s_val.replace(",", ".")
    elif num_dots >= 2 and num_commas == 0:
        s_val = s_val.replace(".", "")
    elif num_commas >= 2 and num_dots == 0:
        s_val = s_val.replace(",", "")
    try:
        return float(s_val)
    except (ValueError, TypeError):
        return None


def format_tipo_with_leading_zero(tipo_str):
    original_tipo_str = str(tipo_str).strip()
    if not original_tipo_str:
        return ""
    parts = original_tipo_str.split()
    if len(parts) >= 2:
        prefix = " ".join(parts[:-1])
        number_part = parts[-1]
        try:
            number_int = int(number_part)
            formatted_number = f"{number_int:02d}"
            return f"{prefix} {formatted_number}"
        except ValueError:
            return original_tipo_str
    else:
        return original_tipo_str


def extract_and_format_number(value_str, default_if_error="XX"):
    """Extrai o primeiro número de uma string e o formata com 2 dígitos (zero à esquerda)."""
    match = re.search(r"\d+", str(value_str))
    if match:
        try:
            return f"{int(match.group(0)):02d}"
        except (ValueError, TypeError):
            return default_if_error
    return default_if_error


def verificar_vaga(g, num_mode):
    if pd.isna(g) or str(g).strip() == "":
        return "01 VAGA"

    if num_mode:
        s = str(g).strip()
        n = 0
        if s:
            # Check for multiple vagas separated by "e" or ","
            for sep in [" e ", ","]:
                if sep in s:
                    n = len([v for v in s.split(sep) if v.strip()])
                    break
            else:
                # If no separator, try to extract the first number from the string
                match = re.search(r"(\d+)", s)
                if match:
                    try:
                        n = int(match.group(1))
                    except ValueError:
                        n = 1  # Fallback if parsing fails
                else:
                    n = 1  # Fallback if no digits found at all

        if n >= 4:
            return "04 VAGAS"
        elif n == 3:
            return "03 VAGAS"
        elif n == 2:
            return "02 VAGAS"
        else:  # Handles n=1, n=0, and any other case
            return "01 VAGA"
    else:  # "Smart" mode
        try:
            # Try to parse as a number. This handles "2", "2.0", "15.5", "15,5"
            # It will fail for "2 VAGAS" or "1 e 2"
            gn = float(str(g).replace(",", ".").strip())

            # Check if it's likely an area (metragem) because it has decimals
            if abs(gn - int(gn)) > 0.001:
                if gn > 35:
                    return "04 VAGAS"
                elif gn > 25:
                    return "03 VAGAS"
                elif gn > 15:
                    return "02 VAGAS"
                else:
                    return "01 VAGA"
            else:
                gi = int(gn)
                if gi >= 4:
                    return "04 VAGAS"
                elif gi == 3:
                    return "03 VAGAS"
                elif gi == 2:
                    return "02 VAGAS"
                else:
                    return "01 VAGA"
        except (ValueError, TypeError):
            # If it's not a number (e.g., "02 VAGAS", "01 e 02"), use the text-based logic (num_mode=True)
            return verificar_vaga(g, True)


def processar_incorporacao_web(input_filepath_or_stream):
    """
    Processa planilha de incorporação.
    - Lógica Padrão: Processa arquivos com títulos de seção 'BLOCO'/'QUADRA'.
    - Lógica Composta: Se encontrar colunas 'BLOCO', 'QUADRA', e 'CASA',
      cria um identificador de unidade. Se 'BLOCO' for 'US', usa o número da casa.
    """
    print(f"(Formatador Incorporação - v12.5 Final US-Rule Corrigido) Iniciando.")
    output = io.BytesIO()
    try:
        print(f"Lendo arquivo/stream...")
        df_raw = pd.read_excel(input_filepath_or_stream, header=None, dtype=str).fillna(
            ""
        )
        if df_raw.empty:
            raise ValueError("Arquivo Excel está vazio.")

        header_row_index = -1
        possible_headers_data = [
            "tipo",
            "casa",
            "apt",
            "apto",
            "bloco",
            "quadra",
            "areaconstruida",
            "area",
            "fracaoideal",
            "valor",
        ]
        for idx, row in df_raw.head(15).iterrows():
            row_values_norm = [
                normalize_text_for_match(str(v)) for v in row.values if pd.notna(v)
            ]
            if sum(h in row_values_norm for h in possible_headers_data) >= 3:
                header_row_index = idx
                print(
                    f"Linha do cabeçalho dos dados encontrada no índice: {header_row_index}"
                )
                break
        if header_row_index == -1:
            raise ValueError(
                "Não foi possível encontrar a linha do cabeçalho dos dados."
            )

        df_columns = df_raw.iloc[header_row_index].fillna("").astype(str).str.strip()
        print(f"Valores do cabeçalho detectado: {df_columns.tolist()}")

        col_map = {}
        concepts_to_find = {
            "TIPO": (["tipo", "tipologia"], True),
            "CASA": (["casa"], False),
            "APT_UNID": (["apt", "apto", "apartamento", "unidade"], False),
            "BLOCO": (["bloco", "blk"], False),
            "QUADRA": (["quadra", "qd"], False),
            "ÁREA CONSTRUIDA": (["areaconstruida", "área construída"], False),
            "QUINTAL": (["quintal"], False),
            "GARAGEM": (["garagem", "garagem e frontal"], False),
            "VAGAS_QTD": (["vagas de garagem", "vagas"], False),
            "ÁREA PRIVATIVA": (["areaprivativa", "área privativa"], False),
            "FRAÇÃO IDEAL": (["fracaoideal", "fração ideal"], False),
        }

        print("--- Mapeando índices numéricos das colunas ---")
        for concept, (keywords, required) in concepts_to_find.items():
            col_name = find_column_flexible(
                df_columns, keywords, concept, required=False
            )
            if col_name:
                col_index = df_columns[df_columns == col_name].index[0]
                col_map[concept] = col_index
                print(
                    f"    -> Mapeado '{concept}' para o índice de coluna: {col_index} (Nome: '{col_name}')"
                )
            else:
                col_map[concept] = None

        has_bloco = col_map.get("BLOCO") is not None
        has_quadra = col_map.get("QUADRA") is not None
        has_casa = col_map.get("CASA") is not None
        has_apt = col_map.get("APT_UNID") is not None

        composite_mode = None
        if has_bloco and has_quadra and has_casa:
            composite_mode = "FULL"  # Modelo Praia (US/BL-QD-CS)
        elif (has_quadra or has_bloco) and has_casa:
            composite_mode = "QD_CS"  # Modelo QDXX - CSXX
        elif (has_bloco or has_quadra) and has_apt:
            composite_mode = "BL_APT"  # Modelo BLXX - APTXX

        is_composite_unit_format = composite_mode is not None

        print(f"--- DETECÇÃO DE FORMATO ---")
        if is_composite_unit_format:
            print(f"Formato detectado: Unidade Composta ({composite_mode}).")
            if col_map.get("TIPO") is None:
                raise ValueError("Formato de Unidade Composta requer a coluna 'TIPO'.")
        else:
            print(
                "Formato detectado: Padrão (cabeçalho de seção ou colunas individuais)."
            )
            if col_map.get("TIPO") is None or (
                col_map.get("CASA") is None and col_map.get("APT_UNID") is None
            ):
                raise ValueError(
                    "Colunas obrigatórias para o formato padrão ('TIPO' e 'CASA' ou 'APT') não encontradas."
                )
        print("----------------------------")

        processed_data = []
        ultimo_bloco_num_str = None
        header_saida_bloco_quadra = "BLOCO"
        final_header_casa_apt = "UNIDADE"

        print(
            f"Iterando pelas linhas de dados a partir do índice {header_row_index + 1}..."
        )
        for idx in range(header_row_index + 1, len(df_raw)):
            row = df_raw.iloc[idx].fillna("")

            if is_composite_unit_format:
                unit_id_final = ""

                if composite_mode == "FULL":
                    casa_val_str = str(row[col_map["CASA"]])
                    if pd.isna(casa_val_str) or casa_val_str.strip() == "":
                        continue
                    bloco_val = str(row[col_map["BLOCO"]])
                    quadra_val = str(row[col_map["QUADRA"]])
                    quadra_num_fmt = extract_and_format_number(quadra_val)
                    casa_num_fmt = extract_and_format_number(casa_val_str)
                    if "us" in normalize_text_for_match(bloco_val):
                        bloco_id_part = f"US{casa_num_fmt}"
                    else:
                        bloco_num_fmt = extract_and_format_number(bloco_val)
                        bloco_id_part = f"BL{bloco_num_fmt}"
                    unit_id_final = (
                        f"{bloco_id_part}-QD{quadra_num_fmt}-CS{casa_num_fmt}"
                    )

                elif composite_mode == "QD_CS":
                    q_idx = col_map["QUADRA"] if has_quadra else col_map["BLOCO"]
                    c_idx = col_map["CASA"]
                    q_val = str(row[q_idx])
                    c_val = str(row[c_idx])
                    if not q_val.strip() or not c_val.strip():
                        continue
                    unit_id_final = f"QD{extract_and_format_number(q_val)} - CASA {extract_and_format_number(c_val)}"

                elif composite_mode == "BL_APT":
                    b_idx = col_map["BLOCO"] if has_bloco else col_map["QUADRA"]
                    a_idx = col_map["APT_UNID"]
                    b_val = str(row[b_idx])
                    a_val = str(row[a_idx])
                    if not b_val.strip() or not a_val.strip():
                        continue
                    unit_id_final = f"BL{extract_and_format_number(b_val)} - APT {extract_and_format_number(a_val)}"

                if not unit_id_final:
                    continue

                tipo_val = format_tipo_with_leading_zero(row[col_map["TIPO"]])
                area_const_val = (
                    row[col_map["ÁREA CONSTRUIDA"]]
                    if col_map.get("ÁREA CONSTRUIDA") is not None
                    else ""
                )
                quintal_val = (
                    row[col_map["QUINTAL"]]
                    if col_map.get("QUINTAL") is not None
                    else ""
                )
                garagem_val = (
                    row[col_map["GARAGEM"]]
                    if col_map.get("GARAGEM") is not None
                    else ""
                )
                area_priv_val = (
                    row[col_map["ÁREA PRIVATIVA"]]
                    if col_map.get("ÁREA PRIVATIVA") is not None
                    else ""
                )
                fracao_val = (
                    row[col_map["FRAÇÃO IDEAL"]]
                    if col_map.get("FRAÇÃO IDEAL") is not None
                    else ""
                )

                # LÓGICA DE VAGAS DE GARAGEM
                vagas_final_str = "01 VAGA"  # Padrão
                vagas_qtd_idx = col_map.get("VAGAS_QTD")
                garagem_area_idx = col_map.get("GARAGEM")

                vagas_source_val = None
                if (
                    vagas_qtd_idx is not None
                    and str(row.get(vagas_qtd_idx, "")).strip()
                ):
                    vagas_source_val = row[vagas_qtd_idx]
                elif (
                    garagem_area_idx is not None
                    and str(row.get(garagem_area_idx, "")).strip()
                ):
                    vagas_source_val = row[garagem_area_idx]

                if vagas_source_val is not None:
                    vagas_final_str = verificar_vaga(vagas_source_val, num_mode=False)

                processed_data.append(
                    {
                        "UNIDADE": unit_id_final,
                        "TIPO": tipo_val,
                        "ÁREA CONSTRUIDA": area_const_val,
                        "QUINTAL": quintal_val,
                        "GARAGEM": garagem_val,
                        "VAGAS DE GARAGEM": vagas_final_str,
                        "ÁREA PRIVATIVA": area_priv_val,
                        "FRAÇÃO IDEAL": fracao_val,
                        "ETAPA": "01",
                    }
                )
            else:
                # Lógica de Fallback (sem alterações)
                cell_val_a = str(row.iloc[0]).strip()
                is_quadra_title = cell_val_a.lower().startswith("quadra")
                is_bloco_title = cell_val_a.lower().startswith("bloco")

                if is_quadra_title or is_bloco_title:
                    header_saida_bloco_quadra = "QUADRA" if is_quadra_title else "BLOCO"
                    ultimo_bloco_num_str = extract_and_format_number(cell_val_a, "??")
                    print(
                        f"  Linha {idx}: Título de Seção '{cell_val_a}'. Número = {ultimo_bloco_num_str}."
                    )
                    continue

                casa_apt_col_idx = (
                    col_map.get("CASA")
                    if col_map.get("CASA") is not None
                    else col_map.get("APT_UNID")
                )
                if (
                    casa_apt_col_idx is None
                    or pd.isna(row[casa_apt_col_idx])
                    or str(row[casa_apt_col_idx]).strip() == ""
                ):
                    continue

                bloco_quadra_col_idx = (
                    col_map.get("BLOCO")
                    if col_map.get("BLOCO") is not None
                    else col_map.get("QUADRA")
                )
                if (
                    bloco_quadra_col_idx is not None
                    and str(row[bloco_quadra_col_idx]).strip()
                ):
                    ultimo_bloco_num_str = extract_and_format_number(
                        row[bloco_quadra_col_idx], "??"
                    )

                if ultimo_bloco_num_str is None:
                    continue

                tipo_val_original = str(row[col_map["TIPO"]])
                casa_apt_val_original = str(row[casa_apt_col_idx]).strip()
                formatted_tipo_output_val = format_tipo_with_leading_zero(
                    tipo_val_original
                )

                is_special_unit = (
                    "pcd" in normalize_text_for_match(tipo_val_original)
                    or "pne" in normalize_text_for_match(tipo_val_original)
                    or "pcd" in normalize_text_for_match(casa_apt_val_original)
                    or "pne" in normalize_text_for_match(casa_apt_val_original)
                )

                unit_number_part = extract_and_format_number(
                    casa_apt_val_original, casa_apt_val_original
                )
                formatted_unit_number_with_pcd = (
                    f"{unit_number_part} (PCD)" if is_special_unit else unit_number_part
                )

                header_casa_apt_orig_val = df_columns[casa_apt_col_idx]
                norm_header = normalize_text_for_match(str(header_casa_apt_orig_val))
                if "casa" in norm_header:
                    final_header_casa_apt = "CASA"
                elif "apt" in norm_header or "apartamento" in norm_header:
                    final_header_casa_apt = "APT"
                else:
                    final_header_casa_apt = "UNIDADE"

                area_const_val = (
                    row[col_map["ÁREA CONSTRUIDA"]]
                    if col_map.get("ÁREA CONSTRUIDA") is not None
                    else ""
                )
                quintal_val = (
                    row[col_map["QUINTAL"]]
                    if col_map.get("QUINTAL") is not None
                    else ""
                )
                garagem_val = (
                    row[col_map["GARAGEM"]]
                    if col_map.get("GARAGEM") is not None
                    else ""
                )
                area_priv_val = (
                    row[col_map["ÁREA PRIVATIVA"]]
                    if col_map.get("ÁREA PRIVATIVA") is not None
                    else ""
                )
                fracao_val = (
                    row[col_map["FRAÇÃO IDEAL"]]
                    if col_map.get("FRAÇÃO IDEAL") is not None
                    else ""
                )

                # LÓGICA DE VAGAS DE GARAGEM
                vagas_final_str = "01 VAGA"  # Padrão
                vagas_qtd_idx = col_map.get("VAGAS_QTD")
                garagem_area_idx = col_map.get("GARAGEM")

                vagas_source_val = None
                if (
                    vagas_qtd_idx is not None
                    and str(row.get(vagas_qtd_idx, "")).strip()
                ):
                    vagas_source_val = row[vagas_qtd_idx]
                elif (
                    garagem_area_idx is not None
                    and str(row.get(garagem_area_idx, "")).strip()
                ):
                    vagas_source_val = row[garagem_area_idx]

                if vagas_source_val is not None:
                    vagas_final_str = verificar_vaga(vagas_source_val, num_mode=False)

                processed_data.append(
                    {
                        header_saida_bloco_quadra: ultimo_bloco_num_str,
                        "TIPO": formatted_tipo_output_val,
                        final_header_casa_apt: formatted_unit_number_with_pcd,
                        "ÁREA CONSTRUIDA": area_const_val,
                        "QUINTAL": quintal_val,
                        "GARAGEM": garagem_val,
                        "VAGAS DE GARAGEM": vagas_final_str,
                        "ÁREA PRIVATIVA": area_priv_val,
                        "FRAÇÃO IDEAL": fracao_val,
                        "ETAPA": "01",
                    }
                )

        # O restante do código (DataFrame, formatação, Excel) não precisa de alterações
        print(f"Iteração concluída. {len(processed_data)} linhas de dados extraídas.")
        if not processed_data:
            raise ValueError("Nenhum dado válido extraído.")
        df_final = pd.DataFrame(processed_data)
        print("--- Formatando Colunas Numéricas ---")
        cols_to_format_final = {
            "ÁREA CONSTRUIDA": 2,
            "QUINTAL": 2,
            "GARAGEM": 2,
            "ÁREA PRIVATIVA": 2,
            "FRAÇÃO IDEAL": 6,
        }
        for col_name, precision in cols_to_format_final.items():
            if col_name in df_final.columns:
                df_final[col_name] = df_final[col_name].apply(
                    lambda x: format_decimal_br(x, precision)
                )
        if is_composite_unit_format:
            ordem_saida = [
                "UNIDADE",
                "TIPO",
                "ÁREA CONSTRUIDA",
                "QUINTAL",
                "GARAGEM",
                "VAGAS DE GARAGEM",
                "ÁREA PRIVATIVA",
                "FRAÇÃO IDEAL",
                "ETAPA",
            ]
        else:
            ordem_saida = [
                header_saida_bloco_quadra,
                "TIPO",
                final_header_casa_apt,
                "ÁREA CONSTRUIDA",
                "QUINTAL",
                "GARAGEM",
                "VAGAS DE GARAGEM",
                "ÁREA PRIVATIVA",
                "FRAÇÃO IDEAL",
                "ETAPA",
            ]
        colunas_finais_real = [col for col in ordem_saida if col in df_final.columns]
        for col in df_final.columns:
            if col not in colunas_finais_real:
                colunas_finais_real.append(col)
        df_final = df_final[colunas_finais_real]
        print(f"Ordem final das colunas: {df_final.columns.tolist()}")
        print("Gerando arquivo Excel e aplicando conversão/formatação numérica...")
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_final.to_excel(
                writer, index=False, header=True, sheet_name="Incorporacao Formatada"
            )
            workbook = writer.book
            worksheet = writer.sheets["Incorporacao Formatada"]
            col_excel_formats = {
                "ÁREA CONSTRUIDA": "0.00",
                "QUINTAL": "0.00",
                "GARAGEM": "0.00",
                "ÁREA PRIVATIVA": "0.00",
                "FRAÇÃO IDEAL": "0.000000000",
            }
            text_columns = [
                "UNIDADE",
                "TIPO",
                "BLOCO",
                "QUADRA",
                "CASA",
                "APT",
                "VAGAS DE GARAGEM",
            ]
            for row_idx in range(2, worksheet.max_row + 1):
                for col_idx_1based in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx_1based)
                    current_col_name = df_final.columns[col_idx_1based - 1]
                    if current_col_name in text_columns:
                        if isinstance(cell.value, str) and cell.value.startswith(
                            ("=", "+", "-", "@")
                        ):
                            cell.value = "'" + cell.value
                            cell.number_format = "@"
                        continue
                    numeric_value = parse_flexible_float(cell.value)
                    if numeric_value is not None:
                        cell.value = numeric_value
                        if current_col_name in col_excel_formats:
                            cell.number_format = col_excel_formats[current_col_name]
                    elif cell.value is None or str(cell.value).strip() == "":
                        cell.value = None
            for i, column_name in enumerate(df_final.columns):
                column_letter = get_column_letter(i + 1)
                try:
                    max_len_data = 0
                    if not df_final[column_name].empty:
                        max_len_data = df_final[column_name].astype(str).map(len).max()
                    max_len_header = len(str(column_name))
                    width = max(max_len_data, max_len_header) + 3
                    worksheet.column_dimensions[column_letter].width = min(width, 60)
                except Exception as e_width:
                    print(
                        f"Aviso: Falha ao ajustar largura da coluna {column_letter}: {e_width}"
                    )
                    worksheet.column_dimensions[column_letter].width = 15
        output.seek(0)
        print("(Formatador Incorporação - v12.5) Arquivo Excel processado gerado.")
        return output

    except ValueError as ve:
        print(f"(Formatador Incorporação) ERRO VALIDAÇÃO: {ve}")
        traceback.print_exc()
        raise ve
    except Exception as e:
        print(f"(Formatador Incorporação) ERRO INESPERADO: {e}")
        traceback.print_exc()
        raise RuntimeError(f"Erro inesperado: {e}") from e
